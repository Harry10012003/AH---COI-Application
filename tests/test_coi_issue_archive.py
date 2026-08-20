from __future__ import annotations

from pathlib import Path
import tempfile
import unittest

from openpyxl import load_workbook

import backend.engine.coi_issue_archive as archive


def _payload(go: str, rows: list[dict]) -> dict:
    columns = [
        {"key": "BRAND", "label": "BRAND"},
        {"key": "GO#", "label": "GO#"},
        {"key": "PPO", "label": "PPO"},
        {"key": "Type", "label": "Type"},
        {"key": "COLOR_CODE", "label": "COLOR_CODE"},
        {"key": "JOB ORDER NO", "label": "JOB ORDER NO"},
        {"key": "Qty (pcs)", "label": "Qty"},
        {"key": "Allocate Q'ty (Yds)", "label": "Allocate Q'ty (Lot)"},
    ]
    return {"go": go, "columns": columns, "rows": rows}


class IssuedCoiArchiveTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)
        self.previous_db = archive.ISSUED_COI_ARCHIVE_DB
        self.previous_issue_dir = archive._ISSUE_DIR
        self.previous_filename = archive.ISSUED_COI_COMBINED_FILENAME
        archive.ISSUED_COI_ARCHIVE_DB = self.root / "archive.db"
        archive._ISSUE_DIR = self.root / "issued"
        archive.ISSUED_COI_COMBINED_FILENAME = "COI-CUTTING-COMBINED.xlsx"

    def tearDown(self) -> None:
        archive.ISSUED_COI_ARCHIVE_DB = self.previous_db
        archive._ISSUE_DIR = self.previous_issue_dir
        archive.ISSUED_COI_COMBINED_FILENAME = self.previous_filename
        self.tempdir.cleanup()

    def _publish(self, payload: dict, issued_at: str, filename: str) -> int:
        saved = archive.archive_issue_snapshot(
            payload,
            issued_at=issued_at,
            filename=filename,
            storage_state="LOCAL_EXPORTED",
        )
        self.assertTrue(saved["ok"])
        published = archive.mark_issue_batch_published(saved["batch_id"], {"file_path": filename})
        self.assertTrue(published["ok"])
        return int(saved["batch_id"])

    def test_combined_workbook_uses_latest_published_issue_per_go(self) -> None:
        self._publish(
            _payload(
                "S26V00001",
                [{"BRAND": "Brand A", "GO#": "S26V00001", "PPO": "PPO-1", "Type": "B", "COLOR_CODE": "RED", "JOB ORDER NO": "26V00001VN01", "Qty (pcs)": 100, "Allocate Q'ty (Yds)": 120}],
            ),
            "2026-07-01 08:00:00",
            "S26V00001-COI-20260701.xlsx",
        )
        latest_version = self._publish(
            _payload(
                "S26V00001",
                [
                    {"BRAND": "Brand A", "GO#": "S26V00001", "PPO": "PPO-1", "Type": "B", "COLOR_CODE": "RED", "JOB ORDER NO": "26V00001VN01", "Qty (pcs)": 100, "Allocate Q'ty (Yds)": 130},
                    {"BRAND": "Brand A", "GO#": "S26V00001", "PPO": "PPO-1", "Type": "B", "COLOR_CODE": "BLUE", "JOB ORDER NO": "26V00001VN01", "Qty (pcs)": 200, "Allocate Q'ty (Yds)": 220},
                ],
            ),
            "2026-07-02 08:00:00",
            "S26V00001-COI-20260702.xlsx",
        )
        self._publish(
            _payload(
                "S26V00002",
                [{"BRAND": "Brand B", "GO#": "S26V00002", "PPO": "PPO-2", "Type": "F", "COLOR_CODE": "WHITE", "JOB ORDER NO": "26V00002VN01", "Qty (pcs)": 50, "Allocate Q'ty (Yds)": 60}],
            ),
            "2026-07-03 08:00:00",
            "S26V00002-COI-20260703.xlsx",
        )

        result = archive.rebuild_combined_issued_coi(sync_to_onedrive=False)

        self.assertTrue(result["ok"])
        self.assertEqual(result["go_count"], 2)
        self.assertEqual(result["row_count"], 3)
        workbook = load_workbook(result["file_path"], read_only=True, data_only=True)
        combined = workbook["COI Combined"]
        summary = workbook["GO Summary"]
        self.assertEqual(combined.max_row - 4, 3)
        self.assertEqual(summary.max_row - 1, 2)
        headers = [combined.cell(4, index).value for index in range(1, combined.max_column + 1)]
        version_column = headers.index("ISSUE VERSION") + 1
        go_column = headers.index("GO#") + 1
        first_go_versions = [
            combined.cell(row, version_column).value
            for row in range(5, combined.max_row + 1)
            if combined.cell(row, go_column).value == "S26V00001"
        ]
        self.assertEqual(first_go_versions, [latest_version, latest_version])
        workbook.close()

    def test_latest_feed_uses_one_published_issue_per_go_and_supports_filters(self) -> None:
        self._publish(
            _payload(
                "S26V00001",
                [{"GO#": "S26V00001", "PPO": "PPO-OLD", "Type": "B", "COLOR_CODE": "RED", "JOB ORDER NO": "26V00001VN01", "Qty (pcs)": 10}],
            ),
            "2026-07-01 08:00:00",
            "S26V00001-COI-20260701.xlsx",
        )
        latest_version = self._publish(
            _payload(
                "S26V00001",
                [{"GO#": "S26V00001", "PPO": "PPO-NEW", "Type": "B", "COLOR_CODE": "BLUE", "JOB ORDER NO": "26V00001VN01", "Qty (pcs)": 20}],
            ),
            "2026-07-02 08:00:00",
            "S26V00001-COI-20260702.xlsx",
        )
        self._publish(
            _payload(
                "S26V00002",
                [{"GO#": "S26V00002", "PPO": "PPO-OTHER", "Type": "F", "COLOR_CODE": "WHITE", "JOB ORDER NO": "26V00002VN01", "Qty (pcs)": 30}],
            ),
            "2026-07-03 08:00:00",
            "S26V00002-COI-20260703.xlsx",
        )

        feed = archive.get_latest_issued_coi_feed(limit=100)

        self.assertTrue(feed["ok"])
        self.assertEqual(feed["mode"], "latest_published_issue_per_go")
        self.assertEqual(feed["pagination"]["total_rows"], 2)
        self.assertEqual(feed["go_count"], 2)
        self.assertEqual({row["PPO"] for row in feed["rows"]}, {"PPO-NEW", "PPO-OTHER"})
        self.assertEqual(
            [row["ISSUE_VERSION"] for row in feed["rows"] if row["GO#"] == "S26V00001"],
            [latest_version],
        )
        self.assertIn("ISSUE_AT", {column["key"] for column in feed["columns"]})

        filtered = archive.get_latest_issued_coi_feed(ppo="ppo-new", limit=100)
        self.assertEqual(filtered["pagination"]["total_rows"], 1)
        self.assertEqual(filtered["rows"][0]["GO#"], "S26V00001")

    def test_published_issue_can_be_synchronized_after_ppo_override(self) -> None:
        batch_id = self._publish(
            _payload(
                "S26V00001",
                [{"GO#": "S26V00001", "PPO": "PPO-OLD", "Type": "B", "COLOR_CODE": "RED", "JOB ORDER NO": "26V00001VN01", "Qty (pcs)": 10}],
            ),
            "2026-07-01 08:00:00",
            "S26V00001-COI-20260701.xlsx",
        )

        synced = archive.replace_latest_published_issue_snapshot(
            "S26V00001",
            _payload(
                "S26V00001",
                [{"GO#": "S26V00001", "PPO": "PPO-NEW", "Type": "B", "COLOR_CODE": "RED", "JOB ORDER NO": "26V00001VN01", "Qty (pcs)": 10}],
            ),
        )

        self.assertTrue(synced["ok"])
        self.assertEqual(synced["batch_id"], batch_id)
        self.assertEqual(synced["sync_revision"], 1)
        self.assertTrue(synced["last_synced_at"])
        feed = archive.get_latest_issued_coi_feed(go="S26V00001", limit=100)
        self.assertEqual(feed["rows"][0]["PPO"], "PPO-NEW")
        self.assertEqual(feed["rows"][0]["ISSUE_VERSION"], batch_id)
        self.assertEqual(feed["rows"][0]["ISSUE_AT"], "2026-07-01 08:00:00")
        self.assertTrue(feed["rows"][0]["ISSUE_SYNC_AT"])
        self.assertEqual(feed["data_version"], f"{batch_id}:1")


if __name__ == "__main__":
    unittest.main()
