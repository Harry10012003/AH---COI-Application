import tempfile
import unittest
from pathlib import Path
import tempfile
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from backend.precoi.clients import MESSummaryClient
from backend.precoi.exceptions import MESRequestError, ValidationError, YPDRequestError
from backend.precoi.excel_exporter import (
    AGGREGATE_KEY_COLUMN,
    ALL_HEADERS,
    COLLAR_ALL_HEADERS,
    COLLAR_SHEET_NAME,
    COLLAR_VISIBLE_HEADERS,
    CM_ALL_HEADERS,
    CM_VISIBLE_HEADERS,
    LEGACY_VISIBLE_HEADERS,
    MASTER_FILE_NAME,
    VISIBLE_HEADERS,
    read_records_from_workbook,
    save_master_workbook,
    write_workbook,
)
from backend.precoi.models import (
    CmQaAggregateRow,
    ExportRecord,
    GOBomBlock,
    GOBomRow,
    GOReportData,
    GoColorSummaryRow,
    GoLotColorRow,
    KnitPpoBulkColorSizeAggregate,
    KnitPpoAggregateRow,
    MarkerRow,
    MESJoRow,
    PpoColorAggregateRow,
    PpoComboAggregateRow,
    WebmergeColorSizeAggregate,
    WovenPpoPartQtyRow,
    WovenPpoYYRow,
)
from backend.precoi.parsers import (
    classify_go_flow,
    parse_cm_qa_aggregate_rows,
    parse_go_batch,
    parse_go_report,
    parse_knit_ppo_bulk_color_size_aggregates,
    parse_mes_jo_rows,
    parse_webmerge_go_color_size_aggregates,
    parse_webmerge_size_rows,
    parse_woven_ppo_part_qty_rows,
    parse_woven_ppo_yy_rows,
    parse_ypd_marker_rows,
    parse_yy_request,
    split_go_batch,
)
from backend.precoi.services import GetYYService, MULTI_PPO_CHECK_MESSAGE

FIXTURE_DIR = Path(__file__).parent / "fixtures"


def read_fixture(name: str) -> str:
    return (FIXTURE_DIR / name).read_text(encoding="utf-8")


def build_mixed_go_report() -> GOReportData:
    bom_blocks = [
        GOBomBlock(
            flow="WOVEN",
            section_title="Woven Fabric BOM Information",
            block_index=0,
            bom_rows=[
                GOBomRow("YF2600407A", "0.4077", "01(01)", "TRIM FAB1", "1602-64221 01(01) Off trim fabric", "Off", "WOVEN", 0, 0),
                GOBomRow("YF2600407A", "0.4077", "02(75)", "TRIM FAB1", "1602-64221 02(75) Navy trim fabric", "Navy", "WOVEN", 0, 1),
            ],
        ),
        GOBomBlock(
            flow="KNIT",
            section_title="Knit Fabric BOM Information",
            block_index=1,
            bom_rows=[
                GOBomRow("YF2600407B(2)", "0.6463", "01(01)", "MAIN BODY1", "C2106100_0006", "Off", "KNIT", 1, 0),
                GOBomRow("YF2600407B(2)", "0.0264", "01(01)", "TRIM RIB1", "C2203379_0045", "Off", "KNIT", 1, 1),
                GOBomRow("YF2600407B(2)", "0.6448", "02(75)", "MAIN BODY1", "C2304547_0009", "Navy", "KNIT", 1, 2),
                GOBomRow("YF2600407B(2)", "0.0264", "02(75)", "TRIM RIB1", "C2203379_0046", "Navy", "KNIT", 1, 3),
                GOBomRow("", "0", "01(01)", "FK COLLAR1", "F2401289_0178", "Off", "KNIT", 1, 4),
            ],
        ),
    ]
    return GOReportData(
        go_no="S26V00420",
        yy_requests=[parse_yy_request("YF2600407A"), parse_yy_request("YF2600407B(2)")],
        bom_rows=bom_blocks[0].bom_rows + bom_blocks[1].bom_rows,
        bom_blocks=bom_blocks,
    )


def make_record(**overrides) -> ExportRecord:
    payload = {
        "go": "S26V00420",
        "yy_req_no": "",
        "marker_yy": "",
        "ppo_yy": "",
        "gmt_color": "",
        "fabric_part": "",
        "color_code": "",
        "color_desc": "",
        "fabric_color": "",
        "jo": "",
        "minus_pct": "0",
        "plus_pct": "0",
        "qty": "",
        "ppo_no": "",
        "ppo_qty": "",
        "go_key": "S26V00420",
        "row_index": 0,
        "flow": "KNIT",
        "combo_name": "",
        "block_index": 0,
        "section_order": 0,
        "part_order": 0,
        "aggregate_key": "",
        "is_separator": False,
    }
    payload.update(overrides)
    return ExportRecord(**payload)


class FakeResponse:
    def __init__(self, text: str, status_code: int = 200) -> None:
        self.text = text
        self.status_code = status_code
        self.encoding = "utf-8"
        self.apparent_encoding = "utf-8"


class FakeYPDSession:
    def __init__(self) -> None:
        self.auth = None
        self.headers: dict[str, str] = {}
        self.get_calls: list[tuple[str, dict[str, str]]] = []
        self.post_calls: list[tuple[str, dict[str, str], dict[str, str]]] = []

    def get(self, url: str, params: dict[str, str], timeout: int):
        self.get_calls.append((url, dict(params)))
        return FakeResponse('<input type="hidden" name="__VIEWSTATE" value="state" /><input type="hidden" name="__EVENTVALIDATION" value="event" />')

    def post(self, url: str, params: dict[str, str], data: dict[str, str], timeout: int):
        self.post_calls.append((url, dict(params), dict(data)))
        return FakeResponse(read_fixture("ypd_report.html"))


class FakeMesFallbackSession:
    def __init__(self, primary_html: str, secondary_html: str) -> None:
        self.primary_html = primary_html
        self.secondary_html = secondary_html
        self.calls: list[str] = []

    def get(self, url: str, params=None, timeout=None):
        self.calls.append(url)
        if "MES_EAV" in url:
            return FakeResponse(self.secondary_html)
        return FakeResponse(self.primary_html)


class ParserTests(unittest.TestCase):
    def test_parser_and_ypd_basics(self) -> None:
        self.assertEqual(parse_go_batch("s25v08131 S25V08131 A11B00001"), ["S25V08131", "A11B00001"])
        valid, invalid = split_go_batch("S25V08131 bad-go Z99X12345")
        self.assertEqual(valid, ["S25V08131", "Z99X12345"])
        self.assertEqual(invalid, ["bad-go"])
        self.assertEqual(classify_go_flow("S25K11192"), "WOVEN")
        self.assertIsNone(parse_yy_request("YF2600107A").workflow_version_no)
        self.assertEqual(parse_ypd_marker_rows(read_fixture("ypd_report.html"))[0].marker_yy, "0.783")
        self.assertEqual(parse_mes_jo_rows(read_fixture("mes_report.html"), "S26V02332")[0].color_code, "102")

    def test_mes_client_falls_back_to_mes_eav_when_primary_mes_has_no_data(self) -> None:
        primary_html = "<html><body>No data!</body></html>"
        secondary_html = """
        <html><body><table>
        <tr><td>COLOR_CODE</td><td>COLOR_NAME</td><td>JO_NO</td><td>OrderQty</td><td>Over/Short% Allowance</td></tr>
        <tr><td>14</td><td>MEDIUM BLUE</td><td>26K00649ES07</td><td>31</td><td>+0/-0</td></tr>
        </table></body></html>
        """
        client = MESSummaryClient(session=FakeMesFallbackSession(primary_html, secondary_html))

        rows = client.fetch_jo_rows("S26K00649")

        self.assertEqual(
            [(row.color_code, row.color_name, row.jo_no, row.order_qty) for row in rows],
            [("14", "MEDIUM BLUE", "26K00649ES07", "31")],
        )
        self.assertEqual(len(client.session.calls), 2)
        self.assertIn("/MES/DCGoSummaryReport.asp", client.session.calls[0])
        self.assertIn("/MES_EAV/DCGoSummaryReport.asp", client.session.calls[1])

    def test_mes_client_keeps_one_result_when_both_sites_return_same_data(self) -> None:
        same_html = """
        <html><body><table>
        <tr><td>COLOR_CODE</td><td>COLOR_NAME</td><td>JO_NO</td><td>OrderQty</td><td>Over/Short% Allowance</td></tr>
        <tr><td>14</td><td>MEDIUM BLUE</td><td>26K00649ES07</td><td>31</td><td>+0/-0</td></tr>
        </table></body></html>
        """
        client = MESSummaryClient(session=FakeMesFallbackSession(same_html, same_html))

        rows = client.fetch_jo_rows("S26K00649")

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].jo_no, "26K00649ES07")

    def test_mes_client_raises_when_both_sites_return_different_data(self) -> None:
        primary_html = """
        <html><body><table>
        <tr><td>COLOR_CODE</td><td>COLOR_NAME</td><td>JO_NO</td><td>OrderQty</td><td>Over/Short% Allowance</td></tr>
        <tr><td>14</td><td>MEDIUM BLUE</td><td>26K00649ES07</td><td>31</td><td>+0/-0</td></tr>
        </table></body></html>
        """
        secondary_html = """
        <html><body><table>
        <tr><td>COLOR_CODE</td><td>COLOR_NAME</td><td>JO_NO</td><td>OrderQty</td><td>Over/Short% Allowance</td></tr>
        <tr><td>14</td><td>MEDIUM BLUE</td><td>26K00649ES07</td><td>32</td><td>+0/-0</td></tr>
        </table></body></html>
        """
        client = MESSummaryClient(session=FakeMesFallbackSession(primary_html, secondary_html))

        with self.assertRaisesRegex(MESRequestError, "MES data mismatch"):
            client.fetch_jo_rows("S26K00649")

    def test_go_parser_is_span_aware_and_keeps_true_blank_yy(self) -> None:
        html = """
        <html><body><span>Woven Fabric BOM Information</span><table>
        <tr><td>GMT Color Code</td><td>GMT Color Desc</td><td>GMT Part</td><td>Combo Name</td><td>YY Req No</td><td>PPO YY</td></tr>
        <tr><td>01(01)</td><td>Off</td><td>MAIN BODY1</td><td>Combo 1</td><td rowspan="2">YF2600407A</td><td rowspan="2">0.4077</td></tr>
        <tr><td>01(01)</td><td>Off</td><td>TRIM FAB1</td><td>Combo 2</td></tr>
        <tr><td>02(75)</td><td>Navy</td><td>TRIM RIB1</td><td>Combo 3</td><td></td><td></td></tr>
        </table></body></html>
        """
        report = parse_go_report(html, "S25K11192")
        self.assertEqual([row.yy_req_no for row in report.bom_blocks[0].bom_rows], ["YF2600407A", "YF2600407A", ""])

    def test_go_parser_reads_distinct_ppo_mapping_numbers(self) -> None:
        html = """
        <html><body>
        <b>PPO Mapping</b>
        <table>
        <tr><td>Lot</td><td>PPO</td></tr>
        <tr><td>0</td><td>PKEK25VE0010386A</td></tr>
        <tr><td>0</td><td>PKEK25VE0010386B</td></tr>
        <tr><td>1</td><td>PKEK25VE0010386A</td></tr>
        <tr><td>2</td><td>PKEK26VE0010386D</td></tr>
        </table>
        <span>Lot Information</span>
        <table>
        <tr><td>Lot No./JO #</td></tr>
        <tr><td>1/25V10386GB01</td></tr>
        <tr><td>2/25V10386GB02</td></tr>
        </table>
        <span>Knit Fabric BOM Information</span>
        <table>
        <tr><td>GMT Color Code</td><td>GMT Color Desc</td><td>GMT Part</td><td>Combo Name</td><td>YY Req No</td><td>PPO YY</td></tr>
        <tr><td>64A</td><td>OXBLOOD</td><td>MAIN BODY1</td><td>Combo 1</td><td>YF2600107A</td><td>1.1</td></tr>
        </table>
        </body></html>
        """
        report = parse_go_report(html, "S25V10386")
        self.assertEqual(report.ppo_mapping_by_lot, {"1": ["PKEK25VE0010386A"], "2": ["PKEK26VE0010386D"]})
        self.assertEqual(report.jo_lot_map, {"25V10386GB01": "1", "25V10386GB02": "2"})
        self.assertEqual(report.ppo_numbers, ["PKEK25VE0010386A", "PKEK26VE0010386D"])

    def test_go_parser_reads_cm_color_summary_without_bom_rows(self) -> None:
        html = """
        <html><body>
        <b>Color Summary</b>
        <table>
        <tr><td>COLOR CODE</td><td>COLOR DESC</td><td>CUST COLOR CODE</td><td>CUST COLOR DESC</td><td>QUOTATION KEY</td><td>QUOTATION PRICE</td><td>TOTAL QUANTITY</td></tr>
        <tr><td>010</td><td>00A</td><td>010</td><td>00A</td><td></td><td></td><td>7430</td></tr>
        <tr><td>084</td><td>0AJ</td><td>084</td><td>0AJ</td><td></td><td></td><td>7</td></tr>
        <tr><td>Color Total :</td><td></td><td></td><td></td><td></td><td></td><td>7437</td></tr>
        </table>
        <b>Lot Information</b>
        <table>
        <tr><td>Lot No./JO #</td></tr>
        <tr><td>2/26V02155US02</td></tr>
        <tr><td>3/26V02155US03</td></tr>
        </table>
        <b>Color Breakdown -Lot : 2</b>
        <table>
        <tr><td>Gmt Color Code</td><td>Gmt Color Desc.</td><td>Cust Color Code</td><td>Cust Color Desc.</td><td>M</td><td>Total</td></tr>
        <tr><td>010</td><td>00A</td><td>010</td><td>00A</td><td>8</td><td>8</td></tr>
        <tr><td colspan="4">Total</td><td>8</td><td>8</td></tr>
        </table>
        <b>Color Breakdown -Lot : 3</b>
        <table>
        <tr><td>Gmt Color Code</td><td>Gmt Color Desc.</td><td>Cust Color Code</td><td>Cust Color Desc.</td><td>M</td><td>Total</td></tr>
        <tr><td>084</td><td>0AJ</td><td>084</td><td>0AJ</td><td>7</td><td>7</td></tr>
        <tr><td colspan="4">Total</td><td>7</td><td>7</td></tr>
        </table>
        <b>PPO Mapping</b>
        <table><tr><td>Lot</td><td>PPO</td></tr></table>
        </body></html>
        """
        report = parse_go_report(html, "S26V02155")
        self.assertEqual(report.bom_rows, [])
        self.assertEqual([(row.color_code, row.color_desc, row.total_quantity) for row in report.color_summary_rows], [("010", "00A", "7430"), ("084", "0AJ", "7")])
        self.assertEqual(report.jo_lot_map, {"26V02155US02": "2", "26V02155US03": "3"})
        self.assertEqual(
            [(row.lot_no, row.color_code, row.color_desc, row.total_quantity) for row in report.lot_color_rows],
            [("2", "010", "00A", "8"), ("3", "084", "0AJ", "7")],
        )

    def test_cm_qa_parser_aggregates_received_qty_by_combo_and_usage(self) -> None:
        html = """
        <html><body><table>
        <tr><td>InvoiceNo</td><td>&nbsp;</td><td>Usage</td><td>BatchNo</td><td>Rolls</td><td>Size</td><td>Received Qty</td><td>AllowQty</td><td>Pattern</td><td>Shade</td><td>MatchingResult</td><td>Remarks</td><td>ReceiptDate</td></tr>
        <tr><td colspan="13">Combo : 00A</td></tr>
        <tr><td>INV1</td><td></td><td>B</td><td>X</td><td>1</td><td></td><td>1815</td><td>0</td><td></td><td>B</td><td></td><td></td><td>03/25/2026</td></tr>
        <tr><td>INV2</td><td></td><td>B</td><td>X</td><td>1</td><td></td><td>43</td><td>0</td><td></td><td>B</td><td></td><td></td><td>03/25/2026</td></tr>
        <tr><td>INV3</td><td></td><td>M1</td><td>X</td><td>1</td><td></td><td>840.3</td><td>0</td><td></td><td>B</td><td></td><td></td><td>03/25/2026</td></tr>
        <tr><td colspan="13">Combo : 0AJ</td></tr>
        <tr><td>INV4</td><td></td><td>B</td><td>X</td><td>1</td><td></td><td>556</td><td>0</td><td></td><td>A</td><td></td><td></td><td>04/03/2026</td></tr>
        </table></body></html>
        """
        rows = parse_cm_qa_aggregate_rows(html, "S26V02155")
        self.assertEqual(
            [(row.combo_name, row.usage, row.received_qty) for row in rows],
            [("00A", "B", "1858"), ("00A", "M1", "840.3"), ("0AJ", "B", "556")],
        )

    def test_woven_ppo_parser_keeps_blank_yy_rows_when_ppo_yy_exists(self) -> None:
        html = """
        <html><body><table>
        <tr><td>Fabric Part</td><td>Fabric Combo</td><td>YY JOB No</td><td>PPO YY</td></tr>
        <tr><td>TRIM FAB1</td><td>Combo A</td><td>YF2600407A</td><td>0.4077</td></tr>
        <tr><td>TRIM FAB1</td><td>Combo B</td><td></td><td>0.1111</td></tr>
        <tr><td>TRIM FAB1</td><td>Combo C</td><td></td><td>0</td></tr>
        </table></body></html>
        """
        rows = parse_woven_ppo_yy_rows(html, "PWGF26SE001192A")
        self.assertEqual([(row.yy_req_no, row.ppo_yy) for row in rows], [("YF2600407A", "0.4077"), ("", "0.1111")])

    def test_woven_ppo_part_qty_parser_inherits_fabric_combo_for_continuation_rows(self) -> None:
        html = """
        <html><body>
        <span class="bigfont1">TRIM FAB1</span>
        <table>
          <tr><td>Fabric Combo</td><td>Fabric Code</td><td>Lot No.</td><td>Order Qty<br>(Yds)</td></tr>
          <tr><td>1602-65232 trim fab</td><td>KC3030</td><td>1</td><td>139</td></tr>
          <tr><td></td><td>KC3030</td><td>2</td><td>6</td></tr>
          <tr><td></td><td>KC3030</td><td>3</td><td>0.6</td></tr>
          <tr><td>Total:</td><td></td><td></td><td>145.6</td></tr>
        </table>
        </body></html>
        """
        rows = parse_woven_ppo_part_qty_rows(html, "PWGF26VB003162A")
        self.assertEqual(
            [(row.fabric_part, row.fabric_combo, row.ppo_qty) for row in rows],
            [("TRIM FAB1", "1602-65232 trim fab", "145.6")],
        )

    def test_ypd_client_omits_workflow_version_for_woven(self) -> None:
        from backend.precoi.clients import YPDClient

        session = FakeYPDSession()
        client = YPDClient(username="DOMAIN\\user", password="secret", session=session)
        client.fetch_marker_rows(parse_yy_request("YF2600107A"))
        self.assertEqual(session.get_calls[0][1], {"WorkflowNo": "YF2600107A", "ReportType": "1"})
        self.assertEqual(session.post_calls[0][1], {"WorkflowNo": "YF2600107A", "ReportType": "1"})

    def test_webmerge_parser_flattens_split_size_tables(self) -> None:
        html = """
        <html><body>
        <table>
          <tr><td>JO #: 25V11174GB01</td></tr>
          <tr><td>
            <table>
              <tr><td>Colorway# Colorway Desc.</td><td>Colorway# Colorway Desc.</td><td>Cust Colorway#</td><td>Cust Colorway Desc.</td><td>XS</td><td>S</td><td>Total</td></tr>
              <tr><td>W54</td><td>BLACK / ECRU / LAUREL</td><td>W54</td><td>BLACK / ECRU / LAUREL</td><td>5</td><td>0</td><td>5</td></tr>
              <tr><td>Total</td><td>Total</td><td></td><td></td><td>5</td><td>0</td><td>5</td></tr>
            </table>
            <table>
              <tr><td>Colorway# Colorway Desc.</td><td>Colorway# Colorway Desc.</td><td>Cust Colorway#</td><td>Cust Colorway Desc.</td><td>XXXL</td><td>Total</td></tr>
              <tr><td>W54</td><td>BLACK / ECRU / LAUREL</td><td>W54</td><td>BLACK / ECRU / LAUREL</td><td>2</td><td>7</td></tr>
            </table>
          </td></tr>
        </table>
        </body></html>
        """
        rows = parse_webmerge_size_rows(html, "S25V11174")
        self.assertEqual(
            [(row.jo_no, row.color_code, row.size, row.qty) for row in rows],
            [("25V11174GB01", "W54", "XS", "5"), ("25V11174GB01", "W54", "XXXL", "2")],
        )

    def test_webmerge_go_parser_reads_garment_level_breakdown_only(self) -> None:
        html = """
        <html><body>
        <table><tr><td>Color/Size Breakdown</td></tr></table>
        <table>
          <tr><td></td><td>Colorway Desc. Colorway#</td><td>Cust Colorway#</td><td>Cust Colorway Desc.</td><td>Is Sensitive Color</td><td>XS</td><td>S</td><td>M</td><td>L</td><td>XL</td><td>XXL</td></tr>
          <tr><td></td><td>64A</td><td>OXBLOOD / ECRU / ECRU</td><td>64A</td><td>OXBLOOD / ECRU / ECRU</td><td>NO</td><td>0</td><td>273</td><td>762</td><td>915</td><td>582</td><td>268</td></tr>
          <tr><td></td><td>W54</td><td>BLACK / ECRU / LAUREL</td><td>W54</td><td>BLACK / ECRU / LAUREL</td><td>NO</td><td>5</td><td>6</td><td>0</td><td>0</td><td>0</td><td>0</td></tr>
        </table>
        <table>
          <tr><td></td><td>Colorway Desc. Colorway#</td><td>Cust Colorway#</td><td>Cust Colorway Desc.</td><td>Is Sensitive Color</td><td>XXXL</td><td>Total</td></tr>
          <tr><td></td><td>64A</td><td>OXBLOOD / ECRU / ECRU</td><td>64A</td><td>OXBLOOD / ECRU / ECRU</td><td>NO</td><td>0</td><td>2800</td></tr>
          <tr><td></td><td>W54</td><td>BLACK / ECRU / LAUREL</td><td>W54</td><td>BLACK / ECRU / LAUREL</td><td>NO</td><td>2</td><td>13</td></tr>
        </table>
        <table><tr><td>JO #: 25V11174GB01</td></tr></table>
        <table>
          <tr><td>Colorway# Colorway Desc.</td><td>Colorway# Colorway Desc.</td><td>Cust Colorway#</td><td>Cust Colorway Desc.</td><td>XS</td><td>S</td><td>Total</td></tr>
          <tr><td>64A</td><td>OXBLOOD / ECRU / ECRU</td><td>64A</td><td>OXBLOOD / ECRU / ECRU</td><td>0</td><td>999</td><td>999</td></tr>
        </table>
        </body></html>
        """
        rows = parse_webmerge_go_color_size_aggregates(html, "S25V11174")
        self.assertEqual(
            [(row.color_code, row.size, row.qty) for row in rows],
            [
                ("64A", "S", "273"),
                ("64A", "M", "762"),
                ("64A", "L", "915"),
                ("64A", "XL", "582"),
                ("64A", "XXL", "268"),
                ("W54", "XS", "5"),
                ("W54", "S", "6"),
                ("W54", "XXXL", "2"),
            ],
        )

    def test_webmerge_go_parser_accepts_truncated_first_desc_when_second_desc_is_full(self) -> None:
        html = """
        <html><body>
        <table><tr><td>Color/Size Breakdown</td></tr></table>
        <table>
          <tr><td></td><td>Colorway Desc. Colorway#</td><td>Cust Colorway#</td><td>Cust Colorway Desc.</td><td>Is Sensitive Color</td><td>XS</td><td>S</td><td>M</td><td>L</td><td>XL</td><td>XXL</td></tr>
          <tr><td></td><td>JEP0459COB</td><td>Cobalt blue solid short sleeve cotton stretch piqu</td><td>JEP0459COB</td><td>Cobalt blue solid short sleeve cotton stretch pique polo</td><td>NO</td><td>5</td><td>97</td><td>347</td><td>387</td><td>221</td><td>101</td></tr>
        </table>
        <table>
          <tr><td></td><td>Colorway Desc. Colorway#</td><td>Cust Colorway#</td><td>Cust Colorway Desc.</td><td>Is Sensitive Color</td><td>XXXL</td><td>Total</td></tr>
          <tr><td></td><td>JEP0459COB</td><td>Cobalt blue solid short sleeve cotton stretch piqu</td><td>JEP0459COB</td><td>Cobalt blue solid short sleeve cotton stretch pique polo</td><td>NO</td><td>43</td><td>1201</td></tr>
        </table>
        </body></html>
        """
        rows = parse_webmerge_go_color_size_aggregates(html, "S25V10386")
        self.assertEqual(
            [(row.color_code, row.color_desc, row.size, row.qty) for row in rows],
            [
                ("JEP0459COB", "Cobalt blue solid short sleeve cotton stretch pique polo", "XS", "5"),
                ("JEP0459COB", "Cobalt blue solid short sleeve cotton stretch pique polo", "S", "97"),
                ("JEP0459COB", "Cobalt blue solid short sleeve cotton stretch pique polo", "M", "347"),
                ("JEP0459COB", "Cobalt blue solid short sleeve cotton stretch pique polo", "L", "387"),
                ("JEP0459COB", "Cobalt blue solid short sleeve cotton stretch pique polo", "XL", "221"),
                ("JEP0459COB", "Cobalt blue solid short sleeve cotton stretch pique polo", "XXL", "101"),
                ("JEP0459COB", "Cobalt blue solid short sleeve cotton stretch pique polo", "XXXL", "43"),
            ],
        )

    def test_webmerge_go_parser_accepts_dot_and_single_character_color_codes(self) -> None:
        html = """
        <html><body>
        <table><tr><td>Color/Size Breakdown</td></tr></table>
        <table>
          <tr><td>Color Code</td><td>Color Desc.</td><td>Cust Color Code</td><td>Cust Color Desc</td><td>XS</td><td>S</td><td>M</td><td>L</td><td>Total</td></tr>
          <tr><td>1.</td><td>BRD</td><td>1.</td><td>BRD</td><td>20</td><td>65</td><td>40</td><td>15</td><td>140</td></tr>
          <tr><td>2</td><td>BKS</td><td>2</td><td>BKS</td><td>30</td><td>90</td><td>60</td><td>20</td><td>200</td></tr>
        </table>
        </body></html>
        """
        rows = parse_webmerge_go_color_size_aggregates(html, "S26V01807")
        self.assertEqual(
            [(row.color_code, row.color_desc, row.size, row.qty) for row in rows],
            [
                ("1.", "BRD", "XS", "20"),
                ("1.", "BRD", "S", "65"),
                ("1.", "BRD", "M", "40"),
                ("1.", "BRD", "L", "15"),
                ("2", "BKS", "XS", "30"),
                ("2", "BKS", "S", "90"),
                ("2", "BKS", "M", "60"),
                ("2", "BKS", "L", "20"),
            ],
        )

    def test_webmerge_go_parser_accepts_numeric_sizes(self) -> None:
        html = """
        <html><body>
        <table><tr><td>Color/Size Breakdown</td></tr></table>
        <table>
          <tr><td>Color Code</td><td>Color Desc.</td><td>Cust Color Code</td><td>Cust Color Desc</td><td>95</td><td>100</td><td>105</td><td>110</td><td>115</td><td>Total</td></tr>
          <tr><td>BKS</td><td>BKS</td><td>BKS</td><td>BKS</td><td>200</td><td>380</td><td>370</td><td>170</td><td>80</td><td>1200</td></tr>
          <tr><td>DGS</td><td>DGS</td><td>DGS</td><td>DGS</td><td>260</td><td>380</td><td>250</td><td>110</td><td>0</td><td>1000</td></tr>
        </table>
        </body></html>
        """
        rows = parse_webmerge_go_color_size_aggregates(html, "S26V00960")
        self.assertEqual(
            [(row.color_code, row.size, row.qty) for row in rows],
            [
                ("BKS", "95", "200"),
                ("BKS", "100", "380"),
                ("BKS", "105", "370"),
                ("BKS", "110", "170"),
                ("BKS", "115", "80"),
                ("DGS", "95", "260"),
                ("DGS", "100", "380"),
                ("DGS", "105", "250"),
                ("DGS", "110", "110"),
            ],
        )

    def test_knit_ppo_bulk_parser_aggregates_color_size_by_part(self) -> None:
        html = """
        <html><body>
        <table><tr><td>FK COLLAR1 - O</td></tr></table>
        <table>
          <tr><td>Gmt Color Code</td><td>Combo\\Size (Length x Height)</td><td>Production Status</td><td>Lock Status</td><td>XS (41.8X 3cm)</td><td>S (43.3X 3cm)</td><td>Total</td></tr>
          <tr><td>64A</td><td>OXBLOOD / ECRU / ECRU</td><td>AR</td><td>Y</td><td>2</td><td>5</td><td>7</td></tr>
          <tr><td>68C</td><td>ECRU / SEAGRASS / DARK AIRFORCE</td><td>AR</td><td>Y</td><td>0</td><td>4</td><td>4</td></tr>
          <tr><td>Total</td><td></td><td></td><td></td><td>2</td><td>9</td><td>11</td></tr>
        </table>
        <table>
          <tr><td>Gmt Color Code</td><td>Combo\\Size (Length x Height)</td><td>Production Status</td><td>Lock Status</td><td>XS (41.8X 3cm)</td><td>S (43.3X 3cm)</td><td>Total</td></tr>
          <tr><td>64A</td><td>OXBLOOD / ECRU / ECRU</td><td>AR</td><td>Y</td><td>1</td><td>0</td><td>1</td></tr>
        </table>
        <table><tr><td>FK CUFF1 - F</td></tr></table>
        <table>
          <tr><td>Gmt Color Code</td><td>Combo\\Size (Length x Height)</td><td>Production Status</td><td>Lock Status</td><td>M (35.1X 3.7cm)</td><td>XXXL (43.1X 3.7cm)</td><td>Total</td></tr>
          <tr><td>64A</td><td>OXBLOOD / ECRU / ECRU</td><td>AR</td><td>Y</td><td>8</td><td>1</td><td>9</td></tr>
        </table>
        </body></html>
        """
        rows = parse_knit_ppo_bulk_color_size_aggregates(html, "PKEK25VE0011174A")
        self.assertEqual(
            [(row.fabric_part, row.color_code, row.size, row.ppo_qty) for row in rows],
            [
                ("FK COLLAR1", "64A", "XS", "3"),
                ("FK COLLAR1", "64A", "S", "5"),
                ("FK COLLAR1", "68C", "S", "4"),
                ("FK CUFF1", "64A", "M", "8"),
                ("FK CUFF1", "64A", "XXXL", "1"),
            ],
        )

    def test_knit_ppo_bulk_parser_accepts_fk_bottom_sections(self) -> None:
        html = """
        <html><body>
        <table><tr><td>FK BOTTOM1 - TC</td></tr></table>
        <table>
          <tr><td>Gmt Color Code</td><td>Combo\\Size (Length x Height)</td><td>Production Status</td><td>Lock Status</td><td>S (43.3X 3cm)</td><td>M (46X3cm)</td><td>Total</td></tr>
          <tr><td>BK</td><td>BLACK</td><td>AR</td><td>Y</td><td>10</td><td>20</td><td>30</td></tr>
        </table>
        <table><tr><td>FK BOTTOM2 - TD</td></tr></table>
        <table>
          <tr><td>Gmt Color Code</td><td>Combo\\Size (Length x Height)</td><td>Production Status</td><td>Lock Status</td><td>S (43.3X 3cm)</td><td>M (46X3cm)</td><td>Total</td></tr>
          <tr><td>BK</td><td>BLACK</td><td>AR</td><td>Y</td><td>11</td><td>21</td><td>32</td></tr>
        </table>
        </body></html>
        """
        rows = parse_knit_ppo_bulk_color_size_aggregates(html, "PKGK26SB0001983A")
        self.assertEqual(
            [(row.fabric_part, row.color_code, row.size, row.ppo_qty) for row in rows],
            [
                ("FK BOTTOM1", "BK", "S", "10"),
                ("FK BOTTOM1", "BK", "M", "20"),
                ("FK BOTTOM2", "BK", "S", "11"),
                ("FK BOTTOM2", "BK", "M", "21"),
            ],
        )


class WorkbookTests(unittest.TestCase):
    def _assert_hidden_columns(self, worksheet, visible_count: int, total_count: int) -> None:
        for column_index in range(visible_count + 1, total_count + 1):
            hidden_by_dimension = any(
                dimension.hidden and (dimension.min or 0) <= column_index <= (dimension.max or 0)
                for dimension in worksheet.column_dimensions.values()
            )
            self.assertTrue(hidden_by_dimension, f"Expected column {get_column_letter(column_index)} to stay hidden on {worksheet.title}")

    def test_workbook_round_trip_uses_new_schema_and_accepts_legacy(self) -> None:
        records = [
            make_record(
                yy_req_no="YF2600407A",
                marker_yy="0.402",
                gmt_color="Off",
                fabric_part="TRIM FAB1",
                color_code="01(01)",
                color_desc="OFF",
                jo="26V00420GB01",
                qty="100",
                flow="WOVEN",
                combo_name="Woven Combo 1",
                aggregate_key="S26V00420|WOVEN|TRIM FAB1|01(01)|WOVEN COMBO 1",
            ),
            make_record(
                row_index=1,
                gmt_color="Off",
                fabric_part="FK COLLAR1",
                color_code="01(01)",
                color_desc="OFF",
                jo="26V00420GB01",
                qty="100",
                ppo_no="PPO001",
                ppo_qty="500",
                flow="KNIT",
                combo_name="Collar Combo 1",
                block_index=1,
                section_order=1,
                aggregate_key="S26V00420|KNIT|FK COLLAR1|01(01)|COLLAR COMBO 1",
            ),
        ]
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / "coi.xlsx"
            write_workbook(file_path, records)
            sheet = load_workbook(file_path)["COI"]
            agg_col = get_column_letter(AGGREGATE_KEY_COLUMN)
            self.assertEqual([sheet.cell(row=1, column=i).value for i in range(1, len(ALL_HEADERS) + 1)], ALL_HEADERS)
            self.assertIsInstance(sheet["C2"].value, float)
            self.assertIsInstance(sheet["K2"].value, int)
            self.assertEqual(sheet["M2"].value, '=IF(OR(C2=0,C2="",K2=0,K2=""),"-",IFERROR(K2*C2,0))')
            self.assertEqual(sheet["N2"].value, f'=IF(OR(L2="",M2="-",M2="",{agg_col}2=""),"",SUMIFS(M:M,{agg_col}:{agg_col},{agg_col}2,L:L,L2))')
            self.assertEqual(sheet["P2"].value, '=IF(OR(L2="",N2=0,N2="",O2=""),"",O2/N2)')
            self.assertEqual([str(item) for item in sheet.conditional_formatting], ["<ConditionalFormatting E2:E1048576>", "<ConditionalFormatting P2:P1048576>"])
            self.assertEqual(sheet["A1"].border.left.style, "thin")
            self.assertEqual(sheet["A2"].border.left.style, "thin")
            self._assert_hidden_columns(sheet, len(VISIBLE_HEADERS), len(ALL_HEADERS))
            self.assertEqual(read_records_from_workbook(file_path)[1].fabric_part, "FK COLLAR1")

            legacy = Path(tmp_dir) / "legacy.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "COI"
            ws.append(LEGACY_VISIBLE_HEADERS)
            ws.append(["S25V11174", "YF2511709B(1)", "0.7914", "0.8018", "", "86C - RICH BROWN", "MAIN BODY1", "B", "86C", "RICH BROWN", "25V11174NL13", "724", "PPO001", "", "", "2277", ""])
            wb.save(legacy)
            self.assertEqual(read_records_from_workbook(legacy)[0].color_code, "86C")

    def test_write_workbook_compacts_duplicate_rows_by_summing_qty_and_ppo_qty(self) -> None:
        records = [
            make_record(
                yy_req_no="YF2600407A",
                marker_yy="0.402",
                ppo_yy="0.4077",
                gmt_color="Off",
                fabric_part="TRIM FAB1",
                color_code="01(01)",
                color_desc="OFF",
                jo="26V00420GB01",
                qty="100",
                ppo_no="PWGF26VB000420A",
                ppo_qty="400",
                flow="WOVEN",
                combo_name="1602-64221 01(01) Off trim fabric",
                aggregate_key="S26V00420|WOVEN|TRIM FAB1|01(01)|1602-64221 01(01) OFF TRIM FABRIC",
            ),
            make_record(
                row_index=1,
                yy_req_no="YF2600407A",
                marker_yy="0.402",
                ppo_yy="0.4077",
                gmt_color="Off",
                fabric_part="TRIM FAB1",
                color_code="01(01)",
                color_desc="OFF",
                jo="26V00420GB01",
                qty="25",
                ppo_no="PWGF26VB000420A",
                ppo_qty="10",
                flow="WOVEN",
                combo_name="1602-64221 01(01) Off trim fabric",
                aggregate_key="S26V00420|WOVEN|TRIM FAB1|01(01)|1602-64221 01(01) OFF TRIM FABRIC",
            ),
        ]
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "compact.xlsx"
            write_workbook(path, records)
            reloaded = read_records_from_workbook(path)

        self.assertEqual(len(reloaded), 1)
        self.assertEqual(reloaded[0].qty, "125")
        self.assertEqual(reloaded[0].ppo_qty, "410")

    def test_write_workbook_creates_collar_sheet_and_round_trips_hidden_yy(self) -> None:
        records = [
            make_record(
                go="S25V11174",
                go_key="S25V11174",
                yy_req_no="YF2511709B(1)",
                marker_yy="0.7914",
                ppo_yy="0.8018",
                gmt_color="BLACK / ECRU / LAUREL",
                fabric_part="FK COLLAR1",
                color_code="W54",
                color_desc="BLACK / ECRU / LAUREL",
                qty="7",
                ppo_no="PKEK25VE0011174A",
                ppo_qty="200",
                flow="KNIT",
                combo_name="Collar Combo 1",
                aggregate_key="S25V11174|COI COLLAR／CUFF|KNIT|FK COLLAR1|W54|COLLAR COMBO 1|XS",
                sheet_kind=COLLAR_SHEET_NAME,
                size="XS",
            )
        ]
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "collar.xlsx"
            write_workbook(path, records)
            workbook = load_workbook(path)
            self.assertIn(COLLAR_SHEET_NAME, workbook.sheetnames)
            collar_sheet = workbook[COLLAR_SHEET_NAME]
            self.assertEqual(
                [collar_sheet.cell(row=1, column=i).value for i in range(1, len(COLLAR_ALL_HEADERS) + 1)],
                COLLAR_ALL_HEADERS,
            )
            self.assertEqual(collar_sheet["A2"].value, "S25V11174")
            self.assertEqual(collar_sheet["F2"].value, "XS")
            self.assertIsInstance(collar_sheet["G2"].value, int)
            self.assertIsInstance(collar_sheet["I2"].value, int)
            self.assertEqual(collar_sheet["J2"].value, '=IF(OR(H2="",G2=0,G2="",I2=""),"",I2/G2)')
            self.assertEqual([str(item) for item in collar_sheet.conditional_formatting], ["<ConditionalFormatting J2:J1048576>"])
            self.assertEqual(collar_sheet["A1"].border.left.style, "thin")
            self.assertEqual(collar_sheet["A2"].border.left.style, "thin")
            self.assertEqual(collar_sheet["T2"].value, "YF2511709B(1)")
            self._assert_hidden_columns(collar_sheet, len(COLLAR_VISIBLE_HEADERS), len(COLLAR_ALL_HEADERS))
            self.assertEqual(read_records_from_workbook(path)[0].sheet_kind, COLLAR_SHEET_NAME)

    def test_write_workbook_cm_uses_single_sheet_with_cm_schema(self) -> None:
        records = [
            make_record(
                go="S26V02155",
                go_key="S26V02155",
                gmt_color="00A",
                fabric_part="B",
                color_code="010",
                color_desc="00A",
                jo="",
                qty="7430",
                flow="CM",
                combo_name="00A",
                aggregate_key="S26V02155|COI|CM|B|010|00A|",
            )
        ]
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "cm.xlsx"
            write_workbook(path, records)
            workbook = load_workbook(path)
            self.assertEqual(workbook.sheetnames, ["COI"])
            sheet = workbook["COI"]
            self.assertEqual(
                [sheet.cell(row=1, column=i).value for i in range(1, len(CM_ALL_HEADERS) + 1)],
                CM_ALL_HEADERS,
            )
            self.assertEqual(sheet["A2"].value, "S26V02155")
            self.assertEqual(sheet["C2"].value, "B")
            self.assertIsInstance(sheet["G2"].value, int)
            self.assertEqual(sheet["J2"].value, '=IF(OR(G2=0,G2="",I2=""),"",I2/G2)')
            self._assert_hidden_columns(sheet, len(CM_VISIBLE_HEADERS), len(CM_ALL_HEADERS))
            self.assertEqual(read_records_from_workbook(path)[0].flow, "CM")

    def test_save_master_workbook_writes_fixed_name(self) -> None:
        records = [make_record(go="S25V11174", go_key="S25V11174", color_code="86C", color_desc="RICH BROWN", jo="25V11174NL13", minus_pct="3", plus_pct="3", qty="724")]
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = save_master_workbook(records, tmp_dir)
            self.assertEqual(path.name, MASTER_FILE_NAME)


class FakeMixedGOClient:
    def fetch_go_report(self, go_no: str) -> GOReportData:
        return build_mixed_go_report()


class FakeMixedMESClient:
    def fetch_jo_rows(self, go_no: str) -> list[MESJoRow]:
        return [
            MESJoRow("26V00420GB01", "100", "01(01)", "OFF", "0", "0", ""),
            MESJoRow("26V00420GB02", "120", "02(75)", "NAVY", "0", "0", ""),
        ]


class FakeMixedYPDClient:
    instances: list["FakeMixedYPDClient"] = []

    def __init__(self, username: str, password: str) -> None:
        self.calls: list[str] = []
        self.__class__.instances.append(self)

    def fetch_marker_rows(self, yy_request) -> list[MarkerRow]:
        self.calls.append(yy_request.raw_value)
        if yy_request.raw_value == "YF2600407A":
            return [MarkerRow("0.402", "01(01) - Off", "TRIM FAB1", "01(01)"), MarkerRow("0.402", "02(75) - Navy", "TRIM FAB1", "02(75)")]
        return [MarkerRow("0.6379", "01(01) - Off", "MAIN BODY1", "01(01)"), MarkerRow("0.0261", "01(01) - Off", "TRIM RIB1", "01(01)"), MarkerRow("0.6364", "02(75) - Navy", "MAIN BODY1", "02(75)"), MarkerRow("0.0261", "02(75) - Navy", "TRIM RIB1", "02(75)")]


class FakeWovenVersionYPDClient:
    instances: list["FakeWovenVersionYPDClient"] = []

    def __init__(self, username: str, password: str) -> None:
        self.calls: list[str] = []
        self.__class__.instances.append(self)

    def fetch_marker_rows(self, yy_request) -> list[MarkerRow]:
        self.calls.append(yy_request.raw_value)
        if yy_request.workflow_version_no is None:
            return [MarkerRow("9.9", "996 - Wrong", "MAIN BODY1", "996")]
        if yy_request.workflow_version_no == "3":
            return [MarkerRow("0.402", "01(01) - Off", "TRIM FAB1", "01(01)"), MarkerRow("0.402", "02(75) - Navy", "TRIM FAB1", "02(75)")]
        raise YPDRequestError("Version not available")


class FakePpoClient:
    def __init__(self, color_responses, knit_part_responses=None, woven_responses=None) -> None:
        self.color_responses = color_responses
        self.knit_part_responses = knit_part_responses or {}
        self.woven_responses = woven_responses or {}

    def fetch_color_aggregates(self, ppo_no: str):
        return self.color_responses.get(ppo_no, [])

    def fetch_knit_part_aggregates(self, ppo_no: str):
        return self.knit_part_responses.get(ppo_no, [])

    def fetch_woven_combo_aggregates(self, ppo_no: str):
        return self.woven_responses.get(ppo_no, [])


class FakeWovenPpoReportClient:
    def __init__(self, responses, collar_cuff_responses=None, part_qty_responses=None) -> None:
        self.responses = responses
        self.collar_cuff_responses = collar_cuff_responses or {}
        self.part_qty_responses = part_qty_responses or {}

    def fetch_woven_ppo_yy_rows(self, ppo_no: str):
        return self.responses.get(ppo_no, [])

    def fetch_knit_collar_cuff_aggregates(self, ppo_no: str):
        return self.collar_cuff_responses.get(ppo_no, [])

    def fetch_woven_part_qty_rows(self, ppo_no: str):
        return self.part_qty_responses.get(ppo_no, [])


class FakeCMQaClient:
    def __init__(self, responses) -> None:
        self.responses = responses

    def fetch_aggregate_rows(self, ppo_no: str):
        return list(self.responses.get(ppo_no, []))


class FakeWebmergeClient:
    def __init__(self, rows, size_rows=None) -> None:
        self.rows = rows
        self.size_rows = size_rows or []

    def fetch_size_rows(self, go_no: str):
        return list(self.size_rows)

    def fetch_go_color_size_aggregates(self, go_no: str):
        return list(self.rows)


class ServiceTests(unittest.TestCase):
    def test_split_ppo_values_extracts_tokens_from_pasted_mapping_text(self) -> None:
        self.assertEqual(
            GetYYService._split_ppo_values("Lot\tPPO\n0\tPKGK26VB0003162A\n1\tPKGK26VB0003162A\n1\tPWGF26VB003162A"),
            ["PKGK26VB0003162A", "PWGF26VB003162A"],
        )

    def test_marker_match_uses_single_part_fallback_for_artwork_color_codes(self) -> None:
        logs: list[str] = []
        row = GetYYService._match_marker_row(
            "S26V04873",
            "YF2605217B(2)",
            GOBomRow("YF2605217B(2)", "0.7078", "001", "MAIN BODY1", "WHITE", "UNISEX-TRANG-PHO"),
            [MarkerRow("0.6951", "1 - signature white", "MAIN BODY1", "1")],
            logs.append,
        )
        self.assertIsNotNone(row)
        self.assertEqual(row.marker_yy, "0.6951")
        self.assertTrue(any("single Marker YY part fallback" in message for message in logs))

    def test_create_output_keeps_blank_yy_rows_and_separator(self) -> None:
        FakeMixedYPDClient.instances = []
        service = GetYYService(
            go_client=FakeMixedGOClient(),
            mes_client=FakeMixedMESClient(),
            webmerge_client=FakeWebmergeClient([]),
            ppo_client=FakePpoClient({}),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            with patch("backend.precoi.services.YPDClient", FakeMixedYPDClient):
                output_path = service.create_output("S26V00420", "user", "secret", tmp_dir)
            sheet = load_workbook(output_path)["COI"]
            records = read_records_from_workbook(output_path)

        self.assertEqual(len(records), 7)
        self.assertIsNone(sheet["A4"].value)
        self.assertEqual(records[-1].fabric_part, "FK COLLAR1")
        self.assertEqual(records[-1].yy_req_no, "")
        self.assertEqual(sorted(set(FakeMixedYPDClient.instances[0].calls)), ["YF2600407A", "YF2600407B(2)"])

    def test_create_output_woven_resolves_version_by_part_and_color(self) -> None:
        block = GOBomBlock("WOVEN", "Woven Fabric BOM Information", 0, [GOBomRow("YF2600407A", "0.4077", "01(01)", "TRIM FAB1", "Woven Combo 1", "Off", "WOVEN", 0, 0), GOBomRow("YF2600407A", "0.4077", "02(75)", "TRIM FAB1", "Woven Combo 2", "Navy", "WOVEN", 0, 1)])

        class SingleGOClient:
            def fetch_go_report(self, go_no: str) -> GOReportData:
                return GOReportData(go_no, [parse_yy_request("YF2600407A")], block.bom_rows, [block])

        service = GetYYService(
            go_client=SingleGOClient(),
            mes_client=FakeMixedMESClient(),
            webmerge_client=FakeWebmergeClient([]),
            ppo_client=FakePpoClient({}),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            with patch("backend.precoi.services.YPDClient", FakeWovenVersionYPDClient):
                output_path = service.create_output("S26V00420", "user", "secret", tmp_dir)
            self.assertEqual(load_workbook(output_path)["COI"]["C2"].value, 0.402)

        self.assertIn("YF2600407A(3)", FakeWovenVersionYPDClient.instances[0].calls)

    def test_update_ppo_qty_is_best_effort_for_blank_yy_rows(self) -> None:
        records = [
            make_record(yy_req_no="YF2600407A", marker_yy="0.402", gmt_color="Off", fabric_part="TRIM FAB1", color_code="01(01)", color_desc="OFF", jo="26V00420GB01", qty="100", ppo_no="PWGF26SE00420A", flow="WOVEN", combo_name="1602-64221 01(01) Off trim fabric", aggregate_key="S26V00420|WOVEN|TRIM FAB1|01(01)|1602-64221 01(01) OFF TRIM FABRIC"),
            make_record(row_index=1, gmt_color="Navy", fabric_part="TRIM FAB1", color_code="02(75)", color_desc="NAVY", jo="26V00420GB02", qty="120", ppo_no="PWGF26SE00420A", flow="WOVEN", combo_name="1602-64221 02(75) Navy trim fabric", part_order=1, aggregate_key="S26V00420|WOVEN|TRIM FAB1|02(75)|1602-64221 02(75) NAVY TRIM FABRIC"),
            make_record(row_index=2, gmt_color="White", fabric_part="TRIM FAB1", color_code="03(99)", color_desc="WHITE", jo="26V00420GB03", qty="90", ppo_no="PWGF26SE00420A", flow="WOVEN", combo_name="UNMAPPED COMBO", part_order=2, aggregate_key="S26V00420|WOVEN|TRIM FAB1|03(99)|UNMAPPED COMBO"),
            make_record(row_index=3, yy_req_no="YF2600407B(2)", marker_yy="0.6379", ppo_yy="0.6463", gmt_color="Off", fabric_part="MAIN BODY1", color_code="01(01)", color_desc="OFF", jo="26V00420GB01", qty="100", ppo_no="PKEK26VE000420A", flow="KNIT", combo_name="C2106100_0006", block_index=1, section_order=1, aggregate_key="S26V00420|KNIT|MAIN BODY1|01(01)|C2106100_0006"),
        ]
        service = GetYYService(
            ppo_client=FakePpoClient(
                {"PKEK26VE000420A": [PpoColorAggregateRow("PKEK26VE000420A", "B", "01(01)", "1000")]},
                {"PKEK26VE000420A": [KnitPpoAggregateRow("PKEK26VE000420A", "B", "MAIN BODY1", "C2106100_0006", "Off", "1000")]},
                {"PWGF26SE00420A": [PpoComboAggregateRow("PWGF26SE00420A", "BD", "1602-64221 01(01) Off trim fabric", "400"), PpoComboAggregateRow("PWGF26SE00420A", "M1", "1602-64221 02(75) Navy trim fabric", "50")]},
            ),
            ppo_report_client=FakeWovenPpoReportClient(
                {"PWGF26SE00420A": [WovenPpoYYRow("PWGF26SE00420A", "YF2600407A", "TRIM FAB1", "1602-64221 01(01) Off trim fabric", "0.4077"), WovenPpoYYRow("PWGF26SE00420A", "", "TRIM FAB1", "1602-64221 02(75) Navy trim fabric", "0.1111")]}
            ),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / MASTER_FILE_NAME
            write_workbook(file_path, records)
            updated = read_records_from_workbook(service.update_ppo_qty_from_workbook(file_path))

        self.assertEqual([updated[0].ppo_yy, updated[1].ppo_yy, updated[2].ppo_yy, updated[3].ppo_qty], ["0.4077", "0.1111", "", "1000"])
        self.assertEqual([updated[0].ppo_qty, updated[1].ppo_qty, updated[2].ppo_qty], ["400", "50", ""])

    def test_create_output_cm_uses_go_color_summary_and_go_as_ppono(self) -> None:
        class CmGOClient:
            def fetch_go_report(self, go_no: str) -> GOReportData:
                return GOReportData(
                    go_no=go_no,
                    yy_requests=[],
                    bom_rows=[],
                    color_summary_rows=[
                        GoColorSummaryRow("010", "00A", "7430"),
                        GoColorSummaryRow("084", "0AJ", "7"),
                    ],
                    jo_lot_map={
                        "26V02155US02": "2",
                        "26V02155US04": "4",
                        "26V02155US06": "6",
                        "26V02155US03": "3",
                    },
                    lot_color_rows=[
                        GoLotColorRow("2", "010", "00A", "8"),
                        GoLotColorRow("4", "010", "00A", "498"),
                        GoLotColorRow("6", "010", "00A", "6924"),
                        GoLotColorRow("3", "084", "0AJ", "7"),
                    ],
                )

        service = GetYYService(
            go_client=CmGOClient(),
            cm_client=FakeCMQaClient(
                {
                    "S26V02155": [
                        CmQaAggregateRow("S26V02155", "00A", "B", "1858"),
                        CmQaAggregateRow("S26V02155", "00A", "M1", "840.3"),
                        CmQaAggregateRow("S26V02155", "0AJ", "B", "556"),
                    ]
                }
            ),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = service.create_output("S26V02155", "", "", tmp_dir)
            workbook = load_workbook(output_path)
            records = read_records_from_workbook(output_path)

        self.assertEqual(workbook.sheetnames, ["COI"])
        self.assertEqual(
            [(row.flow, row.color_desc, row.fabric_part, row.qty, row.jo) for row in records],
            [
                ("CM", "00A", "B", "7430", "26V02155US02"),
                ("CM", "00A", "B", "7430", "26V02155US04"),
                ("CM", "00A", "B", "7430", "26V02155US06"),
                ("CM", "00A", "M1", "7430", "26V02155US02"),
                ("CM", "00A", "M1", "7430", "26V02155US04"),
                ("CM", "00A", "M1", "7430", "26V02155US06"),
                ("CM", "0AJ", "B", "7", "26V02155US03"),
            ],
        )

    def test_update_cm_from_go_fills_cm_qty_and_keeps_visible_ppo_blank(self) -> None:
        class CmGOClient:
            def fetch_go_report(self, go_no: str) -> GOReportData:
                return GOReportData(
                    go_no=go_no,
                    yy_requests=[],
                    bom_rows=[],
                    color_summary_rows=[
                        GoColorSummaryRow("010", "00A", "7430"),
                    ],
                    jo_lot_map={
                        "26V02155US02": "2",
                        "26V02155US04": "4",
                        "26V02155US06": "6",
                    },
                    lot_color_rows=[
                        GoLotColorRow("2", "010", "00A", "8"),
                        GoLotColorRow("4", "010", "00A", "498"),
                        GoLotColorRow("6", "010", "00A", "6924"),
                    ],
                )

        service = GetYYService(
            go_client=CmGOClient(),
            cm_client=FakeCMQaClient(
                {
                    "S26V02155": [
                        CmQaAggregateRow("S26V02155", "00A", "B", "1858"),
                        CmQaAggregateRow("S26V02155", "00A", "M1", "840.3"),
                    ]
                }
            ),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = service.update_cm_from_go("S26V02155", tmp_dir)
            workbook = load_workbook(output_path)
            records = read_records_from_workbook(output_path)

        self.assertEqual(workbook.sheetnames, ["COI"])
        self.assertEqual(
            [(row.fabric_part, row.ppo_no, row.ppo_qty, row.jo) for row in records],
            [
                ("B", "", "1858", "26V02155US02"),
                ("B", "", "1858", "26V02155US04"),
                ("B", "", "1858", "26V02155US06"),
                ("M1", "", "840.3", "26V02155US02"),
                ("M1", "", "840.3", "26V02155US04"),
                ("M1", "", "840.3", "26V02155US06"),
            ],
        )

    def test_update_ppo_qty_from_workbook_for_cm_uses_combo_and_usage(self) -> None:
        records = [
            make_record(
                go="S26V02155",
                go_key="S26V02155",
                gmt_color="00A",
                fabric_part="B",
                color_code="010",
                color_desc="00A",
                jo="",
                qty="7430",
                ppo_no="PPO001",
                flow="CM",
                combo_name="00A",
                aggregate_key="S26V02155|COI|CM|B|010|00A|",
            ),
            make_record(
                row_index=1,
                go="S26V02155",
                go_key="S26V02155",
                gmt_color="00A",
                fabric_part="M1",
                color_code="010",
                color_desc="00A",
                jo="",
                qty="7430",
                ppo_no="PPO001",
                flow="CM",
                combo_name="00A",
                aggregate_key="S26V02155|COI|CM|M1|010|00A|",
            ),
        ]
        service = GetYYService(
            cm_client=FakeCMQaClient(
                {
                    "PPO001": [
                        CmQaAggregateRow("PPO001", "00A", "B", "1858"),
                        CmQaAggregateRow("PPO001", "00A", "M1", "840.3"),
                    ]
                }
            )
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / MASTER_FILE_NAME
            write_workbook(file_path, records)
            updated = read_records_from_workbook(service.update_ppo_qty_from_workbook(file_path))

        self.assertEqual([(row.fabric_part, row.ppo_qty) for row in updated], [("B", "1858"), ("M1", "840.3")])

    def test_update_cm_from_workbook_only_accepts_cm_flow(self) -> None:
        cm_records = [
            make_record(
                go="S26V02155",
                go_key="S26V02155",
                gmt_color="00A",
                fabric_part="B",
                color_code="010",
                color_desc="00A",
                qty="7430",
                ppo_no="PPO001",
                flow="CM",
                combo_name="00A",
                aggregate_key="S26V02155|COI|CM|B|010|00A|",
            )
        ]
        service = GetYYService(
            cm_client=FakeCMQaClient({"PPO001": [CmQaAggregateRow("PPO001", "00A", "B", "1858")]})
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / MASTER_FILE_NAME
            write_workbook(file_path, cm_records)
            updated = read_records_from_workbook(service.update_cm_from_workbook(file_path))

        self.assertEqual(updated[0].ppo_qty, "1858")

    def test_update_cm_from_workbook_rejects_non_cm_workbook(self) -> None:
        records = [
            make_record(
                go="S25V11174",
                go_key="S25V11174",
                yy_req_no="YF2600107A",
                marker_yy="1.6612",
                ppo_yy="3.3715",
                gmt_color="LIGHT BLUE",
                fabric_part="MAIN BODY1",
                color_code="17",
                color_desc="LIGHT BLUE",
                jo="25K11192ES01",
                qty="218",
                ppo_no="PWGF26SE001192A",
                flow="KNIT",
                combo_name="4SP078 4 GERARD 2 C/17X1-2",
                aggregate_key="S25V11174|COI|KNIT|MAIN BODY1|17|4SP078 4 GERARD 2 C/17X1-2|",
            )
        ]
        service = GetYYService()
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / MASTER_FILE_NAME
            write_workbook(file_path, records)
            with self.assertRaisesRegex(ValidationError, "Workbook nay khong phai CM flow."):
                service.update_cm_from_workbook(file_path)

    def test_update_yy_req_no_from_workbook_refreshes_manual_knit_yy(self) -> None:
        FakeMixedYPDClient.instances = []
        records = [
            make_record(
                go="S26V00420",
                go_key="S26V00420",
                gmt_color="Off",
                fabric_part="MAIN BODY1",
                color_code="01(01)",
                color_desc="OFF",
                jo="26V00420GB01",
                qty="100",
                flow="KNIT",
                combo_name="C2106100_0006",
                block_index=1,
                section_order=1,
                aggregate_key="S26V00420|KNIT|MAIN BODY1|01(01)|C2106100_0006",
            )
        ]
        service = GetYYService()
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / MASTER_FILE_NAME
            write_workbook(file_path, records)
            workbook = load_workbook(file_path)
            workbook["COI"]["B2"] = "YF2600407B(2)"
            workbook.save(file_path)
            with patch("backend.precoi.services.YPDClient", FakeMixedYPDClient):
                updated_path = service.update_yy_req_no_from_workbook(file_path, "user", "secret")
            updated = read_records_from_workbook(updated_path)

        self.assertEqual(updated[0].yy_req_no, "YF2600407B(2)")
        self.assertEqual(updated[0].marker_yy, "0.6379")
        self.assertEqual(updated[0].ppo_yy, "")
        self.assertIn("YF2600407B(2)", FakeMixedYPDClient.instances[0].calls)

    def test_update_yy_req_no_from_workbook_resolves_woven_version_for_manual_yy(self) -> None:
        FakeWovenVersionYPDClient.instances = []
        records = [
            make_record(
                go="S26V00420",
                go_key="S26V00420",
                gmt_color="Off",
                fabric_part="TRIM FAB1",
                color_code="01(01)",
                color_desc="OFF",
                jo="26V00420GB01",
                qty="100",
                flow="WOVEN",
                combo_name="Woven Combo 1",
                aggregate_key="S26V00420|WOVEN|TRIM FAB1|01(01)|WOVEN COMBO 1",
            )
        ]
        service = GetYYService()
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / MASTER_FILE_NAME
            write_workbook(file_path, records)
            workbook = load_workbook(file_path)
            workbook["COI"]["B2"] = "YF2600407A"
            workbook.save(file_path)
            with patch("backend.precoi.services.YPDClient", FakeWovenVersionYPDClient):
                updated_path = service.update_yy_req_no_from_workbook(file_path, "user", "secret")
            updated = read_records_from_workbook(updated_path)

        self.assertEqual(updated[0].yy_req_no, "YF2600407A")
        self.assertEqual(updated[0].marker_yy, "0.402")
        self.assertIn("YF2600407A", FakeWovenVersionYPDClient.instances[0].calls)
        self.assertIn("YF2600407A(3)", FakeWovenVersionYPDClient.instances[0].calls)

    def test_create_output_keeps_ppo_blank_even_if_go_contains_ppo_mapping(self) -> None:
        class MappingGOClient:
            def fetch_go_report(self, go_no: str) -> GOReportData:
                block = GOBomBlock(
                    "KNIT",
                    "Knit Fabric BOM Information",
                    0,
                    [GOBomRow("", "0", "64A", "FK COLLAR1", "Collar Combo 1", "OXBLOOD / ECRU / ECRU", "KNIT", 0, 0)],
                )
                return GOReportData(
                    go_no,
                    [],
                    block.bom_rows,
                    [block],
                    ppo_numbers=["PPO001", "PPO002"],
                    ppo_mapping_by_lot={"7": ["PPO001", "PPO002"]},
                    jo_lot_map={"25V11174GB07": "7"},
                )

        class MappingMESClient:
            def fetch_jo_rows(self, go_no: str) -> list[MESJoRow]:
                return [MESJoRow("25V11174GB07", "2800", "64A", "OXBLOOD / ECRU / ECRU", "0", "0", "")]

        service = GetYYService(
            go_client=MappingGOClient(),
            mes_client=MappingMESClient(),
            webmerge_client=FakeWebmergeClient(
                [WebmergeColorSizeAggregate("S25V11174", "64A", "OXBLOOD / ECRU / ECRU", "S", "273")],
            ),
            ppo_client=FakePpoClient({}),
            ppo_report_client=FakeWovenPpoReportClient(
                {},
                {
                    "PPO002": [
                        KnitPpoBulkColorSizeAggregate(
                            "PPO002",
                            "FK COLLAR1",
                            "64A",
                            "OXBLOOD / ECRU / ECRU",
                            "S",
                            "222",
                        )
                    ]
                },
            ),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            with patch("backend.precoi.services.YPDClient", FakeMixedYPDClient):
                output_path = service.create_output("S25V11174", "user", "secret", tmp_dir)
            updated = read_records_from_workbook(output_path)

        self.assertEqual([(row.ppo_no, row.ppo_qty) for row in updated], [("", "")])

    def test_create_output_moves_collar_rows_to_second_sheet_and_aggregates_sizes(self) -> None:
        FakeMixedYPDClient.instances = []
        service = GetYYService(
            go_client=FakeMixedGOClient(),
            mes_client=FakeMixedMESClient(),
            webmerge_client=FakeWebmergeClient(
                [
                    WebmergeColorSizeAggregate("S26V00420", "01(01)", "OFF", "M", "10"),
                    WebmergeColorSizeAggregate("S26V00420", "01(01)", "OFF", "L", "3"),
                ]
            ),
            ppo_client=FakePpoClient({}),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            with patch("backend.precoi.services.YPDClient", FakeMixedYPDClient):
                output_path = service.create_output("S26V00420", "user", "secret", tmp_dir)
            workbook = load_workbook(output_path)
            main_sheet = workbook["COI"]
            collar_sheet = workbook[COLLAR_SHEET_NAME]

        main_parts = [main_sheet[f"G{row}"].value for row in range(2, main_sheet.max_row + 1) if main_sheet[f"A{row}"].value]
        collar_parts = [collar_sheet[f"C{row}"].value for row in range(2, collar_sheet.max_row + 1) if collar_sheet[f"A{row}"].value]
        collar_sizes_qty = [
            (collar_sheet[f"F{row}"].value, collar_sheet[f"G{row}"].value)
            for row in range(2, collar_sheet.max_row + 1)
            if collar_sheet[f"A{row}"].value
        ]

        self.assertNotIn("FK COLLAR1", main_parts)
        self.assertEqual(collar_parts, ["FK COLLAR1", "FK COLLAR1"])
        self.assertEqual(collar_sizes_qty, [("M", 10), ("L", 3)])

    def test_create_output_moves_any_fk_part_to_second_sheet(self) -> None:
        class SingleFkGOClient:
            def fetch_go_report(self, go_no: str) -> GOReportData:
                block = GOBomBlock(
                    "KNIT",
                    "Knit Fabric BOM Information",
                    0,
                    [
                        GOBomRow("", "0", "BK", "MAIN BODY2", "Body Combo", "BK", "KNIT", 0, 0),
                        GOBomRow("", "0", "BK", "FK BOTTOM1", "Bottom Combo 1", "BK", "KNIT", 0, 1),
                        GOBomRow("", "0", "BK", "FK BOTTOM2", "Bottom Combo 2", "BK", "KNIT", 0, 2),
                    ],
                )
                return GOReportData(go_no, [], block.bom_rows, [block])

        class SingleBkMESClient:
            def fetch_jo_rows(self, go_no: str) -> list[MESJoRow]:
                return [MESJoRow("26V01983KR01", "977", "BK", "BK", "0", "0", "")]

        service = GetYYService(
            go_client=SingleFkGOClient(),
            mes_client=SingleBkMESClient(),
            webmerge_client=FakeWebmergeClient(
                [
                    WebmergeColorSizeAggregate("S26V01983", "BK", "BK", "S", "114"),
                    WebmergeColorSizeAggregate("S26V01983", "BK", "BK", "M", "281"),
                ]
            ),
            ppo_client=FakePpoClient({}),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            with patch("backend.precoi.services.YPDClient", FakeMixedYPDClient):
                output_path = service.create_output("S26V01983", "user", "secret", tmp_dir)
            workbook = load_workbook(output_path)
            main_sheet = workbook["COI"]
            collar_sheet = workbook[COLLAR_SHEET_NAME]

        main_parts = [main_sheet[f"G{row}"].value for row in range(2, main_sheet.max_row + 1) if main_sheet[f"A{row}"].value]
        collar_rows = [
            (collar_sheet[f"C{row}"].value, collar_sheet[f"F{row}"].value, collar_sheet[f"G{row}"].value)
            for row in range(2, collar_sheet.max_row + 1)
            if collar_sheet[f"A{row}"].value
        ]
        self.assertEqual(main_parts, ["MAIN BODY2"])
        self.assertEqual(
            collar_rows,
            [
                ("FK BOTTOM1", "S", 114),
                ("FK BOTTOM1", "M", 281),
                ("FK BOTTOM2", "S", 114),
                ("FK BOTTOM2", "M", 281),
            ],
        )

    def test_create_collar_sheet_keeps_garment_qty_per_part_without_double_counting(self) -> None:
        class SingleCollarCuffGOClient:
            def fetch_go_report(self, go_no: str) -> GOReportData:
                block = GOBomBlock(
                    "KNIT",
                    "Knit Fabric BOM Information",
                    0,
                    [
                        GOBomRow("", "0", "64A", "FK COLLAR1", "Collar Combo 1", "OXBLOOD / ECRU / ECRU", "KNIT", 0, 0),
                        GOBomRow("", "0", "64A", "FK CUFF1", "Cuff Combo 1", "OXBLOOD / ECRU / ECRU", "KNIT", 0, 1),
                    ],
                )
                return GOReportData(go_no, [], block.bom_rows, [block])

        class SingleMESClient:
            def fetch_jo_rows(self, go_no: str) -> list[MESJoRow]:
                return [MESJoRow("25V11174GB01", "2800", "64A", "OXBLOOD / ECRU / ECRU", "0", "0", "")]

        service = GetYYService(
            go_client=SingleCollarCuffGOClient(),
            mes_client=SingleMESClient(),
            webmerge_client=FakeWebmergeClient(
                [WebmergeColorSizeAggregate("S25V11174", "64A", "OXBLOOD / ECRU / ECRU", "S", "273")]
            ),
            ppo_client=FakePpoClient({}),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            with patch("backend.precoi.services.YPDClient", FakeMixedYPDClient):
                output_path = service.create_output("S25V11174", "user", "secret", tmp_dir)
            workbook = load_workbook(output_path)
            collar_sheet = workbook[COLLAR_SHEET_NAME]

        rows = [
            (collar_sheet[f"C{row}"].value, collar_sheet[f"D{row}"].value, collar_sheet[f"F{row}"].value, collar_sheet[f"G{row}"].value)
            for row in range(2, collar_sheet.max_row + 1)
            if collar_sheet[f"A{row}"].value
        ]
        self.assertEqual(rows, [("FK COLLAR1", "64A", "S", 273), ("FK CUFF1", "64A", "S", 273)])

    def test_update_collar_sheet_uses_size_level_bulk_report_without_changing_qty(self) -> None:
        records = [
            make_record(
                go="S25V11174",
                go_key="S25V11174",
                gmt_color="BLACK / ECRU / LAUREL WREATH GREEN",
                fabric_part="FK COLLAR1",
                color_code="W54",
                color_desc="BLACK / ECRU / LAUREL WREATH GREEN",
                qty="273",
                ppo_no="PKEK25VE0011174A",
                ppo_qty="",
                flow="KNIT",
                combo_name="Collar Combo 1",
                aggregate_key="S25V11174|COI COLLARï¼CUFF|KNIT|FK COLLAR1|W54|COLLAR COMBO 1|S",
                sheet_kind=COLLAR_SHEET_NAME,
                size="S",
            ),
            make_record(
                row_index=1,
                go="S25V11174",
                go_key="S25V11174",
                gmt_color="BLACK / ECRU / LAUREL WREATH GREEN",
                fabric_part="FK CUFF1",
                color_code="W54",
                color_desc="BLACK / ECRU / LAUREL WREATH GREEN",
                qty="273",
                ppo_no="PKEK25VE0011174A",
                ppo_qty="",
                flow="KNIT",
                combo_name="Cuff Combo 1",
                aggregate_key="S25V11174|COI COLLARï¼CUFF|KNIT|FK CUFF1|W54|CUFF COMBO 1|S",
                sheet_kind=COLLAR_SHEET_NAME,
                size="S",
            ),
        ]
        service = GetYYService(
            ppo_client=FakePpoClient({}),
            ppo_report_client=FakeWovenPpoReportClient(
                {},
                {
                    "PKEK25VE0011174A": [
                        KnitPpoBulkColorSizeAggregate(
                            "PKEK25VE0011174A",
                            "FK COLLAR1",
                            "W54",
                            "BLACK / ECRU / LAUREL WREATH GREEN",
                            "S",
                            "284",
                        ),
                        KnitPpoBulkColorSizeAggregate(
                            "PKEK25VE0011174A",
                            "FK CUFF1",
                            "W54",
                            "BLACK / ECRU / LAUREL WREATH GREEN",
                            "S",
                            "568",
                        ),
                    ]
                },
            ),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / MASTER_FILE_NAME
            write_workbook(file_path, records)
            updated = read_records_from_workbook(service.update_ppo_qty_from_workbook(file_path))

        self.assertEqual(
            [(row.fabric_part, row.qty, row.ppo_qty) for row in updated],
            [("FK COLLAR1", "273", "284"), ("FK CUFF1", "273", "568")],
        )

    def test_update_collar_sheet_splits_comma_separated_ppos_into_multiple_rows(self) -> None:
        records = [
            make_record(
                go="S25V11174",
                go_key="S25V11174",
                gmt_color="BLACK / ECRU / LAUREL WREATH GREEN",
                fabric_part="FK COLLAR1",
                color_code="W54",
                color_desc="BLACK / ECRU / LAUREL WREATH GREEN",
                qty="273",
                ppo_no="PPO001, PPO002, PPO001",
                ppo_qty="",
                flow="KNIT",
                combo_name="Collar Combo 1",
                aggregate_key="S25V11174|COI COLLARÃ¯Â¼ÂCUFF|KNIT|FK COLLAR1|W54|COLLAR COMBO 1|S",
                sheet_kind=COLLAR_SHEET_NAME,
                size="S",
            ),
        ]
        service = GetYYService(
            ppo_client=FakePpoClient({}),
            ppo_report_client=FakeWovenPpoReportClient(
                {},
                {
                    "PPO001": [
                        KnitPpoBulkColorSizeAggregate(
                            "PPO001",
                            "FK COLLAR1",
                            "W54",
                            "BLACK / ECRU / LAUREL WREATH GREEN",
                            "S",
                            "111",
                        ),
                    ],
                    "PPO002": [
                        KnitPpoBulkColorSizeAggregate(
                            "PPO002",
                            "FK COLLAR1",
                            "W54",
                            "BLACK / ECRU / LAUREL WREATH GREEN",
                            "S",
                            "222",
                        ),
                    ],
                },
            ),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / MASTER_FILE_NAME
            write_workbook(file_path, records)
            updated = read_records_from_workbook(service.update_ppo_qty_from_workbook(file_path))

        self.assertEqual(
            [(row.ppo_no, row.fabric_part, row.color_code, row.size, row.qty, row.ppo_qty) for row in updated],
            [
                ("PPO001", "FK COLLAR1", "W54", "S", "273", "111"),
                ("PPO002", "FK COLLAR1", "W54", "S", "273", "222"),
            ],
        )

    def test_update_main_coi_sheet_splits_comma_separated_ppos_into_multiple_rows(self) -> None:
        records = [
            make_record(
                go="S26V00420",
                go_key="S26V00420",
                yy_req_no="YF2600407B(2)",
                marker_yy="0.6379",
                ppo_yy="0.6463",
                gmt_color="Off",
                fabric_part="MAIN BODY1",
                color_code="01(01)",
                color_desc="OFF",
                jo="26V00420GB01",
                qty="100",
                ppo_no="PPO001, PPO002, PPO003",
                flow="KNIT",
                combo_name="C2106100_0006",
                aggregate_key="S26V00420|KNIT|MAIN BODY1|01(01)|C2106100_0006",
            ),
        ]
        service = GetYYService(
            ppo_client=FakePpoClient(
                {},
                {
                    "PPO001": [KnitPpoAggregateRow("PPO001", "B", "MAIN BODY1", "C2106100_0006", "Off", "111")],
                    "PPO003": [KnitPpoAggregateRow("PPO003", "B", "MAIN BODY1", "C2106100_0006", "Off", "333")],
                },
            ),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / MASTER_FILE_NAME
            write_workbook(file_path, records)
            updated = read_records_from_workbook(service.update_ppo_qty_from_workbook(file_path))

        self.assertEqual(
            [(row.ppo_no, row.fabric_part, row.color_code, row.jo, row.qty, row.ppo_qty) for row in updated],
            [
                ("PPO001", "MAIN BODY1", "01(01)", "26V00420GB01", "100", "111"),
                ("PPO003", "MAIN BODY1", "01(01)", "26V00420GB01", "100", "333"),
            ],
        )

    def test_update_knit_row_falls_back_to_woven_part_qty_when_knit_source_is_empty(self) -> None:
        records = [
            make_record(
                go="S26V02747",
                go_key="S26V02747",
                yy_req_no="YF2508858A(4)",
                marker_yy="0.8",
                ppo_yy="0.7824",
                gmt_color="IVD(2)+KAD(2)",
                fabric_part="MAIN BODY1",
                color_code="IVD(2)+KAD(2)",
                color_desc="IVD(2)+KAD(2)",
                jo="26V02747SB01",
                qty="100",
                ppo_no="POUT26SB002747A",
                flow="KNIT",
                combo_name="IVD(2)-TS81/82/KAD(2)-TS7C",
                aggregate_key="S26V02747|KNIT|MAIN BODY1|IVD(2)+KAD(2)|IVD(2)-TS81/82/KAD(2)-TS7C",
            ),
            make_record(
                row_index=1,
                go="S26V02747",
                go_key="S26V02747",
                yy_req_no="YF2508858A(4)",
                marker_yy="0.1",
                ppo_yy="0.0172",
                gmt_color="IVD(2)+KAD(2)",
                fabric_part="TRIM FAB1",
                color_code="IVD(2)+KAD(2)",
                color_desc="IVD(2)+KAD(2)",
                jo="26V02747SB01",
                qty="100",
                ppo_no="POUT26SB002747A",
                flow="KNIT",
                combo_name="IVD(2)-TS81/82",
                aggregate_key="S26V02747|KNIT|TRIM FAB1|IVD(2)+KAD(2)|IVD(2)-TS81/82",
            ),
        ]
        service = GetYYService(
            ppo_client=FakePpoClient({}, knit_part_responses={"POUT26SB002747A": []}, woven_responses={"POUT26SB002747A": [PpoComboAggregateRow("POUT26SB002747A", "BD", "IVD", "30"), PpoComboAggregateRow("POUT26SB002747A", "BD", "IVD-SURCHARGE", "1")]}),
            ppo_report_client=FakeWovenPpoReportClient(
                {},
                {},
                {
                    "POUT26SB002747A": [
                        WovenPpoPartQtyRow("POUT26SB002747A", "MAIN BODY1", "IVD", "30"),
                        WovenPpoPartQtyRow("POUT26SB002747A", "MAIN BODY1", "IVD-SURCHARGE", "1"),
                    ]
                },
            ),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / MASTER_FILE_NAME
            write_workbook(file_path, records)
            updated = read_records_from_workbook(service.update_ppo_qty_from_workbook(file_path))

        self.assertEqual([(row.fabric_part, row.ppo_qty) for row in updated], [("MAIN BODY1", "30"), ("TRIM FAB1", "")])

    def test_update_knit_row_uses_unique_combo_name_fallback_when_part_mapping_differs(self) -> None:
        records = [
            make_record(
                go="S26V02747",
                go_key="S26V02747",
                yy_req_no="YF2508858A(4)",
                marker_yy="0.0172",
                ppo_yy="0.0172",
                gmt_color="IVD(2)+KAD(2)",
                fabric_part="TRIM FAB1",
                color_code="IVD(2)+KAD(2)",
                color_desc="IVD(2)+KAD(2)",
                jo="26V02747KR01",
                qty="1200",
                ppo_no="PKGK26SB0002747A",
                flow="KNIT",
                combo_name="IVD(2)-TS81/82",
                aggregate_key="S26V02747|KNIT|TRIM FAB1|IVD(2)+KAD(2)|IVD(2)-TS81/82",
            ),
        ]
        service = GetYYService(
            ppo_client=FakePpoClient(
                {"PKGK26SB0002747A": [PpoColorAggregateRow("PKGK26SB0002747A", "B", "IVD(2)+KAD(2)", "1000")]},
                {
                    "PKGK26SB0002747A": [
                        KnitPpoAggregateRow("PKGK26SB0002747A", "O", "FK COLLAR1", "IVD(2)+KAD(2)", "IVD(2)-TS81/82", "1235"),
                        KnitPpoAggregateRow("PKGK26SB0002747A", "B", "MAIN BODY1", "IVD(2)+KAD(2)", "IVD(2)-TS81/82/KAD(2)-TS7C", "1000"),
                    ]
                },
            ),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / MASTER_FILE_NAME
            write_workbook(file_path, records)
            updated = read_records_from_workbook(service.update_ppo_qty_from_workbook(file_path))

        self.assertEqual(updated[0].ppo_qty, "1235")

    def test_update_knit_row_prefers_knit_when_woven_fallback_matches_same_qty(self) -> None:
        records = [
            make_record(
                go="S26V00420",
                go_key="S26V00420",
                yy_req_no="YF2600407B(2)",
                marker_yy="0.6379",
                ppo_yy="0.6463",
                gmt_color="Off",
                fabric_part="MAIN BODY1",
                color_code="01(01)",
                color_desc="OFF",
                jo="26V00420GB01",
                qty="100",
                ppo_no="PPO001",
                flow="KNIT",
                combo_name="C2106100_0006",
                aggregate_key="S26V00420|KNIT|MAIN BODY1|01(01)|C2106100_0006",
            ),
        ]
        service = GetYYService(
            ppo_client=FakePpoClient(
                {},
                {"PPO001": [KnitPpoAggregateRow("PPO001", "B", "MAIN BODY1", "C2106100_0006", "Off", "111")]},
                {"PPO001": [PpoComboAggregateRow("PPO001", "BD", "C2106100_0006", "111")]},
            ),
            ppo_report_client=FakeWovenPpoReportClient(
                {},
                {},
                {"PPO001": [WovenPpoPartQtyRow("PPO001", "MAIN BODY1", "C2106100_0006", "111")]},
            ),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / MASTER_FILE_NAME
            write_workbook(file_path, records)
            updated = read_records_from_workbook(service.update_ppo_qty_from_workbook(file_path))

        self.assertEqual([(row.ppo_no, row.ppo_qty) for row in updated], [("PPO001", "111")])

    def test_update_mixed_go_knit_row_does_not_use_woven_fallback(self) -> None:
        records = [
            make_record(
                go="S26V00420",
                go_key="S26V00420",
                yy_req_no="YF2600407B(2)",
                marker_yy="0.6379",
                ppo_yy="0.6463",
                gmt_color="Off",
                fabric_part="MAIN BODY1",
                color_code="01(01)",
                color_desc="OFF",
                jo="26V00420GB01",
                qty="100",
                ppo_no="PPO001",
                flow="KNIT",
                combo_name="C2106100_0006",
                aggregate_key="S26V00420|KNIT|MAIN BODY1|01(01)|C2106100_0006",
            ),
            make_record(
                row_index=1,
                go="S26V00420",
                go_key="S26V00420",
                yy_req_no="YF2600407A",
                marker_yy="0.4",
                ppo_yy="",
                gmt_color="Off",
                fabric_part="TRIM FAB1",
                color_code="01(01)",
                color_desc="OFF",
                jo="26V00420GB01",
                qty="100",
                ppo_no="PWGF26VB000420A",
                flow="WOVEN",
                combo_name="1602-64221 01(01) Off trim fabric",
                aggregate_key="S26V00420|WOVEN|TRIM FAB1|01(01)|1602-64221 01(01) OFF TRIM FABRIC",
            ),
        ]
        service = GetYYService(
            ppo_client=FakePpoClient(
                {},
                {"PPO001": []},
                {"PPO001": [PpoComboAggregateRow("PPO001", "BD", "C2106100_0006", "111")]},
            ),
            ppo_report_client=FakeWovenPpoReportClient(
                {},
                {},
                {"PPO001": [WovenPpoPartQtyRow("PPO001", "MAIN BODY1", "C2106100_0006", "111")]},
            ),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / MASTER_FILE_NAME
            write_workbook(file_path, records)
            updated = read_records_from_workbook(service.update_ppo_qty_from_workbook(file_path))

        self.assertEqual(updated[0].ppo_qty, "")

    def test_update_multi_ppo_marks_row_when_no_candidate_matches(self) -> None:
        records = [
            make_record(
                go="S25V11174",
                go_key="S25V11174",
                gmt_color="BLACK / ECRU / LAUREL WREATH GREEN",
                fabric_part="FK COLLAR1",
                color_code="W54",
                color_desc="BLACK / ECRU / LAUREL WREATH GREEN",
                qty="273",
                ppo_no="PPO001, PPO002",
                ppo_qty="",
                flow="KNIT",
                combo_name="Collar Combo 1",
                aggregate_key="S25V11174|COI COLLARÃ¯Â¼ÂCUFF|KNIT|FK COLLAR1|W54|COLLAR COMBO 1|S",
                sheet_kind=COLLAR_SHEET_NAME,
                size="S",
            ),
        ]
        service = GetYYService(
            ppo_client=FakePpoClient({}),
            ppo_report_client=FakeWovenPpoReportClient({}, {"PPO999": []}),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / MASTER_FILE_NAME
            write_workbook(file_path, records)
            updated = read_records_from_workbook(service.update_ppo_qty_from_workbook(file_path))

        self.assertEqual(len(updated), 1)
        self.assertEqual(updated[0].ppo_no, MULTI_PPO_CHECK_MESSAGE)
        self.assertEqual(updated[0].ppo_qty, "")

    def test_update_main_sheet_multi_ppo_marks_row_when_no_candidate_matches(self) -> None:
        records = [
            make_record(
                go="S26V00420",
                go_key="S26V00420",
                yy_req_no="YF2600407B(2)",
                marker_yy="0.6379",
                ppo_yy="0.6463",
                gmt_color="Off",
                fabric_part="MAIN BODY1",
                color_code="01(01)",
                color_desc="OFF",
                jo="26V00420GB01",
                qty="100",
                ppo_no="PPO001, PPO002",
                flow="KNIT",
                combo_name="C2106100_0006",
                aggregate_key="S26V00420|KNIT|MAIN BODY1|01(01)|C2106100_0006",
            ),
        ]
        service = GetYYService(ppo_client=FakePpoClient({}))
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / MASTER_FILE_NAME
            write_workbook(file_path, records)
            updated = read_records_from_workbook(service.update_ppo_qty_from_workbook(file_path))

        self.assertEqual(len(updated), 1)
        self.assertEqual(updated[0].ppo_no, MULTI_PPO_CHECK_MESSAGE)
        self.assertEqual(updated[0].ppo_qty, "")

    def test_update_woven_uses_fabric_lots_before_db_qty(self) -> None:
        records = [
            make_record(
                go="S26V03162",
                go_key="S26V03162",
                flow="WOVEN",
                sheet_kind="COI",
                yy_req_no="YF2603810A",
                fabric_part="TRIM FAB1",
                combo_name="1602-65232 trim fab",
                color_code="01(02)",
                color_desc="Ivory",
                ppo_no="PWGF26VB003162A",
            )
        ]
        service = GetYYService(
            ppo_client=FakePpoClient(
                {},
                woven_responses={
                    "PWGF26VB003162A": [
                        PpoComboAggregateRow("PWGF26VB003162A", "M1", "1602-65232 trim fab", "999")
                    ]
                },
            ),
            ppo_report_client=FakeWovenPpoReportClient(
                {},
                part_qty_responses={
                    "PWGF26VB003162A": [
                        WovenPpoPartQtyRow("PWGF26VB003162A", "TRIM FAB1", "1602-65232 trim fab", "145.6")
                    ]
                },
            ),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / "woven.xlsx"
            write_workbook(file_path, records)
            updated = read_records_from_workbook(service.update_ppo_qty_from_workbook(file_path))

        self.assertEqual(updated[0].ppo_qty, "145.6")

    def test_create_collar_sheet_keeps_blank_qty_when_webmerge_has_no_match(self) -> None:
        class SingleCollarGOClient:
            def fetch_go_report(self, go_no: str) -> GOReportData:
                block = GOBomBlock(
                    "KNIT",
                    "Knit Fabric BOM Information",
                    0,
                    [GOBomRow("", "0", "01(01)", "FK COLLAR1", "Collar Combo 1", "Off", "KNIT", 0, 0)],
                )
                return GOReportData(go_no, [], block.bom_rows, [block])

        class SingleMESClient:
            def fetch_jo_rows(self, go_no: str) -> list[MESJoRow]:
                return [MESJoRow("25V11174GB01", "100", "99Z", "OTHER", "0", "0", "")]

        service = GetYYService(
            go_client=SingleCollarGOClient(),
            mes_client=SingleMESClient(),
            webmerge_client=FakeWebmergeClient([]),
            ppo_client=FakePpoClient({}),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            with patch("backend.precoi.services.YPDClient", FakeMixedYPDClient):
                output_path = service.create_output("S25V11174", "user", "secret", tmp_dir)
            workbook = load_workbook(output_path)
            collar_sheet = workbook[COLLAR_SHEET_NAME]

        self.assertEqual(collar_sheet["D2"].value, "01(01)")
        self.assertEqual(collar_sheet["F2"].value, None)
        self.assertEqual(collar_sheet["G2"].value, None)

    def test_update_woven_sums_same_combo_across_type_codes(self) -> None:
        record = make_record(yy_req_no="YF2600407A", marker_yy="0.402", gmt_color="Off", fabric_part="TRIM FAB1", color_code="01(01)", color_desc="OFF", jo="26V00420GB01", qty="100", ppo_no="PWGF26VB000420A", flow="WOVEN", combo_name="1602-64221 01(01) Off trim fabric", aggregate_key="S26V00420|WOVEN|TRIM FAB1|01(01)|1602-64221 01(01) OFF TRIM FABRIC")
        service = GetYYService(
            ppo_client=FakePpoClient({}, woven_responses={"PWGF26VB000420A": [PpoComboAggregateRow("PWGF26VB000420A", "M1", "1602-64221 01(01) Off trim fabric", "796"), PpoComboAggregateRow("PWGF26VB000420A", "M2", "1602-64221 01(01) Off trim fabric", "5.6")]}),
            ppo_report_client=FakeWovenPpoReportClient({"PWGF26VB000420A": [WovenPpoYYRow("PWGF26VB000420A", "YF2600407A", "TRIM FAB1", "1602-64221 01(01) Off trim fabric", "0.4077")]}),
        )
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / MASTER_FILE_NAME
            write_workbook(file_path, [record])
            updated = read_records_from_workbook(service.update_ppo_qty_from_workbook(file_path))
        self.assertEqual((updated[0].ppo_yy, updated[0].ppo_qty), ("0.4077", "801.6"))


if __name__ == "__main__":
    unittest.main()
