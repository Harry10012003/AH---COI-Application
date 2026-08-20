from __future__ import annotations

from datetime import datetime
from io import BytesIO
import math
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_EDITABLE_KEYS = {"AH Allocate Q'ty (yds)", "User Remark"}
_PERCENT_KEYS = {"Allocate %"}
_INTEGER_KEYS = {"- %", "+%"}
_DECIMAL_KEYS = {
    "Qty (pcs)",
    "Net YY",
    "PPO YY",
    "Marker YY",
    "Required Q'ty (Yds)",
    "Rcv Q'ty (PPO)",
    "On The Way Q'ty (Yds)",
    "Allocate Q'ty (Yds)",
    "Shortage Q'ty (Yds)",
    "AH Allocate Q'ty (yds)",
    "PPO Order Total (Yds)",
}


def _safe_token(value: object, fallback: str = "COI") -> str:
    text = str(value or "").strip().upper()
    text = re.sub(r"[^A-Z0-9_.-]+", "_", text).strip("._-")
    return text or fallback


def _to_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return float(value)
    text = str(value).replace(",", "").replace("%", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _display_value(key: str, value: object) -> object:
    numeric = _to_float(value)
    if key in _PERCENT_KEYS and numeric is not None:
        pct = numeric * 100 if abs(numeric) <= 10 else numeric
        return f"{pct:.2f}".rstrip("0").rstrip(".") + "%"
    if key in _INTEGER_KEYS and numeric is not None:
        return int(round(numeric))
    if key in _DECIMAL_KEYS and numeric is not None:
        return round(numeric, 6)
    return "" if value is None else value


def _number_format(key: str) -> str:
    if key in _PERCENT_KEYS:
        return "0.##%"
    if key in _INTEGER_KEYS:
        return "0"
    if key in _DECIMAL_KEYS:
        return '#,##0.###'
    return "General"


def _source_fill(source: object) -> PatternFill:
    token = str(source or "").strip().lower().replace("/", "-").replace(" ", "-")
    if token in {"go", "sql", "go-sql"}:
        return PatternFill("solid", fgColor="EEF6EF")
    if token == "mes":
        return PatternFill("solid", fgColor="FDE7CC")
    if token == "calculated":
        return PatternFill("solid", fgColor="E8F0FE")
    if token in {"ui", "system-ui"}:
        return PatternFill("solid", fgColor="FFF4CC")
    return PatternFill("solid", fgColor="F3F7F0")


def _column_width(ui_width: object) -> float:
    try:
        width = float(ui_width or 120)
    except (TypeError, ValueError):
        width = 120
    return max(8.0, min(42.0, width / 7.6))


def _visible_columns(payload: dict[str, Any]) -> list[dict[str, Any]]:
    columns = payload.get("columns") if isinstance(payload.get("columns"), list) else []
    return [dict(item) for item in columns if isinstance(item, dict) and str(item.get("key") or "").strip()]


def _visible_rows(payload: dict[str, Any]) -> list[dict[str, Any]]:
    rows = payload.get("rows") if isinstance(payload.get("rows"), list) else []
    return [dict(item) for item in rows if isinstance(item, dict)]


def build_coi_ui_export_workbook(payload: dict[str, Any]) -> tuple[bytes, str, str]:
    columns = _visible_columns(payload)
    rows = _visible_rows(payload)
    go_no = _safe_token(payload.get("go") or payload.get("selected_go") or "GO", "GO")
    sheet_info = payload.get("sheet") if isinstance(payload.get("sheet"), dict) else {}
    summary = payload.get("summary") if isinstance(payload.get("summary"), dict) else {}
    filters = payload.get("filters") if isinstance(payload.get("filters"), dict) else {}

    wb = Workbook()
    ws = wb.active
    ws.title = "FORMAT COI REQUEST"
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin", color="E3E8DE")
    strong = Side(style="thin", color="AEB9AA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_border = Border(left=thin, right=thin, top=thin, bottom=strong)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ribbon_fill = PatternFill("solid", fgColor="103A2B")
    panel_fill = PatternFill("solid", fgColor="F8F6EF")
    letter_fill = PatternFill("solid", fgColor="E7ECE3")
    corner_fill = PatternFill("solid", fgColor="F1F4ED")
    filter_fill = PatternFill("solid", fgColor="F7FAF4")
    editable_fill = PatternFill("solid", fgColor="FFF9DF")
    warning_fill = PatternFill("solid", fgColor="FFF5BF")
    warning_edit_fill = PatternFill("solid", fgColor="FFE9A8")

    total_cols = max(len(columns) + 1, 8)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=f"FORMAT COI REQUEST - {go_no}")
    title_cell.fill = ribbon_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = left
    ws.row_dimensions[1].height = 24

    meta_text = " | ".join(
        item
        for item in [
            str(payload.get("brand") or "").strip(),
            str(payload.get("factory_code") or "").strip(),
            str(sheet_info.get("name") or "FORMAT COI REQUEST").strip(),
            f"Rows: {len(rows)}",
            f"Export: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        ]
        if item
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.cell(row=2, column=1, value=meta_text).fill = panel_fill

    metrics = [
        ("ROWS", len(rows)),
        ("REQUIRED", summary.get("total_required_qty", "")),
        ("RECEIVED TOTAL", summary.get("total_received_qty", "")),
        ("ON THE WAY", summary.get("total_on_way_qty", "")),
        ("ALLOCATE TOTAL", summary.get("total_effective_allocate_qty", "")),
        ("SHORTAGE", summary.get("total_shortage_qty", "")),
        ("COVERAGE", summary.get("coverage_pct", "")),
        ("MANUAL ROWS", summary.get("manual_override_rows", "")),
    ]
    for index, (label, value) in enumerate(metrics, start=1):
        cell = ws.cell(row=4, column=index, value=label)
        cell.fill = panel_fill
        cell.font = Font(color="536056", bold=True, size=9)
        cell.border = border
        value_cell = ws.cell(row=5, column=index, value=_display_value("Allocate %" if label == "COVERAGE" else "", value))
        value_cell.fill = panel_fill
        value_cell.font = Font(color="1B241F", bold=True, size=11)
        value_cell.border = border

    header_letter_row = 7
    header_label_row = 8
    filter_row = 9
    first_data_row = 10

    ws.cell(row=header_letter_row, column=1, value="").fill = letter_fill
    ws.cell(row=header_label_row, column=1, value=8).fill = corner_fill
    ws.cell(row=filter_row, column=1, value="").fill = filter_fill
    for row_idx in (header_letter_row, header_label_row, filter_row):
        ws.cell(row=row_idx, column=1).border = header_border
        ws.cell(row=row_idx, column=1).alignment = center
    ws.column_dimensions["A"].width = 8

    for idx, col in enumerate(columns, start=2):
        key = str(col.get("key") or "").strip()
        letter = str(col.get("letter") or "").strip()
        label = str(col.get("label") or key).strip()
        column_letter = get_column_letter(idx)
        ws.column_dimensions[column_letter].width = _column_width(col.get("width"))

        letter_cell = ws.cell(row=header_letter_row, column=idx, value=letter)
        letter_cell.fill = letter_fill
        letter_cell.font = Font(color="536056", bold=True, size=9)
        letter_cell.alignment = center
        letter_cell.border = border

        label_cell = ws.cell(row=header_label_row, column=idx, value=label)
        label_cell.fill = _source_fill(col.get("source"))
        label_cell.font = Font(color="1B241F", bold=True)
        label_cell.alignment = left
        label_cell.border = header_border

        filter_cell = ws.cell(row=filter_row, column=idx, value=str(filters.get(key) or ""))
        filter_cell.fill = filter_fill
        filter_cell.font = Font(color="536056", size=9)
        filter_cell.alignment = left
        filter_cell.border = header_border

    for row_offset, row in enumerate(rows):
        excel_row = first_data_row + row_offset
        allocate_pct = _to_float(row.get("Allocate %"))
        warning = allocate_pct is not None and allocate_pct <= 0.99
        row_head = ws.cell(row=excel_row, column=1, value=9 + row_offset)
        row_head.fill = warning_fill if warning else corner_fill
        row_head.font = Font(color="536056", bold=True, size=9)
        row_head.alignment = center
        row_head.border = border

        for idx, col in enumerate(columns, start=2):
            key = str(col.get("key") or "").strip()
            editable = bool(col.get("editable")) or key in _EDITABLE_KEYS
            value = _display_value(key, row.get(key, ""))
            cell = ws.cell(row=excel_row, column=idx, value=value)
            cell.fill = warning_edit_fill if warning and editable else (warning_fill if warning else (editable_fill if editable else PatternFill("solid", fgColor="FFFFFF")))
            cell.alignment = left
            cell.border = border
            cell.number_format = _number_format(key)

    ws.freeze_panes = f"B{first_data_row}"
    ws.auto_filter.ref = f"B{header_label_row}:{get_column_letter(len(columns) + 1)}{max(first_data_row, first_data_row + len(rows) - 1)}"

    output = BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)
    filename = f"{go_no}-COI-UI-Export-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    return output.getvalue(), filename, _XLSX_MIME
