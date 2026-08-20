from __future__ import annotations

from datetime import date, datetime, time
import re
from pathlib import Path
import shutil
import subprocess
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from backend.sources import CACHE_DIR, COI_SAMPLE_XLSX

_MAIN_NS = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
_REL_NS = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}

_AUDIT_CACHE: dict[str, dict] = {}


def _error(message: str, **extra) -> dict:
    return {"ok": False, "error": message, **extra}


def _safe_int(value: object, default: int, minimum: int, maximum: int) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return max(minimum, min(parsed, maximum))


def _to_jsonable(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (list, tuple)):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, dict):
        return {str(k): _to_jsonable(v) for k, v in value.items()}
    return str(value)


def _coerce_cell_input(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, (int, float, bool, datetime, date, time)):
        return value

    text = str(value)
    if not text.strip():
        return None
    if text.startswith("="):
        return text

    number_like = text.replace(",", "").strip()
    if re.fullmatch(r"[+-]?\d+", number_like):
        try:
            return int(number_like)
        except ValueError:
            pass
    if re.fullmatch(r"[+-]?\d+\.\d+", number_like):
        try:
            return float(number_like)
        except ValueError:
            pass
    if number_like.lower() == "true":
        return True
    if number_like.lower() == "false":
        return False
    return text


def _normalize_zip_path(path: str) -> str:
    normalized = str(path or "").replace("\\", "/").lstrip("/")
    return normalized if normalized.startswith("xl/") else f"xl/{normalized}"


def _resolve_workbook_path(workbook_path: str | None) -> Path:
    if workbook_path:
        path = Path(str(workbook_path)).expanduser().resolve()
    else:
        source = Path(COI_SAMPLE_XLSX).resolve()
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        workspace_copy = (CACHE_DIR / "coi_ui_workspace.xlsx").resolve()
        if not workspace_copy.exists() and source.exists():
            shutil.copy2(source, workspace_copy)
        path = workspace_copy if workspace_copy.exists() else source

    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ValueError(f"Unsupported workbook format: {path.suffix}")
    return path


def _sheet_target_map(zipf: zipfile.ZipFile) -> dict[str, str]:
    try:
        workbook_xml = ET.fromstring(zipf.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(zipf.read("xl/_rels/workbook.xml.rels"))
    except KeyError:
        return {}

    rel_map: dict[str, str] = {}
    for rel in rels_xml.findall("r:Relationship", _REL_NS):
        rel_id = rel.attrib.get("Id", "")
        target = rel.attrib.get("Target", "")
        if rel_id and target:
            rel_map[rel_id] = _normalize_zip_path(target)

    output = {}
    rid_key = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    for sheet in workbook_xml.findall("x:sheets/x:sheet", _MAIN_NS):
        name = sheet.attrib.get("name", "")
        rel_id = sheet.attrib.get(rid_key, "")
        target = rel_map.get(rel_id, "")
        if name and target:
            output[name] = target
    return output


def _parse_connections(zipf: zipfile.ZipFile) -> list[dict]:
    if "xl/connections.xml" not in zipf.namelist():
        return []
    try:
        root = ET.fromstring(zipf.read("xl/connections.xml"))
    except ET.ParseError:
        return []

    result = []
    for conn in root.findall("x:connection", _MAIN_NS):
        web_pr = conn.find("x:webPr", _MAIN_NS)
        db_pr = conn.find("x:dbPr", _MAIN_NS)
        parameters = []
        for param in conn.findall("x:parameters/x:parameter", _MAIN_NS):
            parameters.append(
                {
                    "name": param.attrib.get("name", ""),
                    "parameter_type": param.attrib.get("parameterType", ""),
                    "cell": param.attrib.get("cell", ""),
                }
            )
        result.append(
            {
                "id": conn.attrib.get("id", ""),
                "name": conn.attrib.get("name", ""),
                "type": conn.attrib.get("type", ""),
                "description": conn.attrib.get("description", ""),
                "url": web_pr.attrib.get("url", "") if web_pr is not None else "",
                "command": db_pr.attrib.get("command", "") if db_pr is not None else "",
                "parameters": parameters,
            }
        )
    return result


def _parse_query_tables(zipf: zipfile.ZipFile) -> list[dict]:
    names = sorted(
        name
        for name in zipf.namelist()
        if name.startswith("xl/queryTables/queryTable") and name.endswith(".xml")
    )
    result = []
    for name in names:
        try:
            root = ET.fromstring(zipf.read(name))
        except ET.ParseError:
            continue
        result.append(
            {
                "file": name,
                "name": root.attrib.get("name", ""),
                "connection_id": root.attrib.get("connectionId", ""),
            }
        )
    return result


def _scan_sheet_xml_metrics(zipf: zipfile.ZipFile, xml_path: str) -> dict:
    try:
        payload = zipf.read(xml_path)
    except KeyError:
        return {"non_empty_cells": 0, "formula_cells": 0}
    return {
        "non_empty_cells": len(re.findall(rb"<c(?:\s|>)", payload)),
        "formula_cells": len(re.findall(rb"<f(?:\s|>)", payload)),
    }


def get_workbook_overview(workbook_path: str | None = None) -> dict:
    try:
        workbook = _resolve_workbook_path(workbook_path)
    except (FileNotFoundError, ValueError) as exc:
        return _error(str(exc))

    stat = workbook.stat()
    with zipfile.ZipFile(workbook, "r") as zipf:
        names = set(zipf.namelist())
        has_vba = "xl/vbaProject.bin" in names
        query_table_files = sorted(
            name
            for name in names
            if name.startswith("xl/queryTables/queryTable") and name.endswith(".xml")
        )
        has_connections = "xl/connections.xml" in names
        has_data_mashup = any(
            "datamashup" in zipf.read(name).decode("utf-8", errors="ignore").lower()
            for name in names
            if name.startswith("customXml/itemProps") and name.endswith(".xml")
        )
        custom_xml_files = sorted(name for name in names if name.startswith("customXml/"))

    wb = load_workbook(workbook, data_only=False, read_only=True)
    sheets = []
    for ws in wb.worksheets:
        sheets.append(
            {
                "name": ws.title,
                "state": ws.sheet_state,
                "max_row": int(ws.max_row or 0),
                "max_col": int(ws.max_column or 0),
                "dimension": ws.calculate_dimension(),
            }
        )
    wb.close()

    return {
        "ok": True,
        "workbook": {
            "path": str(workbook),
            "filename": workbook.name,
            "size_bytes": stat.st_size,
            "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(sep=" ", timespec="seconds"),
        },
        "sheets": sheets,
        "flags": {
            "has_vba_project": has_vba,
            "has_connections_xml": has_connections,
            "has_power_query_artifacts": bool(query_table_files or has_connections or has_data_mashup),
            "has_data_mashup": has_data_mashup,
        },
        "power_query_files": {
            "query_tables": query_table_files,
            "custom_xml_files": custom_xml_files,
        },
    }


def audit_workbook(workbook_path: str | None = None) -> dict:
    overview = get_workbook_overview(workbook_path)
    if not overview.get("ok"):
        return overview

    workbook = Path(overview["workbook"]["path"])
    stat = workbook.stat()
    cache_key = f"{workbook}|{stat.st_mtime_ns}|{stat.st_size}"
    cached = _AUDIT_CACHE.get(cache_key)
    if cached:
        return cached

    with zipfile.ZipFile(workbook, "r") as zipf:
        sheet_path_by_name = _sheet_target_map(zipf)
        connections = _parse_connections(zipf)
        query_tables = _parse_query_tables(zipf)

        sheets = []
        for sheet in overview.get("sheets", []):
            name = sheet.get("name", "")
            xml_path = sheet_path_by_name.get(name, "")
            metrics = _scan_sheet_xml_metrics(zipf, xml_path) if xml_path else {"non_empty_cells": 0, "formula_cells": 0}
            sheets.append({**sheet, **metrics, "xml_path": xml_path})

        data_mashup_items = []
        for name in sorted(item for item in zipf.namelist() if item.startswith("customXml/itemProps")):
            text = zipf.read(name).decode("utf-8", errors="ignore")
            if "DataMashup" in text:
                data_mashup_items.append(name)

    result = {
        "ok": True,
        "workbook": overview.get("workbook", {}),
        "flags": overview.get("flags", {}),
        "sheets": sheets,
        "power_query": {
            "connections": connections,
            "query_tables": query_tables,
            "data_mashup_item_props": data_mashup_items,
        },
    }
    _AUDIT_CACHE.clear()
    _AUDIT_CACHE[cache_key] = result
    return result


def read_sheet_window(
    *,
    sheet_name: str,
    workbook_path: str | None = None,
    start_row: int = 1,
    row_limit: int = 120,
    start_col: int = 1,
    col_limit: int = 26,
) -> dict:
    try:
        workbook = _resolve_workbook_path(workbook_path)
    except (FileNotFoundError, ValueError) as exc:
        return _error(str(exc))

    sheet = str(sheet_name or "").strip()
    if not sheet:
        return _error("Sheet name required")

    start_row = _safe_int(start_row, default=1, minimum=1, maximum=1_048_576)
    start_col = _safe_int(start_col, default=1, minimum=1, maximum=16_384)
    row_limit = _safe_int(row_limit, default=120, minimum=1, maximum=5000)
    col_limit = _safe_int(col_limit, default=26, minimum=1, maximum=80)

    wb_formula = load_workbook(workbook, data_only=False, read_only=True)
    if sheet not in wb_formula.sheetnames:
        wb_formula.close()
        return _error("Sheet not found", sheet=sheet)

    wb_values = load_workbook(workbook, data_only=True, read_only=True)
    ws_formula = wb_formula[sheet]
    ws_values = wb_values[sheet]

    max_row = max(int(ws_formula.max_row or 1), 1)
    max_col = max(int(ws_formula.max_column or 1), 1)

    start_row = min(start_row, max_row)
    start_col = min(start_col, max_col)
    end_row = min(max_row, start_row + row_limit - 1)
    end_col = min(max_col, start_col + col_limit - 1)

    rows = []
    formula_iter = ws_formula.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
    )
    values_iter = ws_values.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
    )

    for row_offset, (f_row, v_row) in enumerate(zip(formula_iter, values_iter)):
        row_number = start_row + row_offset
        cell_items = []
        for col_offset, (f_cell, v_cell) in enumerate(zip(f_row, v_row)):
            col_number = start_col + col_offset
            address = f"{get_column_letter(col_number)}{row_number}"
            formula_value = getattr(f_cell, "value", None)
            cached_value = getattr(v_cell, "value", None)
            is_formula = isinstance(formula_value, str) and formula_value.startswith("=")
            display_value = cached_value if is_formula else formula_value
            cell_items.append(
                {
                    "address": address,
                    "row": row_number,
                    "col": col_number,
                    "display": _to_jsonable(display_value),
                    "value": _to_jsonable(formula_value),
                    "formula": formula_value if is_formula else "",
                    "is_formula": is_formula,
                }
            )
        rows.append({"row_number": row_number, "cells": cell_items})

    sheet_state = ws_formula.sheet_state
    sheet_dimension = ws_formula.calculate_dimension()
    wb_formula.close()
    wb_values.close()

    columns = [{"index": col_idx, "label": get_column_letter(col_idx)} for col_idx in range(start_col, end_col + 1)]
    return {
        "ok": True,
        "workbook_path": str(workbook),
        "sheet": sheet,
        "sheet_state": sheet_state,
        "dimension": sheet_dimension,
        "max_row": max_row,
        "max_col": max_col,
        "start_row": start_row,
        "end_row": end_row,
        "start_col": start_col,
        "end_col": end_col,
        "columns": columns,
        "rows": rows,
    }


def apply_workbook_edits(edits: list[dict], workbook_path: str | None = None) -> dict:
    try:
        workbook = _resolve_workbook_path(workbook_path)
    except (FileNotFoundError, ValueError) as exc:
        return _error(str(exc))

    if not isinstance(edits, list) or not edits:
        return _error("No edits provided")

    wb = load_workbook(workbook, data_only=False, read_only=False)
    invalid = []
    applied = []

    for item in edits:
        if not isinstance(item, dict):
            invalid.append({"edit": item, "error": "Invalid edit payload"})
            continue
        sheet = str(item.get("sheet") or "").strip()
        cell = str(item.get("cell") or "").strip().upper()
        if not sheet or not cell:
            invalid.append({"edit": item, "error": "sheet and cell required"})
            continue
        if sheet not in wb.sheetnames:
            invalid.append({"edit": item, "error": f"Sheet not found: {sheet}"})
            continue
        if not re.fullmatch(r"[A-Z]{1,3}\d{1,7}", cell):
            invalid.append({"edit": item, "error": f"Invalid cell address: {cell}"})
            continue
        value = _coerce_cell_input(item.get("value"))
        wb[sheet][cell].value = value
        applied.append({"sheet": sheet, "cell": cell})

    if not applied:
        wb.close()
        return _error("No valid edits to apply", invalid=invalid)

    wb.save(workbook)
    wb.close()

    _AUDIT_CACHE.clear()
    saved_stat = workbook.stat()
    return {
        "ok": True,
        "workbook_path": str(workbook),
        "applied_count": len(applied),
        "invalid_count": len(invalid),
        "applied": applied,
        "invalid": invalid,
        "saved_at": datetime.fromtimestamp(saved_stat.st_mtime).isoformat(sep=" ", timespec="seconds"),
    }


def _recalculate_with_win32(workbook: Path) -> str:
    import win32com.client  # type: ignore

    excel = None
    wbk = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wbk = excel.Workbooks.Open(str(workbook), UpdateLinks=0, ReadOnly=False)
        wbk.RefreshAll()
        try:
            excel.CalculateUntilAsyncQueriesDone()
        except Exception:
            pass
        excel.CalculateFullRebuild()
        wbk.Save()
        wbk.Close(SaveChanges=True)
        excel.Quit()
        return "win32com"
    finally:
        if wbk is not None:
            try:
                wbk.Close(SaveChanges=True)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


def _recalculate_with_powershell(workbook: Path) -> str:
    safe_path = str(workbook).replace("'", "''")
    script = f"""
$ErrorActionPreference = 'Stop'
$path = '{safe_path}'
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
$wb = $excel.Workbooks.Open($path, 0, $false)
$wb.RefreshAll()
try {{ $excel.CalculateUntilAsyncQueriesDone() }} catch {{ }}
$excel.CalculateFullRebuild()
$wb.Save()
$wb.Close($true)
$excel.Quit()
"""
    proc = subprocess.run(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
        capture_output=True,
        text=True,
        timeout=900,
    )
    if proc.returncode != 0:
        raise RuntimeError(proc.stderr.strip() or proc.stdout.strip() or "PowerShell COM recalculate failed")
    return "powershell_com"


def recalculate_workbook(workbook_path: str | None = None) -> dict:
    try:
        workbook = _resolve_workbook_path(workbook_path)
    except (FileNotFoundError, ValueError) as exc:
        return _error(str(exc))

    detail = ""
    engine = ""
    try:
        engine = _recalculate_with_win32(workbook)
    except Exception as exc:
        detail = str(exc)
        try:
            engine = _recalculate_with_powershell(workbook)
        except Exception as fallback_exc:
            return _error("Cannot recalculate workbook by Excel engine", detail=f"{detail}; fallback: {fallback_exc}")

    _AUDIT_CACHE.clear()
    stat = workbook.stat()
    return {
        "ok": True,
        "workbook_path": str(workbook),
        "engine": engine,
        "saved_at": datetime.fromtimestamp(stat.st_mtime).isoformat(sep=" ", timespec="seconds"),
    }
