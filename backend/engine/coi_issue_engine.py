from __future__ import annotations

import asyncio
import copy
from datetime import datetime
import os
from pathlib import Path
import re
import shutil
import threading
import uuid
from urllib.parse import parse_qs, quote, unquote, urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from backend.engine.coi_issue_archive import (
    archive_issue_snapshot,
    get_latest_published_issue,
    mark_issue_batch_published,
    rebuild_combined_issued_coi,
    replace_latest_published_issue_snapshot,
)
from backend.engine.sql_live_engine import (
    build_live_coi_sheet,
    get_go_issue_state,
    record_go_issue_event,
)
from backend.sources import CACHE_DIR, ONEDRIVE_COI_FOLDER_PATH


_ISSUE_DIR = CACHE_DIR / "issued_coi"
_SHEET_NAME = "FORMAT COI REQUEST"
_FIELD_S = "AH Allocate Q'ty (yds)"
_FIELD_R = "Allocate Q'ty (Yds)"
_SHAREPOINT_UPLOAD_TIMEOUT_SEC = max(5, int(os.getenv("SHAREPOINT_UPLOAD_TIMEOUT_SEC", "20") or "20"))
_ISSUE_JOB_LOCK = threading.Lock()
_ISSUE_JOBS: dict[str, dict] = {}


def _error(message: str, **extra) -> dict:
    return {"ok": False, "error": message, **extra}


def _issue_job_now() -> str:
    return datetime.now().isoformat(sep=" ", timespec="seconds")


def _copy_issue_job_payload(payload: dict | None) -> dict:
    return copy.deepcopy(payload if isinstance(payload, dict) else {})


def _get_issue_job(go: str) -> dict:
    go_key = str(go or "").strip().upper()
    with _ISSUE_JOB_LOCK:
        return _copy_issue_job_payload(_ISSUE_JOBS.get(go_key))


def _set_issue_job(go: str, **updates) -> dict:
    go_key = str(go or "").strip().upper()
    with _ISSUE_JOB_LOCK:
        current = dict(_ISSUE_JOBS.get(go_key) or {})
        current.update(updates)
        _ISSUE_JOBS[go_key] = current
        return _copy_issue_job_payload(current)


def _sanitize_filename_part(value: object) -> str:
    text = str(value or "").strip().upper()
    text = re.sub(r"[^A-Z0-9\-_.]+", "_", text)
    return text or "GO"


def _issue_filename(go: str, issued_at: datetime) -> str:
    date_token = issued_at.strftime("%Y%m%d")
    return f"{_sanitize_filename_part(go)}-COI-{date_token}.xlsx"


def _to_number(value: object) -> float | None:
    if value is None or value == "":
        return None
    text = str(value).replace(",", "").replace("%", "").strip()
    if not text:
        return None
    try:
        numeric = float(text)
    except (TypeError, ValueError):
        return None
    if numeric != numeric or abs(numeric) == float("inf"):
        return None
    return numeric


def _build_export_rows(rows: list[dict]) -> tuple[list[dict], bool]:
    export_rows = [dict(item) for item in rows or []]
    has_manual_edit = any(str(item.get(_FIELD_S) or "").strip() != "" for item in export_rows)
    if has_manual_edit:
        for item in export_rows:
            if str(item.get(_FIELD_S) or "").strip() == "":
                item[_FIELD_S] = item.get(_FIELD_R, "")
    return export_rows, has_manual_edit


def _effective_allocate_for_row(row: dict) -> float:
    manual_raw = row.get(_FIELD_S)
    if str(manual_raw or "").strip() != "":
        return max(_to_number(manual_raw) or 0.0, 0.0)
    return max(_to_number(row.get(_FIELD_R)) or 0.0, 0.0)


def _write_issue_workbook(payload: dict, output_path: Path) -> dict:
    columns = list(payload.get("columns") or [])
    sheet_meta = dict(payload.get("sheet") or {})
    rows, has_manual_edit = _build_export_rows(payload.get("rows") or [])
    header_row = int(sheet_meta.get("header_row") or 8)
    first_data_row = int(sheet_meta.get("first_data_row") or (header_row + 1))

    wb = Workbook()
    ws = wb.active
    ws.title = _SHEET_NAME
    ws.freeze_panes = f"A{first_data_row}"

    title_fill = PatternFill("solid", fgColor="EAF2E5")
    head_fill = PatternFill("solid", fgColor="DDEAD8")
    border_align = Alignment(vertical="center", horizontal="left")

    ws.cell(row=1, column=1, value=f"GO: {payload.get('go', '')}")
    ws.cell(row=1, column=1).font = Font(size=12, bold=True)
    ws.cell(row=2, column=1, value=f"Issue at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(row=3, column=1, value=f"Factory: {payload.get('factory_code', '')}")

    for idx, col in enumerate(columns, start=1):
        letter = str(col.get("letter") or "").strip()
        key = str(col.get("key") or "").strip()
        label = str(col.get("label") or key)
        width = int(col.get("width") or 100)
        ws.column_dimensions[ws.cell(row=header_row, column=idx).column_letter].width = max(10, min(48, round(width / 8)))

        cell = ws.cell(row=header_row, column=idx, value=label)
        cell.font = Font(bold=True)
        cell.fill = title_fill if key == _FIELD_S else head_fill
        cell.alignment = border_align

        for row_offset, row in enumerate(rows, start=first_data_row):
            value = row.get(key, "")
            data_cell = ws.cell(row=row_offset, column=idx, value=value)
            data_cell.alignment = border_align

        if letter == "S" and not has_manual_edit:
            ws.column_dimensions[ws.cell(row=header_row, column=idx).column_letter].hidden = True
        if letter == "R" and has_manual_edit:
            ws.column_dimensions[ws.cell(row=header_row, column=idx).column_letter].hidden = True

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
    return {
        "ok": True,
        "file_path": str(output_path),
        "filename": output_path.name,
        "row_count": len(rows),
        "has_manual_edit": has_manual_edit,
        "hidden_column": "R" if has_manual_edit else "S",
    }


def _save_issue_workbook_to_onedrive(local_path: Path, filename: str) -> dict:
    folder_path = Path(ONEDRIVE_COI_FOLDER_PATH).expanduser()
    if not str(folder_path).strip():
        return _error("OneDrive COI folder path is empty")
    if not local_path.exists():
        return _error("Issue workbook was not created before OneDrive save", file_path=str(local_path))
    try:
        folder_path.mkdir(parents=True, exist_ok=True)
        target_path = (folder_path / filename).resolve()
        shutil.copy2(local_path, target_path)
    except Exception as exc:
        return _error(
            "Cannot save ISSUE COI workbook to OneDrive folder",
            detail=str(exc),
            folder_path=str(folder_path),
            source_path=str(local_path),
        )
    return {
        "ok": True,
        "method": "onedrive_sync_folder",
        "folder_path": str(folder_path),
        "file_path": str(target_path),
        "filename": filename,
    }


def sync_published_issue_after_sheet_change(go: str, sheet_payload: dict) -> dict:
    """Synchronize an already issued COI after an operator changes its PPO.

    The individual workbook, the SQLite Cutting feed and the combined workbook
    are updated as one application action.  A GO that has never been issued is
    intentionally skipped: editing a draft must not create an ISSUE record.
    """
    go_key = str(go or "").strip().upper()
    if not go_key:
        return _error("GO number required")
    if not isinstance(sheet_payload, dict) or not sheet_payload.get("ok"):
        return _error("Current COI sheet is required before issue synchronization", go=go_key)
    if not [row for row in sheet_payload.get("rows") or [] if isinstance(row, dict)]:
        return _error("Cannot synchronize issued COI with no rows", go=go_key)

    current = get_latest_published_issue(go_key)
    if not current.get("ok"):
        return current
    if not current.get("published"):
        return {
            "ok": True,
            "go": go_key,
            "skipped": True,
            "reason": "GO has not been issued yet",
        }

    raw_filename = Path(str(current.get("filename") or "")).name
    if raw_filename:
        filename = f"{_sanitize_filename_part(Path(raw_filename).stem)}.xlsx"
    else:
        filename = _issue_filename(go_key, datetime.now())
    saved_local = str(current.get("local_file_path") or "").strip()
    output_path = Path(saved_local).expanduser() if saved_local else (_ISSUE_DIR / filename)
    output_path = output_path.resolve()
    temporary_path = output_path.with_name(
        f".{output_path.stem}-{uuid.uuid4().hex}.sync.xlsx"
    )

    try:
        export_result = _write_issue_workbook(sheet_payload, temporary_path)
        if not export_result.get("ok"):
            return export_result
        storage_result = _save_issue_workbook_to_onedrive(temporary_path, filename)
        if not storage_result.get("ok"):
            return {
                "ok": False,
                "go": go_key,
                "error": "Issued COI file was not synchronized to OneDrive",
                "storage": storage_result,
            }
        output_path.parent.mkdir(parents=True, exist_ok=True)
        temporary_path.replace(output_path)
        archive_result = replace_latest_published_issue_snapshot(
            go_key,
            sheet_payload,
            local_file_path=str(output_path),
            shared_file_path=str(storage_result.get("file_path") or ""),
        )
        if not archive_result.get("ok"):
            return {
                "ok": False,
                "go": go_key,
                "error": "Issued COI SQLite feed was not synchronized",
                "archive": archive_result,
            }
        combined_result = rebuild_combined_issued_coi(sync_to_onedrive=True)
        return {
            "ok": True,
            "go": go_key,
            "batch_id": archive_result.get("batch_id"),
            "last_synced_at": archive_result.get("last_synced_at"),
            "sync_revision": archive_result.get("sync_revision"),
            "file_path": str(storage_result.get("file_path") or output_path),
            "local_file_path": str(output_path),
            "storage": storage_result,
            "archive": archive_result,
            "combined": combined_result,
            "combined_sync_failed": not bool(combined_result.get("ok")),
        }
    except Exception as exc:
        return _error("Cannot synchronize issued COI after sheet change", go=go_key, detail=str(exc))
    finally:
        if temporary_path.exists():
            temporary_path.unlink(missing_ok=True)


def get_issue_job_status(go: str) -> dict:
    go_key = str(go or "").strip().upper()
    if not go_key:
        return {"go": "", "status": "IDLE", "job": {}, "issue_state": {"go": "", "issue_count": 0, "last_issued_at": ""}}
    job = _get_issue_job(go_key)
    if not job:
        return {
            "go": go_key,
            "status": "IDLE",
            "job": {},
            "issue_state": get_go_issue_state(go_key),
        }
    return {
        "go": go_key,
        "status": str(job.get("status") or "IDLE"),
        "job": job,
        "issue_state": get_go_issue_state(go_key),
    }


def _run_issue_job(go: str, target_url: str | None = None) -> None:
    go_key = str(go or "").strip().upper()
    _set_issue_job(
        go_key,
        status="RUNNING",
        started_at=_issue_job_now(),
        finished_at="",
        message=f"Issuing COI for {go_key}...",
        error="",
    )
    try:
        result = issue_coi_to_sharepoint(go_key, target_url=target_url)
    except Exception as exc:
        _set_issue_job(
            go_key,
            status="FAILED",
            finished_at=_issue_job_now(),
            message=f"ISSUE COI failed for {go_key}.",
            error=str(exc),
            result=_error("ISSUE COI background job failed", detail=str(exc), go=go_key),
        )
        return

    if result.get("ok"):
        storage = result.get("storage") if isinstance(result.get("storage"), dict) else {}
        upload_label = "saved to OneDrive" if storage.get("ok") else "exported locally"
        _set_issue_job(
            go_key,
            status="COMPLETED",
            finished_at=_issue_job_now(),
            message=f"ISSUE COI done for {go_key} ({upload_label}).",
            error="",
            result=result,
        )
        return

    export_result = result.get("export") if isinstance(result.get("export"), dict) else {}
    if export_result.get("ok"):
        upload_result = result.get("upload") if isinstance(result.get("upload"), dict) else {}
        _set_issue_job(
            go_key,
            status="EXPORTED",
            finished_at=_issue_job_now(),
            message=f"ISSUE COI exported locally for {go_key}.",
            error=str(upload_result.get("error") or result.get("error") or ""),
            result=result,
        )
        return

    _set_issue_job(
        go_key,
        status="FAILED",
        finished_at=_issue_job_now(),
        message=f"ISSUE COI failed for {go_key}.",
        error=str(result.get("error") or result.get("detail") or "Unknown issue error"),
        result=result,
    )


def queue_issue_coi_to_sharepoint(go: str, target_url: str | None = None) -> dict:
    go_key = str(go or "").strip().upper()
    if not go_key:
        return _error("GO number required")

    current = _get_issue_job(go_key)
    current_status = str(current.get("status") or "").strip().upper()
    if current_status in {"QUEUED", "RUNNING"}:
        return {
            "ok": True,
            "go": go_key,
            "queued": False,
            "already_running": True,
            "job": current,
            "issue_state": get_go_issue_state(go_key),
        }

    job = _set_issue_job(
        go_key,
        status="QUEUED",
        created_at=_issue_job_now(),
        started_at="",
        finished_at="",
        message=f"ISSUE COI queued for {go_key}.",
        error="",
        target_url=str(target_url or "").strip(),
        result={},
    )
    worker = threading.Thread(
        target=_run_issue_job,
        args=(go_key, target_url),
        daemon=True,
        name=f"coi-issue-{go_key}",
    )
    worker.start()
    return {
        "ok": True,
        "go": go_key,
        "queued": True,
        "already_running": False,
        "job": job,
        "issue_state": get_go_issue_state(go_key),
    }


def issue_coi_to_sharepoint(go: str, target_url: str | None = None) -> dict:
    go_key = str(go or "").strip().upper()
    if not go_key:
        return _error("GO number required")

    issue_state = get_go_issue_state(go_key)

    sheet_payload = build_live_coi_sheet(
        go_key,
        prefer_mes_cache=True,
        allow_live_mes=False,
        use_snapshot=True,
        persist_snapshot=True,
        allow_inline_build=False,
        snapshot_built_from="issue-precheck",
    )
    if sheet_payload.get("pending") and int(sheet_payload.get("row_count") or 0) <= 0:
        sheet_payload = build_live_coi_sheet(
            go_key,
            prefer_mes_cache=True,
            allow_live_mes=False,
            use_snapshot=False,
            persist_snapshot=True,
            snapshot_built_from="issue-live",
        )
    if not sheet_payload.get("ok"):
        return _error(
            "Cannot load COI sheet for issue",
            go=go_key,
            detail=sheet_payload.get("error") or sheet_payload.get("detail") or "",
        )
    sheet_rows = [row for row in sheet_payload.get("rows") or [] if isinstance(row, dict)]
    if not sheet_rows:
        cache_profile = sheet_payload.get("cache_profile") if isinstance(sheet_payload.get("cache_profile"), dict) else {}
        return _error(
            "Cannot ISSUE COI because COI sheet has no rows",
            go=go_key,
            detail=str(cache_profile.get("reason") or "PPO/fabric data is not ready").strip(),
            cache_profile=cache_profile,
            export={"ok": False, "row_count": 0},
        )

    issued_at = datetime.now()
    filename = _issue_filename(go_key, issued_at)
    output_path = (_ISSUE_DIR / filename).resolve()
    export_result = _write_issue_workbook(sheet_payload, output_path)
    if not export_result.get("ok"):
        return export_result

    archive_result = archive_issue_snapshot(
        sheet_payload,
        issued_at=issued_at.isoformat(sep=" ", timespec="seconds"),
        filename=filename,
        local_file_path=str(output_path),
        storage_state="LOCAL_EXPORTED",
    )

    storage_result = _save_issue_workbook_to_onedrive(output_path, filename)
    if storage_result.get("ok"):
        archive_publish_result = (
            mark_issue_batch_published(archive_result.get("batch_id"), storage_result)
            if archive_result.get("ok")
            else _error("ISSUE COI archive was not created", archive_error=archive_result.get("error", ""))
        )
        cutting_data_result = (
            {
                "ok": True,
                "storage": "SQLite",
                "mode": "latest_published_issue_per_go",
                "batch_id": archive_publish_result.get("batch_id"),
            }
            if archive_publish_result.get("ok")
            else _error("Cutting COI SQLite feed was not updated", archive_error=archive_publish_result.get("error", ""))
        )
        record_go_issue_event(go_key, issued_at=issued_at.isoformat(sep=" ", timespec="seconds"))
        return {
            "ok": True,
            "go": go_key,
            "issued_at": issued_at.isoformat(sep=" ", timespec="seconds"),
            "filename": filename,
            "file_path": storage_result.get("file_path") or str(output_path),
            "local_file_path": str(output_path),
            "issue_state": get_go_issue_state(go_key),
            "storage": storage_result,
            "upload": storage_result,
            "export": export_result,
            "archive": archive_result,
            "archive_publish": archive_publish_result,
            "cutting_data": cutting_data_result,
        }

    return {
        "ok": False,
        "go": go_key,
        "issued_at": issued_at.isoformat(sep=" ", timespec="seconds"),
        "filename": filename,
        "file_path": str(output_path),
        "issue_state": issue_state,
        "storage": storage_result,
        "upload": storage_result,
        "export": export_result,
        "archive": archive_result,
    }

