from __future__ import annotations

from contextlib import contextmanager
from datetime import datetime
import hashlib
import json
import os
from pathlib import Path
import re
import shutil
import sqlite3
import threading
import uuid
from typing import Any
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.sources import CACHE_DIR, ONEDRIVE_COI_FOLDER_PATH


ISSUED_COI_ARCHIVE_DB = Path(
    os.getenv("ISSUED_COI_ARCHIVE_DB", str(CACHE_DIR / "issued_coi_archive_v1.db"))
).expanduser()
ISSUED_COI_COMBINED_FILENAME = str(
    os.getenv("ISSUED_COI_COMBINED_FILENAME", "COI-CUTTING-COMBINED.xlsx") or "COI-CUTTING-COMBINED.xlsx"
).strip()
_ISSUE_DIR = CACHE_DIR / "issued_coi"
_COMBINED_SHEET = "COI Combined"
_SUMMARY_SHEET = "GO Summary"
_ABOUT_SHEET = "About"
_ARCHIVE_LOCK = threading.RLock()


_CANONICAL_COLUMNS = [
    ("BRAND", "BRAND"),
    ("GO#", "GO#"),
    ("PPO", "PPO"),
    ("Type", "Type"),
    ("COLOR_CODE", "COLOR_CODE"),
    ("COLOR_DESC", "COLOR_DESC"),
    ("FABRIC COLOR (For piecing only)", "FABRIC COLOR (For piecing only)"),
    ("JOB ORDER NO", "JOB ORDER NO"),
    ("LOT", "LOT"),
    ("SIZE", "SIZE"),
    ("- %", "- %"),
    ("+%", "+%"),
    ("Qty (pcs)", "Qty"),
    ("BUYER_PO_DEL_DATE", "BUYER_PO_DEL_DATE"),
    ("Net YY", "Net YY"),
    ("PPO YY", "PPO YY"),
    ("Marker YY", "Marker YY"),
    ("Required Q'ty (Yds)", "Required Q'ty (Yds)"),
    ("Rcv Q'ty (PPO)", "Rcv Q'ty (PPO)"),
    ("On The Way Q'ty (Yds)", "On The Way Q'ty (Yds)"),
    ("Allocate Q'ty (Yds)", "Allocate Q'ty (Lot)"),
    ("Shortage Q'ty (Yds)", "Shortage Q'ty (Yds)"),
    ("AH Allocate Q'ty (yds)", "AH Allocate Q'ty (yds)"),
    ("Allocate %", "Allocate %"),
    ("ETD Fabric", "ETD Fabric"),
    ("User Remark", "User Remark"),
    ("PPO Order Total (Yds)", "PPO Q'ty"),
    ("SAMPLE STATUS", "SAMPLE STATUS"),
]
_LABEL_TO_KEY = {label.casefold(): key for key, label in _CANONICAL_COLUMNS}
_KEY_TO_LABEL = {key: label for key, label in _CANONICAL_COLUMNS}
_METADATA_COLUMNS = [
    ("ISSUE_AT", "ISSUE AT"),
    ("ISSUE_SYNC_AT", "LAST SYNC AT"),
    ("ISSUE_FILE", "ISSUE FILE"),
    ("ISSUE_VERSION", "ISSUE VERSION"),
]
_FEED_DEFAULT_LIMIT = 5000
_FEED_MAX_LIMIT = 10000


def _now() -> str:
    return datetime.now().isoformat(sep=" ", timespec="seconds")


def _error(message: str, **extra: Any) -> dict[str, Any]:
    return {"ok": False, "error": message, **extra}


@contextmanager
def _connect():
    ISSUED_COI_ARCHIVE_DB.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(ISSUED_COI_ARCHIVE_DB), timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def ensure_issued_coi_archive() -> None:
    with _connect() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS issued_coi_batches (
                batch_id INTEGER PRIMARY KEY AUTOINCREMENT,
                go_no TEXT NOT NULL,
                issued_at TEXT NOT NULL,
                filename TEXT NOT NULL,
                local_file_path TEXT NOT NULL DEFAULT '',
                shared_file_path TEXT NOT NULL DEFAULT '',
                storage_state TEXT NOT NULL DEFAULT 'LOCAL_EXPORTED',
                row_count INTEGER NOT NULL DEFAULT 0,
                has_manual_edit INTEGER NOT NULL DEFAULT 0,
                source_hash TEXT NOT NULL DEFAULT '',
                source_file_hash TEXT NOT NULL DEFAULT '',
                columns_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                last_synced_at TEXT NOT NULL DEFAULT '',
                sync_revision INTEGER NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS issued_coi_rows (
                batch_id INTEGER NOT NULL,
                row_index INTEGER NOT NULL,
                row_key TEXT NOT NULL DEFAULT '',
                ppo_no TEXT NOT NULL DEFAULT '',
                jo_no TEXT NOT NULL DEFAULT '',
                fabric_type TEXT NOT NULL DEFAULT '',
                color_code TEXT NOT NULL DEFAULT '',
                lot_no TEXT NOT NULL DEFAULT '',
                row_json TEXT NOT NULL,
                PRIMARY KEY (batch_id, row_index),
                FOREIGN KEY (batch_id) REFERENCES issued_coi_batches(batch_id) ON DELETE CASCADE
            );
            """
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_issued_coi_batches_go ON issued_coi_batches(go_no, issued_at DESC, batch_id DESC)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_issued_coi_batches_state ON issued_coi_batches(storage_state, issued_at DESC)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_issued_coi_batches_file_hash ON issued_coi_batches(source_file_hash)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_issued_coi_rows_batch ON issued_coi_rows(batch_id, row_index)"
        )
        # The first archive schema predates automatic PPO re-sync.  Keep this
        # migration local and idempotent so existing issued history is usable.
        batch_columns = {
            str(row["name"] or "").strip().lower()
            for row in conn.execute("PRAGMA table_info(issued_coi_batches)").fetchall()
        }
        if "last_synced_at" not in batch_columns:
            conn.execute(
                "ALTER TABLE issued_coi_batches ADD COLUMN last_synced_at TEXT NOT NULL DEFAULT ''"
            )
        if "sync_revision" not in batch_columns:
            conn.execute(
                "ALTER TABLE issued_coi_batches ADD COLUMN sync_revision INTEGER NOT NULL DEFAULT 0"
            )


def _json_value(value: object) -> object:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, (list, tuple)):
        return [_json_value(item) for item in value]
    if isinstance(value, dict):
        return {str(key): _json_value(item) for key, item in value.items()}
    return str(value)


def _canonical_key(key: object, label: object = "") -> str:
    raw_key = str(key or "").strip()
    raw_label = str(label or "").strip()
    if raw_label.casefold() in _LABEL_TO_KEY:
        return _LABEL_TO_KEY[raw_label.casefold()]
    if raw_key.casefold() in _LABEL_TO_KEY:
        return _LABEL_TO_KEY[raw_key.casefold()]
    for canonical, preferred_label in _CANONICAL_COLUMNS:
        if raw_key.casefold() == canonical.casefold() or raw_label.casefold() == canonical.casefold():
            return canonical
        if raw_key.casefold() == preferred_label.casefold():
            return canonical
    return raw_key or raw_label


def _normalise_columns(columns: object) -> list[dict[str, str]]:
    result: list[dict[str, str]] = []
    seen: set[str] = set()
    for item in columns if isinstance(columns, list) else []:
        if not isinstance(item, dict):
            continue
        source_key = str(item.get("key") or item.get("label") or "").strip()
        label = str(item.get("label") or source_key).strip()
        key = _canonical_key(source_key, label)
        if not key or key in seen:
            continue
        seen.add(key)
        result.append({"key": key, "label": _KEY_TO_LABEL.get(key, label or key), "source_key": source_key})
    if result:
        return result
    return [
        {"key": key, "label": label, "source_key": key}
        for key, label in _CANONICAL_COLUMNS
    ]


def _normalise_rows(rows: object, columns: list[dict[str, str]], fallback_go: str = "") -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for row in rows if isinstance(rows, list) else []:
        if not isinstance(row, dict):
            continue
        output: dict[str, Any] = {}
        for column in columns:
            key = column["key"]
            source_key = column["source_key"]
            value = row.get(source_key)
            if value is None and source_key != key:
                value = row.get(key)
            output[key] = _json_value(value)
        if not str(output.get("GO#") or "").strip() and fallback_go:
            output["GO#"] = fallback_go
        if not any(value not in (None, "") for value in output.values()):
            continue
        output["_row_key"] = str(row.get("_row_key") or "").strip()
        normalized.append(output)
    return normalized


def _snapshot_hash(columns: list[dict[str, str]], rows: list[dict[str, Any]]) -> str:
    payload = {
        "columns": [{"key": item["key"], "label": item["label"]} for item in columns],
        "rows": rows,
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def _file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def archive_issue_snapshot(
    payload: dict[str, Any],
    *,
    issued_at: str,
    filename: str,
    local_file_path: str = "",
    source_file_hash: str = "",
    storage_state: str = "LOCAL_EXPORTED",
) -> dict[str, Any]:
    go_no = str(payload.get("go") or "").strip().upper()
    if not go_no:
        return _error("GO number required for ISSUE archive")
    columns = _normalise_columns(payload.get("columns"))
    rows = _normalise_rows(payload.get("rows"), columns, fallback_go=go_no)
    if not rows:
        return _error("Cannot archive ISSUE COI with no rows", go=go_no)
    if not source_file_hash and local_file_path:
        try:
            candidate_path = Path(local_file_path)
            if candidate_path.exists():
                source_file_hash = _file_hash(candidate_path)
        except Exception:
            source_file_hash = ""
    source_hash = _snapshot_hash(columns, rows)
    has_manual_edit = any(str(row.get("AH Allocate Q'ty (yds)") or "").strip() for row in rows)
    with _ARCHIVE_LOCK:
        ensure_issued_coi_archive()
        with _connect() as conn:
            if source_file_hash:
                existing = conn.execute(
                    "SELECT batch_id FROM issued_coi_batches WHERE source_file_hash = ? ORDER BY batch_id DESC LIMIT 1",
                    (source_file_hash,),
                ).fetchone()
                if existing:
                    return {
                        "ok": True,
                        "batch_id": int(existing["batch_id"]),
                        "go": go_no,
                        "row_count": len(rows),
                        "deduplicated": True,
                    }
            cursor = conn.execute(
                """
                INSERT INTO issued_coi_batches (
                    go_no, issued_at, filename, local_file_path, storage_state,
                    row_count, has_manual_edit, source_hash, source_file_hash,
                    columns_json, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    go_no,
                    str(issued_at or _now()),
                    str(filename or "").strip(),
                    str(local_file_path or "").strip(),
                    str(storage_state or "LOCAL_EXPORTED").strip().upper(),
                    len(rows),
                    int(bool(has_manual_edit)),
                    source_hash,
                    str(source_file_hash or "").strip(),
                    json.dumps(
                        [{"key": item["key"], "label": item["label"]} for item in columns],
                        ensure_ascii=False,
                    ),
                    _now(),
                ),
            )
            batch_id = int(cursor.lastrowid)
            params = []
            for row_index, row in enumerate(rows, start=1):
                params.append(
                    (
                        batch_id,
                        row_index,
                        str(row.get("_row_key") or ""),
                        str(row.get("PPO") or "").strip().upper(),
                        str(row.get("JOB ORDER NO") or "").strip().upper(),
                        str(row.get("Type") or "").strip().upper(),
                        str(row.get("COLOR_CODE") or "").strip().upper(),
                        str(row.get("LOT") or "").strip(),
                        json.dumps(row, ensure_ascii=False, default=str),
                    )
                )
            conn.executemany(
                """
                INSERT INTO issued_coi_rows (
                    batch_id, row_index, row_key, ppo_no, jo_no, fabric_type,
                    color_code, lot_no, row_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                params,
            )
    return {
        "ok": True,
        "batch_id": batch_id,
        "go": go_no,
        "row_count": len(rows),
        "deduplicated": False,
        "source_hash": source_hash,
    }


def get_latest_published_issue(go: object) -> dict[str, Any]:
    """Locate the issued COI that is currently exposed to Cutting for one GO."""
    go_no = str(go or "").strip().upper()
    if not go_no:
        return _error("GO number required")
    with _ARCHIVE_LOCK:
        ensure_issued_coi_archive()
        with _connect() as conn:
            row = conn.execute(
                """
                SELECT *
                FROM issued_coi_batches
                WHERE go_no = ? AND storage_state = 'PUBLISHED'
                ORDER BY issued_at DESC, batch_id DESC
                LIMIT 1
                """,
                (go_no,),
            ).fetchone()
    if not row:
        return {
            "ok": True,
            "go": go_no,
            "published": False,
            "reason": "GO has not been issued yet",
        }
    return {
        "ok": True,
        "go": go_no,
        "published": True,
        "batch_id": int(row["batch_id"]),
        "filename": str(row["filename"] or ""),
        "local_file_path": str(row["local_file_path"] or ""),
        "shared_file_path": str(row["shared_file_path"] or ""),
        "issued_at": str(row["issued_at"] or ""),
        "last_synced_at": str(row["last_synced_at"] or ""),
        "sync_revision": int(row["sync_revision"] or 0),
    }


def replace_latest_published_issue_snapshot(
    go: object,
    payload: dict[str, Any],
    *,
    local_file_path: str = "",
    shared_file_path: str = "",
) -> dict[str, Any]:
    """Replace the *current* issued representation after a UI PPO edit.

    The original issue timestamp and batch id remain intact.  This is important:
    Cutting sees the corrected current data while the record still identifies the
    original issue instead of pretending the user issued a different COI.
    """
    go_no = str(go or payload.get("go") or "").strip().upper()
    if not go_no:
        return _error("GO number required")
    columns = _normalise_columns(payload.get("columns"))
    rows = _normalise_rows(payload.get("rows"), columns, fallback_go=go_no)
    if not rows:
        return _error("Cannot synchronize issued COI with no rows", go=go_no)
    source_hash = _snapshot_hash(columns, rows)
    has_manual_edit = any(str(row.get("AH Allocate Q'ty (yds)") or "").strip() for row in rows)
    synced_at = _now()

    with _ARCHIVE_LOCK:
        ensure_issued_coi_archive()
        with _connect() as conn:
            batch = conn.execute(
                """
                SELECT batch_id, filename, local_file_path, shared_file_path, sync_revision
                FROM issued_coi_batches
                WHERE go_no = ? AND storage_state = 'PUBLISHED'
                ORDER BY issued_at DESC, batch_id DESC
                LIMIT 1
                """,
                (go_no,),
            ).fetchone()
            if not batch:
                return {
                    "ok": True,
                    "go": go_no,
                    "published": False,
                    "skipped": True,
                    "reason": "GO has not been issued yet",
                }
            batch_id = int(batch["batch_id"])
            resolved_local_path = str(local_file_path or batch["local_file_path"] or "").strip()
            resolved_shared_path = str(shared_file_path or batch["shared_file_path"] or "").strip()
            cursor = conn.execute(
                """
                UPDATE issued_coi_batches
                SET local_file_path = ?,
                    shared_file_path = ?,
                    row_count = ?,
                    has_manual_edit = ?,
                    source_hash = ?,
                    columns_json = ?,
                    last_synced_at = ?,
                    sync_revision = COALESCE(sync_revision, 0) + 1
                WHERE batch_id = ?
                """,
                (
                    resolved_local_path,
                    resolved_shared_path,
                    len(rows),
                    int(bool(has_manual_edit)),
                    source_hash,
                    json.dumps(
                        [{"key": item["key"], "label": item["label"]} for item in columns],
                        ensure_ascii=False,
                    ),
                    synced_at,
                    batch_id,
                ),
            )
            if int(cursor.rowcount or 0) <= 0:
                return _error("ISSUE archive batch was not found", batch_id=batch_id)
            conn.execute("DELETE FROM issued_coi_rows WHERE batch_id = ?", (batch_id,))
            conn.executemany(
                """
                INSERT INTO issued_coi_rows (
                    batch_id, row_index, row_key, ppo_no, jo_no, fabric_type,
                    color_code, lot_no, row_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                [
                    (
                        batch_id,
                        row_index,
                        str(row.get("_row_key") or ""),
                        str(row.get("PPO") or "").strip().upper(),
                        str(row.get("JOB ORDER NO") or "").strip().upper(),
                        str(row.get("Type") or "").strip().upper(),
                        str(row.get("COLOR_CODE") or "").strip().upper(),
                        str(row.get("LOT") or "").strip(),
                        json.dumps(row, ensure_ascii=False, default=str),
                    )
                    for row_index, row in enumerate(rows, start=1)
                ],
            )
            revision_row = conn.execute(
                "SELECT sync_revision FROM issued_coi_batches WHERE batch_id = ?",
                (batch_id,),
            ).fetchone()
    return {
        "ok": True,
        "go": go_no,
        "published": True,
        "batch_id": batch_id,
        "row_count": len(rows),
        "last_synced_at": synced_at,
        "sync_revision": int(revision_row["sync_revision"] or 0) if revision_row else 0,
        "source_hash": source_hash,
    }


def mark_issue_batch_published(batch_id: object, storage_result: dict[str, Any] | None = None) -> dict[str, Any]:
    try:
        resolved_batch_id = int(batch_id)
    except (TypeError, ValueError):
        return _error("Invalid ISSUE archive batch id")
    storage = storage_result if isinstance(storage_result, dict) else {}
    shared_file_path = str(storage.get("file_path") or "").strip()
    with _ARCHIVE_LOCK:
        ensure_issued_coi_archive()
        with _connect() as conn:
            cursor = conn.execute(
                """
                UPDATE issued_coi_batches
                SET storage_state = 'PUBLISHED', shared_file_path = ?
                WHERE batch_id = ?
                """,
                (shared_file_path, resolved_batch_id),
            )
            if int(cursor.rowcount or 0) <= 0:
                return _error("ISSUE archive batch was not found", batch_id=resolved_batch_id)
    return {"ok": True, "batch_id": resolved_batch_id, "storage_state": "PUBLISHED"}


def _latest_published_batches(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute(
        """
        WITH ranked AS (
            SELECT *,
                   ROW_NUMBER() OVER (
                       PARTITION BY go_no
                       ORDER BY issued_at DESC, batch_id DESC
                   ) AS version_rank
            FROM issued_coi_batches
            WHERE storage_state = 'PUBLISHED'
        )
        SELECT * FROM ranked
        WHERE version_rank = 1
        ORDER BY go_no ASC
        """
    ).fetchall()


def _combined_columns(batches: list[sqlite3.Row]) -> list[dict[str, str]]:
    seen: set[str] = set()
    columns: list[dict[str, str]] = []
    for canonical_key, label in _CANONICAL_COLUMNS:
        seen.add(canonical_key)
        columns.append({"key": canonical_key, "label": label})
    for batch in batches:
        try:
            source_columns = json.loads(str(batch["columns_json"] or "[]"))
        except json.JSONDecodeError:
            source_columns = []
        for item in source_columns if isinstance(source_columns, list) else []:
            if not isinstance(item, dict):
                continue
            key = _canonical_key(item.get("key"), item.get("label"))
            if not key or key.startswith("_") or key in seen:
                continue
            seen.add(key)
            columns.append({"key": key, "label": str(item.get("label") or key)})
    return columns


def _latest_published_snapshot() -> tuple[list[sqlite3.Row], list[dict[str, str]], list[dict[str, Any]]]:
    ensure_issued_coi_archive()
    with _connect() as conn:
        batches = _latest_published_batches(conn)
        if not batches:
            return [], [], []
        columns = _combined_columns(batches)
        rows: list[dict[str, Any]] = []
        for batch in batches:
            issue_rows = conn.execute(
                "SELECT row_index, row_json FROM issued_coi_rows WHERE batch_id = ? ORDER BY row_index",
                (int(batch["batch_id"]),),
            ).fetchall()
            for issue_row in issue_rows:
                try:
                    row = json.loads(str(issue_row["row_json"] or "{}"))
                except json.JSONDecodeError:
                    row = {}
                if not isinstance(row, dict):
                    continue
                row["ISSUE_AT"] = str(batch["issued_at"] or "")
                row["ISSUE_SYNC_AT"] = str(batch["last_synced_at"] or batch["issued_at"] or "")
                row["ISSUE_FILE"] = str(batch["filename"] or "")
                row["ISSUE_VERSION"] = int(batch["batch_id"])
                rows.append(row)
    return batches, columns, rows


def _feed_filter(value: object) -> str:
    """Keep feed filters bounded and consistent with the normalized archive keys."""
    return str(value or "").strip().upper()[:160]


def _feed_pagination(value: object, *, default: int, minimum: int, maximum: int) -> int:
    try:
        parsed = int(str(value or default).strip())
    except (TypeError, ValueError):
        parsed = default
    return max(minimum, min(maximum, parsed))


def get_latest_issued_coi_feed(
    *,
    go: object = "",
    ppo: object = "",
    jo: object = "",
    color_code: object = "",
    limit: object = _FEED_DEFAULT_LIMIT,
    offset: object = 0,
) -> dict[str, Any]:
    """Return the latest published ISSUE snapshot for every GO as JSON-ready data.

    The archive keeps every ISSUE version for audit purposes.  This feed deliberately
    exposes only one version per GO: the latest successfully published ISSUE.  It is
    the database replacement for the old combined Excel download.
    """
    filters = {
        "go": _feed_filter(go),
        "ppo": _feed_filter(ppo),
        "jo": _feed_filter(jo),
        "color_code": _feed_filter(color_code),
    }
    page_limit = _feed_pagination(limit, default=_FEED_DEFAULT_LIMIT, minimum=1, maximum=_FEED_MAX_LIMIT)
    page_offset = _feed_pagination(offset, default=0, minimum=0, maximum=10_000_000)
    row_conditions: list[str] = []
    row_params: list[object] = []
    if filters["go"]:
        row_conditions.append("b.go_no = ?")
        row_params.append(filters["go"])
    if filters["ppo"]:
        row_conditions.append("r.ppo_no = ?")
        row_params.append(filters["ppo"])
    if filters["jo"]:
        row_conditions.append("r.jo_no = ?")
        row_params.append(filters["jo"])
    if filters["color_code"]:
        row_conditions.append("r.color_code = ?")
        row_params.append(filters["color_code"])
    where_clause = f"WHERE {' AND '.join(row_conditions)}" if row_conditions else ""
    latest_cte = """
        WITH ranked AS (
            SELECT *,
                   ROW_NUMBER() OVER (
                       PARTITION BY go_no
                       ORDER BY issued_at DESC, batch_id DESC
                   ) AS version_rank
            FROM issued_coi_batches
            WHERE storage_state = 'PUBLISHED'
        ), latest_batches AS (
            SELECT * FROM ranked WHERE version_rank = 1
        )
    """

    with _ARCHIVE_LOCK:
        ensure_issued_coi_archive()
        with _connect() as conn:
            batch_where = "WHERE go_no = ?" if filters["go"] else ""
            batch_params: list[object] = [filters["go"]] if filters["go"] else []
            batches = conn.execute(
                latest_cte
                + f"SELECT * FROM latest_batches {batch_where} ORDER BY go_no ASC",
                batch_params,
            ).fetchall()
            columns = _combined_columns(batches)
            columns.extend({"key": key, "label": label} for key, label in _METADATA_COLUMNS)

            counts = conn.execute(
                latest_cte
                + """
                SELECT
                    COUNT(*) AS row_count,
                    COUNT(DISTINCT b.batch_id) AS go_count,
                    MAX(b.issued_at) AS latest_issued_at,
                    MAX(COALESCE(NULLIF(b.last_synced_at, ''), b.issued_at)) AS latest_synced_at,
                    MAX(b.batch_id) AS latest_batch_id,
                    SUM(COALESCE(b.sync_revision, 0)) AS sync_revision_total
                FROM latest_batches b
                JOIN issued_coi_rows r ON r.batch_id = b.batch_id
                """
                + where_clause,
                row_params,
            ).fetchone()
            source_rows = conn.execute(
                latest_cte
                + """
                SELECT b.batch_id, b.go_no, b.issued_at, b.last_synced_at, b.filename, r.row_index, r.row_json
                FROM latest_batches b
                JOIN issued_coi_rows r ON r.batch_id = b.batch_id
                """
                + where_clause
                + " ORDER BY b.go_no ASC, r.row_index ASC LIMIT ? OFFSET ?",
                [*row_params, page_limit, page_offset],
            ).fetchall()

    count_data = dict(counts) if counts is not None else {}
    rows: list[dict[str, Any]] = []
    for source_row in source_rows:
        try:
            saved_row = json.loads(str(source_row["row_json"] or "{}"))
        except json.JSONDecodeError:
            saved_row = {}
        if not isinstance(saved_row, dict):
            saved_row = {}
        output = {
            column["key"]: _json_value(saved_row.get(column["key"]))
            for column in columns
            if not str(column["key"]).startswith("ISSUE_")
        }
        output["ISSUE_AT"] = str(source_row["issued_at"] or "")
        output["ISSUE_SYNC_AT"] = str(source_row["last_synced_at"] or source_row["issued_at"] or "")
        output["ISSUE_FILE"] = str(source_row["filename"] or "")
        output["ISSUE_VERSION"] = int(source_row["batch_id"] or 0)
        rows.append(output)

    total_rows = int(count_data.get("row_count") or 0)
    return {
        "ok": True,
        "feed": "cutting-coi-latest",
        "mode": "latest_published_issue_per_go",
        "columns": columns,
        "rows": rows,
        "filters": filters,
        "pagination": {
            "limit": page_limit,
            "offset": page_offset,
            "returned": len(rows),
            "total_rows": total_rows,
            "has_more": page_offset + len(rows) < total_rows,
        },
        "go_count": int(count_data.get("go_count") or 0),
        "latest_issued_at": str(count_data.get("latest_issued_at") or ""),
        "last_synced_at": str(count_data.get("latest_synced_at") or ""),
        "data_version": "{}:{}".format(
            int(count_data.get("latest_batch_id") or 0),
            int(count_data.get("sync_revision_total") or 0),
        ),
    }


def _write_combined_workbook(output_path: Path, batches: list[sqlite3.Row], columns: list[dict[str, str]], rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = _COMBINED_SHEET
    ws.sheet_view.showGridLines = False

    title_fill = PatternFill("solid", fgColor="103A2B")
    header_fill = PatternFill("solid", fgColor="DDEAD8")
    metadata_fill = PatternFill("solid", fgColor="E8F0FE")
    title_columns = len(columns) + len(_METADATA_COLUMNS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, title_columns))
    title = ws.cell(row=1, column=1, value="CUTTING - COMBINED ISSUED COI (LATEST VERSION PER GO)")
    title.fill = title_fill
    title.font = Font(color="FFFFFF", bold=True, size=13)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, title_columns))
    ws.cell(
        row=2,
        column=1,
        value=f"Generated: {_now()} | Latest published GO: {len(batches)} | COI rows: {len(rows)}",
    )
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="center")

    header_row = 4
    all_columns = columns + [{"key": key, "label": label} for key, label in _METADATA_COLUMNS]
    for index, column in enumerate(all_columns, start=1):
        cell = ws.cell(row=header_row, column=index, value=column["label"])
        cell.font = Font(bold=True)
        cell.fill = metadata_fill if column["key"].startswith("ISSUE_") else header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(index)].width = max(11, min(34, len(column["label"]) + 5))

    for row_index, row in enumerate(rows, start=header_row + 1):
        for column_index, column in enumerate(all_columns, start=1):
            value = row.get(column["key"], "")
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if column["key"] in {"Allocate %"} and isinstance(value, (int, float)):
                cell.number_format = "0.00%"
            elif column["key"] in {"Qty (pcs)", "Net YY", "PPO YY", "Marker YY", "Required Q'ty (Yds)", "Rcv Q'ty (PPO)", "On The Way Q'ty (Yds)", "Allocate Q'ty (Yds)", "Shortage Q'ty (Yds)", "AH Allocate Q'ty (yds)", "PPO Order Total (Yds)"}:
                cell.number_format = "#,##0.###"

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(all_columns))}{max(header_row, header_row + len(rows))}"

    summary = wb.create_sheet(_SUMMARY_SHEET)
    summary_headers = ["GO#", "BRAND", "LATEST ISSUE AT", "LAST SYNC AT", "ISSUE FILE", "ROWS", "ISSUE VERSION", "SHARED FILE"]
    for column_index, label in enumerate(summary_headers, start=1):
        cell = summary.cell(row=1, column=column_index, value=label)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        summary.column_dimensions[get_column_letter(column_index)].width = max(14, min(48, len(label) + 12))
    for row_index, batch in enumerate(batches, start=2):
        brand_row = next(
            (row for row in rows if str(row.get("GO#") or "").strip().upper() == str(batch["go_no"] or "").strip().upper()),
            {},
        )
        values = [
            batch["go_no"],
            brand_row.get("BRAND", ""),
            batch["issued_at"],
            batch["last_synced_at"] or batch["issued_at"],
            batch["filename"],
            batch["row_count"],
            batch["batch_id"],
            batch["shared_file_path"],
        ]
        for column_index, value in enumerate(values, start=1):
            summary.cell(row=row_index, column=column_index, value=value)
    summary.freeze_panes = "A2"
    summary.auto_filter.ref = f"A1:{get_column_letter(len(summary_headers))}{max(1, len(batches) + 1)}"

    about = wb.create_sheet(_ABOUT_SHEET)
    about_rows = [
        ("Purpose", "One latest successfully shared ISSUE COI per GO for Cutting."),
        ("Source", "ISSUE COI snapshots stored by the TEST application; not a folder scan at export time."),
        ("Deduplication", "A later successful ISSUE for the same GO replaces that GO's rows in COI Combined; history remains in SQLite."),
        ("Data columns", "The original ISSUE COI columns are preserved. ISSUE AT, LAST SYNC AT, ISSUE FILE and ISSUE VERSION are appended for traceability."),
        ("Generated at", _now()),
    ]
    for row_index, (label, value) in enumerate(about_rows, start=1):
        about.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        about.cell(row=row_index, column=2, value=value)
    about.column_dimensions["A"].width = 18
    about.column_dimensions["B"].width = 118

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_name(f".{output_path.stem}-{uuid.uuid4().hex}.tmp.xlsx")
    try:
        wb.save(temporary_path)
        temporary_path.replace(output_path)
    finally:
        wb.close()
        if temporary_path.exists():
            temporary_path.unlink(missing_ok=True)


def rebuild_combined_issued_coi(*, sync_to_onedrive: bool = True) -> dict[str, Any]:
    with _ARCHIVE_LOCK:
        batches, columns, rows = _latest_published_snapshot()
        if not batches:
            return _error("No published ISSUE COI snapshots are available for the combined workbook")
        filename = ISSUED_COI_COMBINED_FILENAME or "COI-CUTTING-COMBINED.xlsx"
        if not filename.lower().endswith(".xlsx"):
            filename += ".xlsx"
        output_path = (_ISSUE_DIR / filename).resolve()
        try:
            _write_combined_workbook(output_path, batches, columns, rows)
        except Exception as exc:
            return _error("Cannot create combined ISSUE COI workbook", detail=str(exc))

        storage: dict[str, Any] = {"ok": False, "skipped": not sync_to_onedrive}
        if sync_to_onedrive:
            folder_path = Path(ONEDRIVE_COI_FOLDER_PATH).expanduser()
            try:
                folder_path.mkdir(parents=True, exist_ok=True)
                shared_path = (folder_path / filename).resolve()
                shutil.copy2(output_path, shared_path)
                storage = {
                    "ok": True,
                    "method": "onedrive_sync_folder",
                    "file_path": str(shared_path),
                    "filename": filename,
                }
            except Exception as exc:
                storage = _error(
                    "Cannot save combined ISSUE COI workbook to OneDrive folder",
                    detail=str(exc),
                    folder_path=str(folder_path),
                    source_path=str(output_path),
                )
        return {
            "ok": True,
            "filename": filename,
            "file_path": str(output_path),
            "row_count": len(rows),
            "go_count": len(batches),
            "storage": storage,
        }


def issued_coi_archive_status() -> dict[str, Any]:
    ensure_issued_coi_archive()
    with _connect() as conn:
        published_batches = _latest_published_batches(conn)
        history_count = int(conn.execute("SELECT COUNT(*) FROM issued_coi_batches").fetchone()[0])
        history_rows = int(conn.execute("SELECT COUNT(*) FROM issued_coi_rows").fetchone()[0])
    filename = ISSUED_COI_COMBINED_FILENAME or "COI-CUTTING-COMBINED.xlsx"
    local_path = (_ISSUE_DIR / filename).resolve()
    shared_path = (Path(ONEDRIVE_COI_FOLDER_PATH).expanduser() / filename).resolve()
    return {
        "ok": True,
        "db_file": str(ISSUED_COI_ARCHIVE_DB),
        "published_go_count": len(published_batches),
        "published_row_count": sum(int(row["row_count"] or 0) for row in published_batches),
        "history_batch_count": history_count,
        "history_row_count": history_rows,
        "latest_issued_at": max((str(row["issued_at"] or "") for row in published_batches), default=""),
        "latest_synced_at": max(
            (str(row["last_synced_at"] or row["issued_at"] or "") for row in published_batches),
            default="",
        ),
        "filename": filename,
        "local_file_path": str(local_path),
        "local_file_exists": local_path.exists(),
        "shared_file_path": str(shared_path),
        "shared_file_exists": shared_path.exists(),
    }


def _find_issue_header_row(ws) -> int:
    for row_index in range(1, min(int(ws.max_row or 0), 20) + 1):
        labels = {
            str(ws.cell(row=row_index, column=column_index).value or "").strip().casefold()
            for column_index in range(1, min(int(ws.max_column or 0), 80) + 1)
        }
        if "go#" in labels and "job order no" in labels and "ppo" in labels:
            return row_index
    return 0


def _column_index_from_reference(reference: object) -> int:
    match = re.match(r"([A-Z]+)", str(reference or "").upper())
    if not match:
        return 0
    result = 0
    for character in match.group(1):
        result = result * 26 + (ord(character) - ord("A") + 1)
    return result


def _xml_cell_value(cell, shared_strings: list[str]) -> object:
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    cell_type = str(cell.attrib.get("t") or "").strip()
    if cell_type == "inlineStr":
        return "".join(str(item.text or "") for item in cell.findall(f".//{namespace}t"))
    raw_value = cell.findtext(f"{namespace}v")
    if raw_value is None:
        return ""
    if cell_type == "s":
        try:
            return shared_strings[int(raw_value)]
        except (IndexError, TypeError, ValueError):
            return ""
    if cell_type in {"str", "e"}:
        return raw_value
    try:
        numeric = float(raw_value)
    except (TypeError, ValueError):
        return raw_value
    return int(numeric) if numeric.is_integer() else numeric


def _issue_sheet_cells_from_xlsx(path: Path) -> dict[tuple[int, int], object]:
    main_ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    rel_ns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    package_rel_ns = "{http://schemas.openxmlformats.org/package/2006/relationships}"
    with ZipFile(path) as archive:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
            for entry in shared_root.findall(f"{main_ns}si"):
                shared_strings.append("".join(str(item.text or "") for item in entry.findall(f".//{main_ns}t")))

        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        sheet_rel_id = ""
        for sheet in workbook_root.findall(f".//{main_ns}sheet"):
            if str(sheet.attrib.get("name") or "").strip() == "FORMAT COI REQUEST":
                sheet_rel_id = str(sheet.attrib.get(f"{rel_ns}id") or "")
                break
        if not sheet_rel_id:
            first_sheet = workbook_root.find(f".//{main_ns}sheet")
            sheet_rel_id = str(first_sheet.attrib.get(f"{rel_ns}id") or "") if first_sheet is not None else ""
        if not sheet_rel_id:
            raise ValueError("Workbook has no worksheet")

        relationships_root = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        target = ""
        for relationship in relationships_root.findall(f"{package_rel_ns}Relationship"):
            if str(relationship.attrib.get("Id") or "") == sheet_rel_id:
                target = str(relationship.attrib.get("Target") or "")
                break
        if not target:
            raise ValueError("Cannot resolve worksheet relationship")
        normalized_target = target.lstrip("/")
        sheet_path = normalized_target if normalized_target.startswith("xl/") else "xl/" + normalized_target
        sheet_root = ElementTree.fromstring(archive.read(sheet_path))

    cells: dict[tuple[int, int], object] = {}
    for cell in sheet_root.findall(f".//{main_ns}sheetData/{main_ns}row/{main_ns}c"):
        reference = str(cell.attrib.get("r") or "")
        match = re.match(r"([A-Z]+)(\d+)$", reference.upper())
        if not match:
            continue
        row_index = int(match.group(2))
        column_index = _column_index_from_reference(match.group(1))
        if column_index > 0:
            cells[(row_index, column_index)] = _xml_cell_value(cell, shared_strings)
    return cells


def _parse_issue_workbook_fast(path: Path) -> dict[str, Any]:
    cells = _issue_sheet_cells_from_xlsx(path)
    if not cells:
        return _error("ISSUE COI worksheet is empty", file_path=str(path))
    max_row = max(row for row, _column in cells)
    max_column = max(column for _row, column in cells)
    header_row = 0
    for row_index in range(1, min(max_row, 20) + 1):
        labels = {
            str(cells.get((row_index, column_index), "") or "").strip().casefold()
            for column_index in range(1, min(max_column, 80) + 1)
        }
        if "go#" in labels and "job order no" in labels and "ppo" in labels:
            header_row = row_index
            break
    if not header_row:
        return _error("ISSUE COI header row was not found", file_path=str(path))

    source_columns = []
    for column_index in range(1, max_column + 1):
        label = str(cells.get((header_row, column_index), "") or "").strip()
        if not label:
            continue
        key = _canonical_key(label, label)
        source_columns.append(
            {
                "key": key,
                "label": _KEY_TO_LABEL.get(key, label),
                "source_key": label,
                "column_index": column_index,
            }
        )
    title_go = str(cells.get((1, 1), "") or "").replace("GO:", "").strip().upper()
    issued_at_raw = str(cells.get((2, 1), "") or "").replace("Issue at:", "").strip()
    try:
        issued_at = datetime.fromisoformat(issued_at_raw).isoformat(sep=" ", timespec="seconds")
    except ValueError:
        issued_at = datetime.fromtimestamp(path.stat().st_mtime).isoformat(sep=" ", timespec="seconds")
    rows = []
    for row_index in range(header_row + 1, max_row + 1):
        row = {column["key"]: _json_value(cells.get((row_index, column["column_index"]))) for column in source_columns}
        if not any(value not in (None, "") for value in row.values()):
            continue
        if not str(row.get("GO#") or "").strip() and title_go:
            row["GO#"] = title_go
        rows.append(row)
    go_no = str(title_go or (rows[0].get("GO#") if rows else "") or "").strip().upper()
    if not go_no or not rows:
        return _error("ISSUE COI contains no GO rows", file_path=str(path))
    return {
        "ok": True,
        "go": go_no,
        "issued_at": issued_at,
        "filename": path.name,
        "columns": [{"key": item["key"], "label": item["label"]} for item in source_columns],
        "rows": rows,
    }


def _parse_issue_workbook_openpyxl(path: Path) -> dict[str, Any]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except Exception as exc:
        return _error("Cannot open ISSUE COI workbook", file_path=str(path), detail=str(exc))
    try:
        worksheet = workbook["FORMAT COI REQUEST"] if "FORMAT COI REQUEST" in workbook.sheetnames else workbook.active
        header_row = _find_issue_header_row(worksheet)
        if not header_row:
            return _error("ISSUE COI header row was not found", file_path=str(path))
        source_columns = []
        for column_index in range(1, int(worksheet.max_column or 0) + 1):
            label = str(worksheet.cell(row=header_row, column=column_index).value or "").strip()
            if not label:
                continue
            source_columns.append({"key": _canonical_key(label, label), "label": _KEY_TO_LABEL.get(_canonical_key(label, label), label), "source_key": label, "column_index": column_index})
        if not source_columns:
            return _error("ISSUE COI has no columns", file_path=str(path))

        title_go = str(worksheet.cell(row=1, column=1).value or "").replace("GO:", "").strip().upper()
        issued_at_raw = str(worksheet.cell(row=2, column=1).value or "").replace("Issue at:", "").strip()
        try:
            issued_at = datetime.fromisoformat(issued_at_raw).isoformat(sep=" ", timespec="seconds")
        except ValueError:
            issued_at = datetime.fromtimestamp(path.stat().st_mtime).isoformat(sep=" ", timespec="seconds")
        rows = []
        for row_index in range(header_row + 1, int(worksheet.max_row or 0) + 1):
            row: dict[str, Any] = {}
            for column in source_columns:
                row[column["key"]] = _json_value(worksheet.cell(row=row_index, column=column["column_index"]).value)
            if not any(value not in (None, "") for value in row.values()):
                continue
            if not str(row.get("GO#") or "").strip() and title_go:
                row["GO#"] = title_go
            rows.append(row)
        go_no = str(title_go or (rows[0].get("GO#") if rows else "") or "").strip().upper()
        if not go_no or not rows:
            return _error("ISSUE COI contains no GO rows", file_path=str(path))
        return {
            "ok": True,
            "go": go_no,
            "issued_at": issued_at,
            "filename": path.name,
            "columns": [{"key": item["key"], "label": item["label"]} for item in source_columns],
            "rows": rows,
        }
    finally:
        workbook.close()


def _parse_issue_workbook(path: Path) -> dict[str, Any]:
    try:
        parsed = _parse_issue_workbook_fast(path)
        if parsed.get("ok"):
            return parsed
    except Exception:
        pass
    return _parse_issue_workbook_openpyxl(path)


def backfill_local_issued_coi() -> dict[str, Any]:
    with _ARCHIVE_LOCK:
        ensure_issued_coi_archive()
        files = sorted(
            [
                path
                for path in _ISSUE_DIR.glob("*-COI-*.xlsx")
                if path.is_file() and path.name.casefold() != ISSUED_COI_COMBINED_FILENAME.casefold()
            ],
            key=lambda path: path.stat().st_mtime,
        )
        imported = 0
        skipped = 0
        failed: list[dict[str, str]] = []
        for path in files:
            try:
                source_file_hash = _file_hash(path)
            except Exception as exc:
                failed.append({"filename": path.name, "error": str(exc)})
                continue
            with _connect() as conn:
                existing = conn.execute(
                    "SELECT batch_id FROM issued_coi_batches WHERE source_file_hash = ? LIMIT 1",
                    (source_file_hash,),
                ).fetchone()
            if existing:
                skipped += 1
                continue
            parsed = _parse_issue_workbook(path)
            if not parsed.get("ok"):
                failed.append({"filename": path.name, "error": str(parsed.get("error") or "Unknown error")})
                continue
            archive = archive_issue_snapshot(
                parsed,
                issued_at=str(parsed.get("issued_at") or _now()),
                filename=str(parsed.get("filename") or path.name),
                local_file_path=str(path.resolve()),
                source_file_hash=source_file_hash,
                storage_state="BACKFILLED",
            )
            if not archive.get("ok"):
                failed.append({"filename": path.name, "error": str(archive.get("error") or "Archive error")})
                continue
            published = mark_issue_batch_published(archive.get("batch_id"), {"file_path": str(path.resolve())})
            if not published.get("ok"):
                failed.append({"filename": path.name, "error": str(published.get("error") or "Publish error")})
                continue
            imported += 0 if archive.get("deduplicated") else 1
        return {
            "ok": True,
            "scanned_count": len(files),
            "imported_count": imported,
            "skipped_count": skipped,
            "failed_count": len(failed),
            "failed": failed,
        }
