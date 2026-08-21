from __future__ import annotations

from dataclasses import replace
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .exceptions import WorkbookFormatError
from .models import ExportRecord
from .parsers import classify_go_flow, normalize_lookup_text, normalize_number
from .runtime import recalculate_workbook_with_excel

VISIBLE_HEADERS = [
    "GO",
    "YY Req No",
    "Marker YY",
    "PPO YY",
    "PPO YY Wastage",
    "Gmt Color",
    "Fabric Part",
    "COLOR_CODE",
    "COLOR_DESC",
    "JOB ORDER NO",
    "Qty",
    "PPO",
    "Require YY (for each JO#)",
    "Require YY (all JO#)",
    "PPO Q'ty",
    "Short/Over",
]
CM_VISIBLE_HEADERS = [
    "GO",
    "Gmt Color",
    "Fabric Part",
    "COLOR_CODE",
    "COLOR_DESC",
    "JOB ORDER NO",
    "Qty",
    "PPO",
    "PPO Q'ty",
    "Short/Over",
]
COLLAR_VISIBLE_HEADERS = [
    "GO",
    "Gmt Color",
    "Fabric Part",
    "COLOR_CODE",
    "COLOR_DESC",
    "Size",
    "Qty",
    "PPO",
    "PPO Q'ty",
    "Short/Over",
]
LEGACY_VISIBLE_HEADERS = [
    "GO",
    "YY Req No",
    "Marker YY",
    "PPO YY",
    "PPO YY Wastage",
    "Gmt Color",
    "Fabric Part",
    "Type",
    "COLOR_CODE",
    "COLOR_DESC",
    "JOB ORDER NO",
    "Qty",
    "PPO",
    "Require YY (for each JO#)",
    "Require YY (all JO#)",
    "PPO Q'ty",
    "Short/Over",
]
METADATA_HEADERS = [
    "Flow",
    "Combo Name",
    "Block Index",
    "Section Order",
    "Part Order",
    "Aggregate Key",
    "Is Separator",
    "Sheet Kind",
    "Size",
    "Hidden YY Req No",
    "Hidden Marker YY",
    "Hidden PPO YY",
]
ALL_HEADERS = VISIBLE_HEADERS + METADATA_HEADERS
CM_ALL_HEADERS = CM_VISIBLE_HEADERS + METADATA_HEADERS
COLLAR_ALL_HEADERS = COLLAR_VISIBLE_HEADERS + METADATA_HEADERS

COI_SHEET_NAME = "COI"
COLLAR_SHEET_NAME = "COI Collar／Cuff"
VISIBLE_SHEET_NAME = COI_SHEET_NAME
MASTER_FILE_NAME = "COI Master.xlsx"
MULTI_PPO_CHECK_MESSAGE = "vui long check lai PPO matching"

AGGREGATE_KEY_COLUMN = len(VISIBLE_HEADERS) + METADATA_HEADERS.index("Aggregate Key") + 1
SEPARATOR_COLUMN = len(VISIBLE_HEADERS) + METADATA_HEADERS.index("Is Separator") + 1


def save_master_workbook(records: list[ExportRecord], output_dir: str | Path) -> Path:
    output_path = Path(output_dir).expanduser().resolve()
    output_path.mkdir(parents=True, exist_ok=True)
    file_path = output_path / MASTER_FILE_NAME
    write_workbook(file_path, records)
    return file_path


def write_workbook(file_path: str | Path, records: list[ExportRecord]) -> Path:
    compacted_records = _compact_records(records)
    cm_mode = bool(compacted_records) and all(record.flow == "CM" for record in compacted_records)
    main_records = [record for record in compacted_records if record.sheet_kind != COLLAR_SHEET_NAME]
    collar_records = [record for record in compacted_records if record.sheet_kind == COLLAR_SHEET_NAME]

    workbook = Workbook()
    main_sheet = workbook.active
    main_sheet.title = COI_SHEET_NAME

    _write_main_sheet(main_sheet, main_records, cm_mode=cm_mode)
    if not cm_mode:
        collar_sheet = workbook.create_sheet(COLLAR_SHEET_NAME)
        _write_collar_sheet(collar_sheet, collar_records)

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    path = Path(file_path).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    recalculate_workbook_with_excel(path)
    return path


def read_records_from_workbook(file_path: str | Path) -> list[ExportRecord]:
    workbook = load_workbook(file_path)
    records: list[ExportRecord] = []

    if COI_SHEET_NAME in workbook.sheetnames:
        records.extend(_read_main_sheet(workbook[COI_SHEET_NAME]))
    if COLLAR_SHEET_NAME in workbook.sheetnames:
        records.extend(_read_collar_sheet(workbook[COLLAR_SHEET_NAME]))

    if not records:
        raise WorkbookFormatError("Workbook does not contain any GetYY metadata rows.")

    return [replace(record, row_index=index) for index, record in enumerate(records)]


def _write_main_sheet(worksheet, records: list[ExportRecord], *, cm_mode: bool = False) -> None:
    worksheet.append(CM_ALL_HEADERS if cm_mode else ALL_HEADERS)
    previous_record: ExportRecord | None = None
    for record in records:
        if _should_insert_separator(previous_record, record):
            _append_main_separator_row(worksheet)
        _append_main_record_row(worksheet, record, cm_mode=cm_mode)
        previous_record = record
    if cm_mode:
        _style_cm_main_sheet(worksheet)
    else:
        _style_main_sheet(worksheet)


def _write_collar_sheet(worksheet, records: list[ExportRecord]) -> None:
    worksheet.append(COLLAR_ALL_HEADERS)
    for record in records:
        _append_collar_record_row(worksheet, record)
    _style_collar_sheet(worksheet)


def _read_main_sheet(worksheet) -> list[ExportRecord]:
    headers = [_cell_text(worksheet.cell(row=1, column=index).value) for index in range(1, worksheet.max_column + 1)]
    detected_headers = _detect_main_visible_headers(headers)
    cm_mode = detected_headers == CM_VISIBLE_HEADERS
    visible_count = len(detected_headers)
    visible_index = {header: index for index, header in enumerate(detected_headers)}
    metadata_index = _build_metadata_index(headers[visible_count:], visible_count)

    records: list[ExportRecord] = []
    for worksheet_row in range(2, worksheet.max_row + 1):
        row_values = [
            _cell_text(worksheet.cell(row=worksheet_row, column=column_index).value)
            for column_index in range(1, worksheet.max_column + 1)
        ]
        visible_values = row_values[:visible_count]
        is_separator = _metadata_value(row_values, metadata_index, "Is Separator").upper() == "Y"
        if is_separator or not any(visible_values):
            continue

        go_value = _visible_value(visible_values, visible_index, "GO")
        if not go_value:
            continue

        flow_value = _resolve_flow(go_value, row_values, metadata_index)
        if cm_mode:
            records.append(
                ExportRecord(
                    go=go_value,
                    yy_req_no=_metadata_value(row_values, metadata_index, "Hidden YY Req No"),
                    marker_yy=_metadata_value(row_values, metadata_index, "Hidden Marker YY"),
                    ppo_yy=_metadata_value(row_values, metadata_index, "Hidden PPO YY"),
                    gmt_color=_visible_value(visible_values, visible_index, "Gmt Color"),
                    fabric_part=_visible_value(visible_values, visible_index, "Fabric Part"),
                    color_code=_visible_value(visible_values, visible_index, "COLOR_CODE").upper(),
                    color_desc=_visible_value(visible_values, visible_index, "COLOR_DESC"),
                    fabric_color="",
                    jo=_visible_value(visible_values, visible_index, "JOB ORDER NO").upper(),
                    minus_pct="",
                    plus_pct="",
                    qty=_visible_value(visible_values, visible_index, "Qty"),
                    ppo_no=_normalize_ppo_no(_visible_value(visible_values, visible_index, "PPO")),
                    ppo_qty=_visible_value(visible_values, visible_index, "PPO Q'ty"),
                    go_key=go_value,
                    row_index=len(records),
                    flow=flow_value or "CM",
                    combo_name=_metadata_value(row_values, metadata_index, "Combo Name"),
                    block_index=_safe_int(_metadata_value(row_values, metadata_index, "Block Index")),
                    section_order=_safe_int(_metadata_value(row_values, metadata_index, "Section Order")),
                    part_order=_safe_int(_metadata_value(row_values, metadata_index, "Part Order")),
                    aggregate_key=_metadata_value(row_values, metadata_index, "Aggregate Key"),
                    is_separator=False,
                    sheet_kind=_metadata_value(row_values, metadata_index, "Sheet Kind") or COI_SHEET_NAME,
                    size=_metadata_value(row_values, metadata_index, "Size"),
                )
            )
        else:
            records.append(
                ExportRecord(
                    go=go_value,
                    yy_req_no=_visible_value(visible_values, visible_index, "YY Req No"),
                    marker_yy=_visible_value(visible_values, visible_index, "Marker YY"),
                    ppo_yy=_visible_value(visible_values, visible_index, "PPO YY"),
                    gmt_color=_visible_value(visible_values, visible_index, "Gmt Color"),
                    fabric_part=_visible_value(visible_values, visible_index, "Fabric Part"),
                    color_code=_visible_value(visible_values, visible_index, "COLOR_CODE").upper(),
                    color_desc=_visible_value(visible_values, visible_index, "COLOR_DESC"),
                    fabric_color="",
                    jo=_visible_value(visible_values, visible_index, "JOB ORDER NO").upper(),
                    minus_pct="",
                    plus_pct="",
                    qty=_visible_value(visible_values, visible_index, "Qty"),
                    ppo_no=_normalize_ppo_no(_visible_value(visible_values, visible_index, "PPO")),
                    ppo_qty=_visible_value(visible_values, visible_index, "PPO Q'ty"),
                    go_key=go_value,
                    row_index=len(records),
                    flow=flow_value,
                    combo_name=_metadata_value(row_values, metadata_index, "Combo Name"),
                    block_index=_safe_int(_metadata_value(row_values, metadata_index, "Block Index")),
                    section_order=_safe_int(_metadata_value(row_values, metadata_index, "Section Order")),
                    part_order=_safe_int(_metadata_value(row_values, metadata_index, "Part Order")),
                    aggregate_key=_metadata_value(row_values, metadata_index, "Aggregate Key"),
                    is_separator=False,
                    sheet_kind=_metadata_value(row_values, metadata_index, "Sheet Kind") or COI_SHEET_NAME,
                    size=_metadata_value(row_values, metadata_index, "Size"),
                )
            )
    return records


def _read_collar_sheet(worksheet) -> list[ExportRecord]:
    headers = [_cell_text(worksheet.cell(row=1, column=index).value) for index in range(1, worksheet.max_column + 1)]
    detected_headers = _detect_collar_visible_headers(headers)
    visible_count = len(detected_headers)
    visible_index = {header: index for index, header in enumerate(detected_headers)}
    metadata_index = _build_metadata_index(headers[visible_count:], visible_count)

    records: list[ExportRecord] = []
    for worksheet_row in range(2, worksheet.max_row + 1):
        row_values = [
            _cell_text(worksheet.cell(row=worksheet_row, column=column_index).value)
            for column_index in range(1, worksheet.max_column + 1)
        ]
        visible_values = row_values[:visible_count]
        if not any(visible_values):
            continue

        go_value = _visible_value(visible_values, visible_index, "GO")
        if not go_value:
            continue

        flow_value = _resolve_flow(go_value, row_values, metadata_index)
        records.append(
            ExportRecord(
                go=go_value,
                yy_req_no=_metadata_value(row_values, metadata_index, "Hidden YY Req No"),
                marker_yy=_metadata_value(row_values, metadata_index, "Hidden Marker YY"),
                ppo_yy=_metadata_value(row_values, metadata_index, "Hidden PPO YY"),
                gmt_color=_visible_value(visible_values, visible_index, "Gmt Color"),
                fabric_part=_visible_value(visible_values, visible_index, "Fabric Part"),
                color_code=_visible_value(visible_values, visible_index, "COLOR_CODE").upper(),
                color_desc=_visible_value(visible_values, visible_index, "COLOR_DESC"),
                fabric_color="",
                jo="",
                minus_pct="",
                plus_pct="",
                qty=_visible_value(visible_values, visible_index, "Qty"),
                ppo_no=_normalize_ppo_no(_visible_value(visible_values, visible_index, "PPO")),
                ppo_qty=_visible_value(visible_values, visible_index, "PPO Q'ty"),
                go_key=go_value,
                row_index=len(records),
                flow=flow_value,
                combo_name=_metadata_value(row_values, metadata_index, "Combo Name"),
                block_index=_safe_int(_metadata_value(row_values, metadata_index, "Block Index")),
                section_order=_safe_int(_metadata_value(row_values, metadata_index, "Section Order")),
                part_order=_safe_int(_metadata_value(row_values, metadata_index, "Part Order")),
                aggregate_key=_metadata_value(row_values, metadata_index, "Aggregate Key"),
                is_separator=False,
                sheet_kind=_metadata_value(row_values, metadata_index, "Sheet Kind") or COLLAR_SHEET_NAME,
                size=_visible_value(visible_values, visible_index, "Size")
                or _metadata_value(row_values, metadata_index, "Size"),
            )
        )
    return records


def _append_main_record_row(worksheet, record: ExportRecord, *, cm_mode: bool = False) -> None:
    row_index = worksheet.max_row + 1
    if cm_mode:
        worksheet.append(
            [
                record.go,
                record.gmt_color,
                record.fabric_part,
                record.color_code,
                record.color_desc,
                record.jo,
                _excel_number(record.qty),
                record.ppo_no,
                _excel_number(record.ppo_qty),
                _cm_short_over_formula(row_index),
                record.flow,
                record.combo_name,
                _excel_number(str(record.block_index)),
                _excel_number(str(record.section_order)),
                _excel_number(str(record.part_order)),
                record.aggregate_key or _build_aggregate_key(record),
                "N",
                record.sheet_kind or COI_SHEET_NAME,
                record.size,
                record.yy_req_no,
                _excel_number(record.marker_yy),
                _excel_number(record.ppo_yy),
            ]
        )
        return

    worksheet.append(
        [
            record.go,
            record.yy_req_no,
            _excel_number(record.marker_yy),
            _excel_number(record.ppo_yy),
            _ppo_yy_wastage_formula(row_index),
            record.gmt_color,
            record.fabric_part,
            record.color_code,
            record.color_desc,
            record.jo,
            _excel_number(record.qty),
            record.ppo_no,
            _require_yy_each_formula(row_index),
            _require_yy_all_formula(row_index),
            _excel_number(record.ppo_qty),
            _main_short_over_formula(row_index),
            record.flow,
            record.combo_name,
            _excel_number(str(record.block_index)),
            _excel_number(str(record.section_order)),
            _excel_number(str(record.part_order)),
            record.aggregate_key or _build_aggregate_key(record),
            "N",
            record.sheet_kind or COI_SHEET_NAME,
            record.size,
            record.yy_req_no,
            _excel_number(record.marker_yy),
            _excel_number(record.ppo_yy),
        ]
    )


def _append_collar_record_row(worksheet, record: ExportRecord) -> None:
    row_index = worksheet.max_row + 1
    worksheet.append(
        [
            record.go,
            record.gmt_color,
            record.fabric_part,
            record.color_code,
            record.color_desc,
            record.size,
            _excel_number(record.qty),
            record.ppo_no,
            _excel_number(record.ppo_qty),
            _collar_short_over_formula(row_index),
            record.flow,
            record.combo_name,
            _excel_number(str(record.block_index)),
            _excel_number(str(record.section_order)),
            _excel_number(str(record.part_order)),
            record.aggregate_key or _build_aggregate_key(record),
            "N",
            record.sheet_kind or COLLAR_SHEET_NAME,
            record.size,
            record.yy_req_no,
            _excel_number(record.marker_yy),
            _excel_number(record.ppo_yy),
        ]
    )


def _append_main_separator_row(worksheet) -> None:
    worksheet.append([None] * (len(ALL_HEADERS) - 6) + ["Y", COI_SHEET_NAME, "", "", "", ""])


def _detect_main_visible_headers(headers: list[str]) -> list[str]:
    if headers[: len(VISIBLE_HEADERS)] == VISIBLE_HEADERS:
        return VISIBLE_HEADERS
    if headers[: len(CM_VISIBLE_HEADERS)] == CM_VISIBLE_HEADERS:
        return CM_VISIBLE_HEADERS
    if headers[: len(LEGACY_VISIBLE_HEADERS)] == LEGACY_VISIBLE_HEADERS:
        return LEGACY_VISIBLE_HEADERS
    raise WorkbookFormatError("Workbook format is not compatible with GetYY update flow.")


def _detect_collar_visible_headers(headers: list[str]) -> list[str]:
    if headers[: len(COLLAR_VISIBLE_HEADERS)] == COLLAR_VISIBLE_HEADERS:
        return COLLAR_VISIBLE_HEADERS
    raise WorkbookFormatError("Workbook format is not compatible with GetYY update flow.")


def _build_metadata_index(raw_metadata_headers: list[str], visible_count: int) -> dict[str, int]:
    metadata_index: dict[str, int] = {}
    for offset, header in enumerate(raw_metadata_headers, start=visible_count):
        normalized = header.strip()
        if not normalized:
            continue
        if normalized not in METADATA_HEADERS:
            raise WorkbookFormatError("Workbook format is not compatible with GetYY update flow.")
        metadata_index[normalized] = offset
    return metadata_index


def _metadata_value(row_values: list[str], metadata_index: dict[str, int], header: str) -> str:
    index = metadata_index.get(header)
    if index is None or index >= len(row_values):
        return ""
    return row_values[index]


def _visible_value(row_values: list[str], visible_index: dict[str, int], header: str) -> str:
    index = visible_index.get(header)
    if index is None or index >= len(row_values):
        return ""
    return row_values[index]


def _resolve_flow(go_key: str, row_values: list[str], metadata_index: dict[str, int]) -> str:
    flow_value = _metadata_value(row_values, metadata_index, "Flow").upper()
    if flow_value:
        return flow_value
    try:
        return classify_go_flow(go_key)
    except Exception as exc:
        raise WorkbookFormatError(f"Cannot infer flow for GO '{go_key}': {exc}") from exc


def _style_main_sheet(worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    yy_fill = PatternFill(fill_type="solid", fgColor="F8FBFE")
    coi_fill = PatternFill(fill_type="solid", fgColor="FDFBF4")
    ppo_fill = PatternFill(fill_type="solid", fgColor="F6FAF8")
    blank_fill = PatternFill(fill_type=None)
    header_font = Font(bold=True)
    top_alignment = Alignment(vertical="top", wrap_text=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    table_border = _table_border()

    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        if cell.column <= len(VISIBLE_HEADERS):
            cell.border = table_border

    for row in worksheet.iter_rows(min_row=2):
        is_separator = _cell_text(row[SEPARATOR_COLUMN - 1].value).upper() == "Y"
        for cell in row:
            cell.alignment = top_alignment
            if is_separator and cell.column <= len(VISIBLE_HEADERS):
                cell.fill = blank_fill
                cell.border = Border()
                continue
            if 1 <= cell.column <= 7:
                cell.fill = yy_fill
            elif 8 <= cell.column <= 11:
                cell.fill = coi_fill
            elif 12 <= cell.column <= 16:
                cell.fill = ppo_fill
            if cell.column <= len(VISIBLE_HEADERS):
                cell.border = table_border
            if cell.column in {5, 16}:
                cell.number_format = "0.0%"
            elif cell.column in {3, 4, 11, 13, 14, 15}:
                cell.number_format = "0.####"
        if is_separator:
            worksheet.row_dimensions[row[0].row].height = 8
        else:
            worksheet.row_dimensions[row[0].row].height = 21

    worksheet.freeze_panes = "A2"
    _autosize_columns(worksheet, len(VISIBLE_HEADERS))
    _apply_alert_rules(worksheet, ["E", "P"])
    _hide_metadata_columns(worksheet, len(VISIBLE_HEADERS), len(ALL_HEADERS))


def _style_cm_main_sheet(worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    info_fill = PatternFill(fill_type="solid", fgColor="F8FBFE")
    qty_fill = PatternFill(fill_type="solid", fgColor="FDFBF4")
    ppo_fill = PatternFill(fill_type="solid", fgColor="F6FAF8")
    header_font = Font(bold=True)
    top_alignment = Alignment(vertical="top", wrap_text=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    table_border = _table_border()

    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        if cell.column <= len(CM_VISIBLE_HEADERS):
            cell.border = table_border

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = top_alignment
            if 1 <= cell.column <= 6:
                cell.fill = info_fill
            elif cell.column == 7:
                cell.fill = qty_fill
            elif 8 <= cell.column <= 10:
                cell.fill = ppo_fill
            if cell.column <= len(CM_VISIBLE_HEADERS):
                cell.border = table_border
            if cell.column == 10:
                cell.number_format = "0.0%"
            elif cell.column in {7, 9}:
                cell.number_format = "0.####"
        worksheet.row_dimensions[row[0].row].height = 21

    worksheet.freeze_panes = "A2"
    _autosize_columns(worksheet, len(CM_VISIBLE_HEADERS))
    _apply_alert_rules(worksheet, ["J"])
    _hide_metadata_columns(worksheet, len(CM_VISIBLE_HEADERS), len(CM_ALL_HEADERS))


def _style_collar_sheet(worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    info_fill = PatternFill(fill_type="solid", fgColor="F8FBFE")
    qty_fill = PatternFill(fill_type="solid", fgColor="FDFBF4")
    ppo_fill = PatternFill(fill_type="solid", fgColor="F6FAF8")
    header_font = Font(bold=True)
    top_alignment = Alignment(vertical="top", wrap_text=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    table_border = _table_border()

    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        if cell.column <= len(COLLAR_VISIBLE_HEADERS):
            cell.border = table_border

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = top_alignment
            if 1 <= cell.column <= 6:
                cell.fill = info_fill
            elif 7 <= cell.column <= 10:
                cell.fill = qty_fill if cell.column == 7 else ppo_fill
            if cell.column <= len(COLLAR_VISIBLE_HEADERS):
                cell.border = table_border
            if cell.column == 10:
                cell.number_format = "0.0%"
            elif cell.column in {7, 9}:
                cell.number_format = "0.####"
        worksheet.row_dimensions[row[0].row].height = 21

    worksheet.freeze_panes = "A2"
    _autosize_columns(worksheet, len(COLLAR_VISIBLE_HEADERS))
    _apply_alert_rules(worksheet, ["J"])
    _hide_metadata_columns(worksheet, len(COLLAR_VISIBLE_HEADERS), len(COLLAR_ALL_HEADERS))


def _hide_metadata_columns(worksheet, visible_count: int, total_count: int) -> None:
    for column_index in range(visible_count + 1, total_count + 1):
        column_dimension = worksheet.column_dimensions[get_column_letter(column_index)]
        column_dimension.hidden = True
        column_dimension.width = 0


def _autosize_columns(worksheet, visible_count: int) -> None:
    for column_index in range(1, visible_count + 1):
        max_length = 0
        for cell in worksheet[get_column_letter(column_index)]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        if visible_count == len(VISIBLE_HEADERS):
            if column_index in {9, 10}:
                adjusted_width = 28
            elif column_index in {6, 7, 12}:
                adjusted_width = min(max(max_length + 2, 18), 28)
            elif column_index in {1, 2, 3, 4, 5, 11, 13, 14, 15, 16}:
                adjusted_width = min(max(max_length + 2, 12), 20)
            else:
                adjusted_width = min(max(max_length + 2, 10), 18)
        elif visible_count == len(CM_VISIBLE_HEADERS):
            if column_index in {2, 5, 6, 8}:
                adjusted_width = min(max(max_length + 2, 18), 30)
            elif column_index == 3:
                adjusted_width = min(max(max_length + 2, 14), 22)
            else:
                adjusted_width = min(max(max_length + 2, 10), 18)
        else:
            if column_index in {2, 5, 8}:
                adjusted_width = min(max(max_length + 2, 18), 32)
            elif column_index == 3:
                adjusted_width = min(max(max_length + 2, 14), 22)
            else:
                adjusted_width = min(max(max_length + 2, 10), 18)
        worksheet.column_dimensions[get_column_letter(column_index)].width = adjusted_width


def _table_border() -> Border:
    side = Side(style="thin", color="B7C3D0")
    return Border(left=side, right=side, top=side, bottom=side)


def _should_insert_separator(previous_record: ExportRecord | None, current_record: ExportRecord) -> bool:
    if previous_record is None:
        return False
    if previous_record.sheet_kind != COI_SHEET_NAME or current_record.sheet_kind != COI_SHEET_NAME:
        return False
    return previous_record.go == current_record.go and previous_record.flow == "WOVEN" and current_record.flow == "KNIT"


def _build_aggregate_key(record: ExportRecord) -> str:
    return "|".join(
        [
            record.go.upper(),
            record.sheet_kind.upper(),
            record.flow.upper(),
            normalize_lookup_text(record.fabric_part),
            record.color_code.upper(),
            normalize_lookup_text(record.combo_name),
            normalize_lookup_text(record.size),
        ]
    )


def _cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _excel_number(value: str) -> object:
    text = (value or "").strip()
    if not text:
        return ""

    normalized = normalize_number(text)
    if normalized == "":
        return text

    try:
        decimal_value = Decimal(normalized)
    except (InvalidOperation, ValueError):
        return text

    if decimal_value == decimal_value.to_integral_value():
        return int(decimal_value)
    return float(decimal_value)


def _safe_int(value: str) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _normalize_ppo_no(value: str) -> str:
    trimmed = value.strip()
    if normalize_lookup_text(trimmed) == normalize_lookup_text(MULTI_PPO_CHECK_MESSAGE):
        return MULTI_PPO_CHECK_MESSAGE
    return trimmed.upper()


def _apply_alert_rules(worksheet, columns: list[str]) -> None:
    alert_font = Font(color="000000")
    alert_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    for column in columns:
        cell_range = f"{column}2:{column}1048576"
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="greaterThan", formula=["1.04"], font=alert_font, fill=alert_fill),
        )


def _ppo_yy_wastage_formula(row_index: int) -> str:
    return f'=IF(OR(C{row_index}=0,C{row_index}="",D{row_index}=0,D{row_index}=""),"",D{row_index}/C{row_index})'


def _require_yy_each_formula(row_index: int) -> str:
    return f'=IF(OR(C{row_index}=0,C{row_index}="",K{row_index}=0,K{row_index}=""),"-",IFERROR(K{row_index}*C{row_index},0))'


def _require_yy_all_formula(row_index: int) -> str:
    aggregate_column = get_column_letter(AGGREGATE_KEY_COLUMN)
    return (
        f'=IF(OR(L{row_index}="",M{row_index}="-",M{row_index}="",{aggregate_column}{row_index}=""),"",'
        f'SUMIFS(M:M,{aggregate_column}:{aggregate_column},{aggregate_column}{row_index},L:L,L{row_index}))'
    )


def _main_short_over_formula(row_index: int) -> str:
    return f'=IF(OR(L{row_index}="",N{row_index}=0,N{row_index}="",O{row_index}=""),"",O{row_index}/N{row_index})'


def _cm_short_over_formula(row_index: int) -> str:
    return f'=IF(OR(G{row_index}=0,G{row_index}="",I{row_index}=""),"",I{row_index}/G{row_index})'


def _collar_short_over_formula(row_index: int) -> str:
    return f'=IF(OR(H{row_index}="",G{row_index}=0,G{row_index}="",I{row_index}=""),"",I{row_index}/G{row_index})'


def _compact_records(records: list[ExportRecord]) -> list[ExportRecord]:
    compacted: list[ExportRecord] = []
    index_by_key: dict[tuple[object, ...], int] = {}

    for record in records:
        key = _compaction_key(record)
        existing_index = index_by_key.get(key)
        if existing_index is None:
            index_by_key[key] = len(compacted)
            compacted.append(replace(record, row_index=len(compacted)))
            continue

        existing = compacted[existing_index]
        compacted[existing_index] = replace(
            existing,
            qty=_sum_numeric_text(existing.qty, record.qty),
            ppo_qty=_sum_numeric_text(existing.ppo_qty, record.ppo_qty),
            row_index=existing_index,
        )

    return compacted


def _compaction_key(record: ExportRecord) -> tuple[object, ...]:
    return (
        record.sheet_kind,
        record.go,
        record.yy_req_no,
        record.marker_yy,
        record.ppo_yy,
        record.gmt_color,
        record.fabric_part,
        record.color_code,
        record.color_desc,
        record.fabric_color,
        record.jo,
        record.minus_pct,
        record.plus_pct,
        record.ppo_no,
        record.go_key,
        record.flow,
        record.combo_name,
        record.block_index,
        record.section_order,
        record.part_order,
        record.aggregate_key,
        record.is_separator,
        record.size,
    )


def _sum_numeric_text(left: str, right: str) -> str:
    left_text = (left or "").strip()
    right_text = (right or "").strip()
    if not left_text and not right_text:
        return ""
    try:
        total = Decimal(normalize_number(left_text) or "0") + Decimal(normalize_number(right_text) or "0")
        return normalize_number(str(total))
    except (InvalidOperation, ValueError):
        return left_text or right_text
